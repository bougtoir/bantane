import logging

from pathlib import Path
import sys
import platform
import traceback
from typing import Optional, Tuple, Dict, List, Set

import pandas as pd
import numpy as np
import re
import calendar
import datetime as dt

from PySide6.QtWidgets import (
    QApplication, QWidget, QFileDialog, QPushButton, QLabel, QLineEdit,
    QHBoxLayout, QVBoxLayout, QMessageBox, QTextEdit, QComboBox, QGroupBox,
    QTabWidget, QSpinBox, QProgressBar, QDialog
)
from PySide6.QtCore import Qt, QMimeData, QTimer
from PySide6.QtGui import QFont, QFontMetrics, QResizeEvent, QDragEnterEvent, QDropEvent

import pulp
import matplotlib.pyplot as plt
import jpholiday

logging.basicConfig(filename="shift_app.log", level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")


def _is_bundled_exe() -> bool:
    """Return True if running as a bundled executable (PyInstaller/Nuitka)."""
    if getattr(sys, "frozen", False):
        return True
    try:
        exe_stem = Path(sys.executable).stem.lower()
        if exe_stem not in ("python", "python3", "pythonw", "python3w") \
                and not exe_stem.startswith("python3."):
            return True
    except Exception:
        pass
    return False


def get_app_dir() -> Path:
    """Get the application directory.

    PyInstaller sets sys.frozen; Nuitka does not, but sys.executable
    points to the compiled .exe rather than a Python interpreter.
    In Nuitka --onefile mode sys.executable may point to a temp
    extraction directory; sys.argv[0] preserves the original path.
    """
    if _is_bundled_exe():
        argv0_dir = Path(sys.argv[0]).resolve().parent
        exe_dir = Path(sys.executable).resolve().parent
        logging.info(
            'get_app_dir: argv0=%s, executable=%s', argv0_dir, exe_dir,
        )
        return argv0_dir
    # Normal script execution
    return Path(__file__).resolve().parent


def get_system_japanese_font():
    """Get appropriate Japanese font name for the current OS"""
    if platform.system() == "Windows":
        return "Yu Gothic UI"
    elif platform.system() == "Darwin":
        return "Hiragino Sans"
    else:
        return ""


def get_excel_japanese_font():
    """Get appropriate Japanese font name for Excel output"""
    if platform.system() == "Windows":
        return "Yu Gothic"
    else:
        return "IPAGothic"


def configure_matplotlib_fonts():
    """Configure matplotlib to use appropriate Japanese fonts for the current OS"""
    import matplotlib.pyplot as plt
    
    if platform.system() == "Windows":
        plt.rcParams['font.family'] = ['Yu Gothic', 'Yu Gothic UI', 'Meiryo', 'Meiryo UI', 'MS Gothic', 'MS PGothic', 'sans-serif']
        plt.rcParams['font.size'] = 10
        plt.rcParams['axes.unicode_minus'] = False
        print("Using Windows Japanese fonts: Yu Gothic, Meiryo, MS Gothic")
        return True
    elif platform.system() == "Darwin":
        plt.rcParams['font.family'] = ['Hiragino Sans', 'Hiragino Kaku Gothic ProN', 'sans-serif']
        plt.rcParams['font.size'] = 10
        plt.rcParams['axes.unicode_minus'] = False
        print("Using macOS Japanese fonts: Hiragino Sans")
        return True
    else:
        try:
            import japanize_matplotlib
            print("Using japanize-matplotlib for Japanese fonts")
            return True
        except ImportError:
            import os
            import matplotlib.font_manager as fm
            
            linux_font_paths = [
                '/usr/share/fonts/opentype/ipafont-gothic/ipag.ttf',
                '/usr/share/fonts/opentype/ipafont-gothic/ipagp.ttf',
                '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc'
            ]
            
            for font_path in linux_font_paths:
                if os.path.exists(font_path):
                    try:
                        prop = fm.FontProperties(fname=font_path)
                        font_name = prop.get_name()
                        plt.rcParams['font.family'] = font_name
                        plt.rcParams['font.size'] = 10
                        plt.rcParams['axes.unicode_minus'] = False
                        print(f"Using Linux font: {font_name}")
                        return True
                    except Exception as e:
                        print(f"Error loading font {font_path}: {e}")
                        continue
            
            plt.rcParams['font.family'] = ['IPAGothic', 'DejaVu Sans', 'sans-serif']
            plt.rcParams['axes.unicode_minus'] = False
            print("Using fallback font list")
            return False


class ValidationIssue:
    """Represents a validation issue found in input files"""
    def __init__(self, severity: str, code: str, file_name: str, sheet: str, 
                 location: str, field: str, value: str, expected: str, hint: str):
        self.severity = severity  # "ERROR" or "WARNING"
        self.code = code
        self.file_name = file_name
        self.sheet = sheet
        self.location = location  # e.g., "行3" or "セルA5"
        self.field = field
        self.value = value
        self.expected = expected
        self.hint = hint
    
    def to_string(self) -> str:
        """Format issue as a readable string"""
        return (f"[{self.severity}] {self.code}\n"
                f"ファイル: {self.file_name}\n"
                f"シート: {self.sheet}\n"
                f"場所: {self.location}\n"
                f"項目: {self.field}\n"
                f"現在の値: {self.value}\n"
                f"期待される形式: {self.expected}\n"
                f"対処方法: {self.hint}\n")


def hhmm_to_minutes(hhmm: int) -> int:
    h = hhmm // 100
    m = hhmm % 100
    return h * 60 + m


def minutes_to_hhmm(mins: int, round10=True):
    if round10:
        r = ((mins + 9) // 10) * 10
        h = r // 60
        m = r % 60
        return h * 100 + m
    else:
        r = mins
        h = r // 60
        m = r % 60
        return f"{h:02d}:{m:02d}"


def parse_tokens(s: str) -> Set[str]:
    if pd.isna(s) or not str(s).strip():
        return set()
    return set([t.strip() for t in str(s).split(",") if t.strip()])

WEEKDAYS = ["sun", "mon", "tue", "wed", "thu", "fri", "sat"]


def normalize_weekday_token(s: str) -> str:
    if pd.isna(s):
        return ""
    s = str(s).strip().lower()
    alias = {
        "sunday": "sun",
        "monday": "mon",
        "tuesday": "tue",
        "wednesday": "wed",
        "thursday": "thu",
        "friday": "fri",
        "saturday": "sat",
    }
    if s in alias:
        s = alias[s]
    elif len(s) >= 3:
        s = s[:3]
    return s if s in WEEKDAYS else ""


_re_iso = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_re_dd = re.compile(r"^\d{1,2}$")

PERIOD_CUTOFF_DAY = 21


def get_next_period_start(today: dt.date, cutoff: int = PERIOD_CUTOFF_DAY) -> dt.date:
    """Get the start date of the next period (cutoff day of current or next month)"""
    if today.day < cutoff:
        return today.replace(day=cutoff)
    else:
        first_next = (today.replace(day=1) + dt.timedelta(days=32)).replace(day=1)
        return first_next.replace(day=cutoff)


def parse_date_token(tok: str, year: int, month: int, use_period: bool = True, cutoff: int = PERIOD_CUTOFF_DAY) -> str:
    """
    Parse date token to ISO format.
    
    If use_period=True (default):
    - Selected year/month represents the period starting on the cutoff day (default 21st)
    - Days >= cutoff are in the selected month
    - Days < cutoff are in the next month
    - Example: year=2025, month=10 means period 2025-10-21 to 2025-11-20
      - dd=21 -> 2025-10-21
      - dd=31 -> 2025-10-31
      - dd=1 -> 2025-11-01
      - dd=20 -> 2025-11-20
    
    If use_period=False:
    - Traditional behavior: all days are in the selected month
    """
    t = str(tok).strip()
    if _re_iso.match(t):
        return t
    if _re_dd.match(t):
        day = int(t)
        
        if use_period:
            if day >= cutoff:
                target_year = year
                target_month = month
            else:
                if month == 12:
                    target_year = year + 1
                    target_month = 1
                else:
                    target_year = year
                    target_month = month + 1
        else:
            target_year = year
            target_month = month
        
        last = calendar.monthrange(target_year, target_month)[1]
        if not (1 <= day <= last):
            raise ValueError(f"日付(dd)が不正です: {t} (year={target_year}, month={target_month})")
        return f"{target_year:04d}-{target_month:02d}-{day:02d}"
    raise ValueError(f"日付形式は dd もしくは yyyy-mm-dd にしてください: {tok}")


def to_dd(date_str: str) -> str:
    return date_str.split("-")[-1]


def weekday_of(date_str: str) -> str:
    y, m, d = map(int, date_str.split("-"))
    idx = (dt.date(y, m, d).weekday() + 1) % 7
    return WEEKDAYS[idx]


def get_japanese_weekday(weekday_idx) -> str:
    """Get Japanese weekday name from weekday index (0=Monday, 6=Sunday) or date string (YYYY-MM-DD format)"""
    if isinstance(weekday_idx, str):
        y, m, d = map(int, weekday_idx.split("-"))
        weekday_idx = dt.date(y, m, d).weekday()
    japanese_weekdays = ["月", "火", "水", "木", "金", "土", "日"]
    return japanese_weekdays[weekday_idx]

def count_prefer_entries(prefer_tokens) -> int:
    """Count the number of prefer entries for prioritization"""
    return len(prefer_tokens) if prefer_tokens else 0


def validate_excel_file(file_path: Path, file_type: str, target_year: int, target_month: int, kintai_path: Optional[Path] = None) -> List[ValidationIssue]:
    """
    Validate an Excel input file and return a list of issues.
    
    Args:
        file_path: Path to the Excel file
        file_type: Type of file ("work", "jobA", "jobB", "duty", "setting")
        target_year: Target year for date validation
        target_month: Target month for date validation (period start month)
        kintai_path: Optional path to kintai file (if provided, shift sheet is not required)
    
    Returns:
        List of ValidationIssue objects
    """
    issues = []
    file_name = file_path.name
    
    try:
        xl = pd.read_excel(file_path, sheet_name=None)
    except Exception as e:
        issues.append(ValidationIssue(
            severity="ERROR",
            code="E001",
            file_name=file_name,
            sheet="",
            location="ファイル全体",
            field="",
            value="",
            expected="有効なExcelファイル (.xlsx)",
            hint=f"ファイルを開けません。ファイルが破損しているか、他のプログラムで開かれている可能性があります。エラー: {str(e)}"
        ))
        return issues
    
    if file_type == "work":
        if "work" not in xl:
            issues.append(ValidationIssue(
                severity="ERROR",
                code="E101",
                file_name=file_name,
                sheet="",
                location="ファイル全体",
                field="workシート",
                value="見つかりません",
                expected="workシートが必要",
                hint="業務一覧ファイルには「work」という名前のシートが必要です。シート名を確認してください。"
            ))
        else:
            work_df = xl["work"]
            required_cols = ["id", "date", "start", "duration", "need_A", "need_B", "dept"]
            for col in required_cols:
                if col not in work_df.columns:
                    issues.append(ValidationIssue(
                        severity="ERROR",
                        code="E102",
                        file_name=file_name,
                        sheet="work",
                        location="列ヘッダー",
                        field=col,
                        value="見つかりません",
                        expected=f"列「{col}」が必要",
                        hint=f"workシートに「{col}」列を追加してください。必須列: {', '.join(required_cols)}"
                    ))
            
            if "date" in work_df.columns:
                for idx, row in work_df.iterrows():
                    try:
                        date_val = str(row["date"]).strip()
                        if pd.isna(row["date"]) or not date_val:
                            issues.append(ValidationIssue(
                                severity="ERROR",
                                code="E211",
                                file_name=file_name,
                                sheet="work",
                                location=f"行{idx+2}",
                                field="date",
                                value="空欄",
                                expected="dd (例: 1, 15, 21) または yyyy-mm-dd (例: 2025-10-21)",
                                hint=f"日付を入力してください。このアプリでは「{target_year}年{target_month}月」を選択すると、{target_year}年{target_month}月21日〜{target_month+1 if target_month<12 else 1}月20日の期間になります。dd=1〜20は翌月として扱われます。"
                            ))
                        else:
                            parse_date_token(date_val, target_year, target_month)
                    except ValueError as e:
                        issues.append(ValidationIssue(
                            severity="ERROR",
                            code="E211",
                            file_name=file_name,
                            sheet="work",
                            location=f"行{idx+2}",
                            field="date",
                            value=str(row["date"]),
                            expected="dd (例: 1, 15, 21) または yyyy-mm-dd (例: 2025-10-21)",
                            hint=f"日付の形式が正しくありません。このアプリでは「{target_year}年{target_month}月」を選択すると、{target_year}年{target_month}月21日〜{target_month+1 if target_month<12 else 1}月20日の期間になります。dd=1〜20は翌月として扱われます。エラー: {str(e)}"
                        ))
        
        if "dict" not in xl:
            issues.append(ValidationIssue(
                severity="ERROR",
                code="E101",
                file_name=file_name,
                sheet="",
                location="ファイル全体",
                field="dictシート",
                value="見つかりません",
                expected="dictシートが必要",
                hint="業務一覧ファイルには「dict」という名前のシートが必要です。シート名を確認してください。"
            ))
        else:
            dict_df = xl["dict"]
            if "aka" not in dict_df.columns or "name" not in dict_df.columns:
                issues.append(ValidationIssue(
                    severity="ERROR",
                    code="E102",
                    file_name=file_name,
                    sheet="dict",
                    location="列ヘッダー",
                    field="aka, name",
                    value="見つかりません",
                    expected="列「aka」と「name」が必要",
                    hint="dictシートには「aka」列と「name」列が必要です。"
                ))
    
    elif file_type in ["jobA", "jobB"]:
        if "members" not in xl:
            issues.append(ValidationIssue(
                severity="ERROR",
                code="E101",
                file_name=file_name,
                sheet="",
                location="ファイル全体",
                field="membersシート",
                value="見つかりません",
                expected="membersシートが必要",
                hint=f"職種{file_type[-1]}設定ファイルには「members」という名前のシートが必要です。"
            ))
        else:
            members_df = xl["members"]
            required_cols = ["aka"]
            for col in required_cols:
                if col not in members_df.columns:
                    issues.append(ValidationIssue(
                        severity="ERROR",
                        code="E102",
                        file_name=file_name,
                        sheet="members",
                        location="列ヘッダー",
                        field=col,
                        value="見つかりません",
                        expected=f"列「{col}」が必要",
                        hint=f"membersシートに「{col}」列を追加してください。"
                    ))
            
            if "aka" in members_df.columns:
                aka_counts = members_df["aka"].value_counts()
                duplicates = aka_counts[aka_counts > 1]
                if len(duplicates) > 0:
                    for aka, count in duplicates.items():
                        issues.append(ValidationIssue(
                            severity="ERROR",
                            code="E103",
                            file_name=file_name,
                            sheet="members",
                            location="aka列",
                            field="aka",
                            value=str(aka),
                            expected="一意の値",
                            hint=f"aka「{aka}」が{count}回出現しています。各メンバーのakaは一意である必要があります。"
                        ))
        
        if "shift" not in xl:
            if kintai_path and kintai_path.exists():
                pass
            else:
                issues.append(ValidationIssue(
                    severity="ERROR",
                    code="E101",
                    file_name=file_name,
                    sheet="",
                    location="ファイル全体",
                    field="shiftシート",
                    value="見つかりません",
                    expected="shiftシートが必要",
                    hint=f"職種{file_type[-1]}設定ファイルには「shift」という名前のシートが必要です。kintaiファイルを指定するか、shiftシートを追加してください。"
                ))
        else:
            shift_df = xl["shift"]
            required_cols = ["date", "staff"]
            for col in required_cols:
                if col not in shift_df.columns:
                    issues.append(ValidationIssue(
                        severity="ERROR",
                        code="E102",
                        file_name=file_name,
                        sheet="shift",
                        location="列ヘッダー",
                        field=col,
                        value="見つかりません",
                        expected=f"列「{col}」が必要",
                        hint=f"shiftシートに「{col}」列を追加してください。"
                    ))
            
            if "date" in shift_df.columns:
                for idx, row in shift_df.iterrows():
                    try:
                        date_val = str(row["date"]).strip()
                        if pd.isna(row["date"]) or not date_val:
                            continue  # Empty dates are allowed in shift
                        parse_date_token(date_val, target_year, target_month)
                    except ValueError as e:
                        issues.append(ValidationIssue(
                            severity="ERROR",
                            code="E211",
                            file_name=file_name,
                            sheet="shift",
                            location=f"行{idx+2}",
                            field="date",
                            value=str(row["date"]),
                            expected="dd (例: 1, 15, 21) または yyyy-mm-dd (例: 2025-10-21)",
                            hint=f"日付の形式が正しくありません。このアプリでは「{target_year}年{target_month}月」を選択すると、{target_year}年{target_month}月21日〜{target_month+1 if target_month<12 else 1}月20日の期間になります。dd=1〜20は翌月として扱われます。"
                        ))
        
        if "dict" not in xl:
            issues.append(ValidationIssue(
                severity="ERROR",
                code="E101",
                file_name=file_name,
                sheet="",
                location="ファイル全体",
                field="dictシート",
                value="見つかりません",
                expected="dictシートが必要",
                hint=f"職種{file_type[-1]}設定ファイルには「dict」という名前のシートが必要です。"
            ))
        else:
            dict_df = xl["dict"]
            if "aka" not in dict_df.columns or "name" not in dict_df.columns:
                issues.append(ValidationIssue(
                    severity="ERROR",
                    code="E102",
                    file_name=file_name,
                    sheet="dict",
                    location="列ヘッダー",
                    field="aka, name",
                    value="見つかりません",
                    expected="列「aka」と「name」が必要",
                    hint="dictシートには「aka」列と「name」列が必要です。"
                ))
    
    return issues

def derive_shift_from_kintai(kintai_df: pd.DataFrame, members_df: pd.DataFrame, 
                             target_year: int, target_month: int, cutoff_day: int = 21) -> Tuple[pd.DataFrame, Dict[str, Dict[str, str]]]:
    """
    Derive shift availability data from kintai (attendance) file.
    
    Args:
        kintai_df: DataFrame with columns [name, aka, 21, 22, ..., 31, 1, 2, ..., 20]
                   Cell values:
                   - Empty = fully available
                   - "1" = full day off (no assignments)
                   - "A" = AM off (no assignments before noon)
                   - "P" = PM off (no assignments after noon)
                   - "D" = duty off (no close duties)
                   - Other non-empty = full day off
        members_df: DataFrame with aka column (list of valid member IDs)
        target_year: Target year
        target_month: Target month
        cutoff_day: Day of month where period starts (default 21)
    
    Returns:
        Tuple of:
        - DataFrame with columns [date, day, staff] matching the old shift sheet format
        - Dict mapping date_iso -> aka -> leave_type ('1', 'A', 'P', 'D', or '')
    """
    import calendar
    
    if 'aka' not in kintai_df.columns:
        raise ValueError("E102: kintaiファイルに'aka'列が見つかりません。")
    
    if 'aka' not in members_df.columns:
        raise ValueError("E102: membersシートに'aka'列が見つかりません。")
    
    all_members_aka = set(members_df['aka'].astype(str))
    # NaNや空文字列を除外
    all_members_aka = {aka for aka in all_members_aka if aka and aka.lower() != 'nan' and aka.strip() != ''}
    kintai_aka = set(kintai_df['aka'].astype(str))
    # NaNや空文字列を除外
    kintai_aka = {aka for aka in kintai_aka if aka and aka.lower() != 'nan' and aka.strip() != ''}
    
    extra_aka = kintai_aka - all_members_aka
    if extra_aka:
        logging.warning(f"kintaiファイルに未知のaka: {', '.join(sorted(extra_aka))}")
        # Resolve aka to display names for GUI warning
        for aka in sorted(extra_aka):
            rows = kintai_df[kintai_df['aka'].astype(str) == aka]
            if not rows.empty and 'name' in kintai_df.columns:
                name = str(rows.iloc[0]['name']).strip()
                if name and name.lower() != 'nan':
                    _UNREGISTERED_STAFF.append(name)
                else:
                    _UNREGISTERED_STAFF.append(aka)
            else:
                _UNREGISTERED_STAFF.append(aka)
    
    # load列からreserve（ダミー）スタッフを特定
    reserve_aka = set()
    if 'load' in members_df.columns:
        for _, r in members_df.iterrows():
            aka_str = str(r['aka']).strip()
            if aka_str and aka_str.lower() != 'nan' and str(r.get('load', '')).strip().lower() == 'reserve':
                reserve_aka.add(aka_str)

    missing_aka = all_members_aka - kintai_aka
    missing_real = missing_aka - reserve_aka
    missing_reserve = missing_aka & reserve_aka
    if missing_real:
        logging.warning(f"kintaiファイルに存在しない実メンバー（全日不可として扱います）: {', '.join(sorted(missing_real))}")
    if missing_reserve:
        logging.info(f"kintaiファイルに存在しないダミースタッフ（全日可能として扱います）: {', '.join(sorted(missing_reserve))}")
    
    last_day_of_month = calendar.monthrange(target_year, target_month)[1]
    next_month = target_month + 1 if target_month < 12 else 1
    next_year = target_year if target_month < 12 else target_year + 1
    
    period_dates = []
    for day in range(cutoff_day, last_day_of_month + 1):
        date_obj = dt.date(target_year, target_month, day)
        period_dates.append((day, date_obj))
    
    for day in range(1, cutoff_day):
        date_obj = dt.date(next_year, next_month, day)
        period_dates.append((day, date_obj))
    
    shift_rows = []
    leave_type_by_date: Dict[str, Dict[str, str]] = {}
    
    # akaの数値順でソート（A1, A2, ..., A10, A11, ...）
    def aka_sort_key(aka_val):
        """akaから数値部分を抽出してソートキーを生成"""
        aka_str = str(aka_val)
        prefix = ''
        num_str = ''
        for c in aka_str:
            if c.isdigit():
                num_str += c
            else:
                if not num_str:
                    prefix += c
        num = int(num_str) if num_str else 0
        return (prefix, num)
    
    sorted_members_aka = sorted(all_members_aka, key=aka_sort_key)
    
    for day_num, date_obj in period_dates:
        weekday_name = get_japanese_weekday(date_obj.weekday())
        date_iso = date_obj.strftime("%Y-%m-%d")
        
        available_staff = []
        leave_type_by_date[date_iso] = {}
        
        for aka in sorted_members_aka:
            leave_type = ''
            is_available = True
            
            if aka not in kintai_aka and aka in missing_real:
                # settingにいるがinputにいない実メンバーは全日不可
                leave_type = '1'
                is_available = False
            elif aka in kintai_aka:
                kintai_row = kintai_df[kintai_df['aka'] == aka].iloc[0]
                
                if day_num in kintai_row.index:
                    cell_value = kintai_row[day_num]
                    
                    if pd.notna(cell_value) and str(cell_value).strip() != '':
                        cell_str = str(cell_value).strip().upper()
                        if cell_str == '1':
                            leave_type = '1'
                            is_available = False
                        elif cell_str == 'A':
                            leave_type = 'A'
                            is_available = True
                        elif cell_str == 'P':
                            leave_type = 'P'
                            is_available = True
                        elif cell_str == 'D':
                            leave_type = 'D'
                            is_available = True
                        else:
                            leave_type = '1'
                            is_available = False
            
            leave_type_by_date[date_iso][aka] = leave_type
            
            if is_available:
                available_staff.append(aka)
        
        staff_str = ', '.join(available_staff)
        
        shift_rows.append({
            'date': day_num,
            'day': weekday_name,
            'staff': staff_str
        })
    
    shift_df = pd.DataFrame(shift_rows)
    return shift_df, leave_type_by_date


# 勤務入力表からkintai/dutyを抽出するための定数
KINMU_REST_CODES = {'A', 'B', 'C', 'D', 'E', '四季', '有', '除外', '欠', '代', '指'}
KINMU_SPECIAL_CODES = {'PM', '締め×', '締め\n×'}


def normalize_kinmu_value(val) -> str:
    """勤務入力表のセル値を正規化（改行除去、空白トリム）"""
    if pd.isna(val) or val == '':
        return ''
    return str(val).replace('\n', '').strip()


def is_kinmu_rest_code(val) -> bool:
    """休み系コードかどうか判定"""
    normalized = normalize_kinmu_value(val)
    return normalized in KINMU_REST_CODES


def extract_kintai_value_from_kinmu(am_val, pm_val) -> Tuple[str, Optional[int], bool]:
    """AM/PM値からkintai値を抽出
    
    Args:
        am_val: AM列の値
        pm_val: PM列の値
    
    Returns:
        tuple: (kintai_value, end_time_constraint, no_close_duty)
        - kintai_value: '1', 'A', 'P', 'D', or ''
        - end_time_constraint: int (e.g., 1430) or None
        - no_close_duty: True if 締め業務に割り当て禁止
    """
    am_norm = normalize_kinmu_value(am_val)
    pm_norm = normalize_kinmu_value(pm_val)
    
    # 「PM」「PM」パターン → A（午前休み、PMのみ入れる）
    if am_norm.upper() == 'PM' and pm_norm.upper() == 'PM':
        return 'A', None, False
    
    # 「14時」「30分」パターン → 終了時間制約
    if '時' in am_norm and '分' in pm_norm:
        try:
            hour = int(am_norm.replace('時', ''))
            minute = int(pm_norm.replace('分', ''))
            end_time = hour * 100 + minute
            return '', end_time, False
        except ValueError:
            pass
    
    # 「直入り」「直入」パターン（当直入り）→ 締め業務禁止のみ、kintai値なし
    if '直入' in am_norm or '直入' in pm_norm:
        return '', None, True
    
    # 「直明け」「直明」パターン（当直明け）→ D（業務割り当てられにくい）
    if '直明' in am_norm or '直明' in pm_norm:
        return 'D', None, False
    
    # AM/PM両方に休み系コード → 1（1日休み）
    am_is_rest = is_kinmu_rest_code(am_val)
    pm_is_rest = is_kinmu_rest_code(pm_val)
    
    if am_is_rest and pm_is_rest:
        return '1', None, False
    
    # AMのみ休み系コード、PM空欄 → A（午前休み）
    if am_is_rest and pm_norm == '':
        return 'A', None, False
    
    # AM空欄、PMのみ休み系コード → P（午後休み）
    if am_norm == '' and pm_is_rest:
        return 'P', None, False
    
    return '', None, False


def extract_duty_from_kinmu_cell(val, duty_times: Dict[str, Dict[str, int]], time_slot: str) -> Optional[Tuple[str, int, int]]:
    """勤務入力表のセル値からduty情報を抽出
    
    Args:
        val: セル値
        duty_times: duty名 -> {'start': int, 'end': int} のマッピング
        time_slot: 'AM' or 'PM'
    
    Returns:
        tuple: (duty_name, start, end) or None
    """
    if pd.isna(val) or val == '':
        return None
    
    val_str = str(val).strip()
    val_normalized = val_str.replace('\n', '')
    
    # 数字のみの値はdutyとみなさない
    if val_normalized.isdigit():
        return None
    
    # 休み系コードは除外
    if val_normalized in KINMU_REST_CODES:
        return None
    
    # 特殊コードは除外（PM, 締め×）
    if val_normalized in {'PM', '締め×'}:
        return None
    
    # 時間パターンは除外
    if '時' in val_normalized or '分' in val_normalized:
        return None
    
    # 直入りは除外（kintaiで処理）
    if '直入' in val_normalized:
        return None
    
    # duty_templateに定義されている業務
    if val_normalized in duty_times:
        return (val_normalized, duty_times[val_normalized]['start'], duty_times[val_normalized]['end'])
    
    # 定義されていない業務はデフォルト時間帯で個人業務として扱う
    # AM: 830-1230, PM: 1230-1700
    if time_slot == 'AM':
        return (val_normalized, 830, 1230)
    else:
        return (val_normalized, 1230, 1700)


def extract_kintai_and_duty_from_kinmu(
    kinmu_path: Path,
    duty_template_path: Optional[Path],
    members_df: pd.DataFrame,
    target_year: int,
    target_month: int
) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, Dict[str, int]], Dict[str, Set[str]]]:
    """勤務入力表からkintaiとdutyデータを抽出
    
    Args:
        kinmu_path: 勤務入力表ファイルのパス
        duty_template_path: duty_templateファイルのパス（時間帯参照用）
        members_df: メンバー情報DataFrame（aka列を含む）
        target_year: 対象年
        target_month: 対象月
    
    Returns:
        tuple: (kintai_df, duty_df, end_time_constraints, no_close_duty_constraints)
        - kintai_df: kintaiデータ（name, aka, 21, 22, ..., 20）
        - duty_df: dutyデータ（duty, from, till, start, end, name）
        - end_time_constraints: {date_iso: {aka: end_time}} 終了時間制約
        - no_close_duty_constraints: {date_iso: set(aka)} 締め業務禁止制約（当直入り）
    """
    # 勤務入力表を読み込み
    try:
        xl = pd.ExcelFile(kinmu_path)
    except Exception as e:
        raise ValueError(f"E201: 勤務入力表の読み込みに失敗しました: {e}")
    
    # 「勤務入力表」シートを探す
    sheet_name = None
    for name in xl.sheet_names:
        if '勤務入力表' in name:
            sheet_name = name
            break
    
    if sheet_name is None:
        raise ValueError("E202: 勤務入力表シートが見つかりません。")
    
    df = pd.read_excel(xl, sheet_name=sheet_name, header=None)
    
    # duty_templateから時間帯を取得
    duty_times: Dict[str, Dict[str, int]] = {}
    if duty_template_path and duty_template_path.exists():
        try:
            duty_xl = pd.ExcelFile(duty_template_path)
            duty_df_template = pd.read_excel(duty_xl, sheet_name='fix')
            for _, row in duty_df_template.iterrows():
                duty_name = str(row['duty']).strip()
                if duty_name not in duty_times:
                    duty_times[duty_name] = {'start': int(row['start']), 'end': int(row['end'])}
        except Exception as e:
            logging.warning(f"duty_templateの読み込みに失敗しました: {e}")
    
    # メンバーのakaマッピングを作成
    name_to_aka: Dict[str, str] = {}
    aka_to_display_name: Dict[str, str] = {}  # akaからjobAのnameへのマッピング
    surname_to_akas: Dict[str, List[Tuple[str, str]]] = {}  # 姓 -> [(name, aka), ...]（同姓対応）
    import re as _re
    if 'name' in members_df.columns and 'aka' in members_df.columns:
        for _, row in members_df.iterrows():
            name = str(row['name']).strip()
            aka = str(row['aka']).strip()
            if not aka or aka.lower() == 'nan':
                continue
            # 括弧を除去して姓を抽出（例: 後藤（奈）→ 後藤）
            name_base = _re.sub(r'[（(].+?[）)]', '', name).strip()
            name_short = name_base.split()[0] if ' ' in name_base else name_base.split('　')[0] if '　' in name_base else name_base
            name_to_aka[name] = aka
            surname_to_akas.setdefault(name_short, []).append((name, aka))
            # 括弧付き名（例: 後藤（奈））から区別文字を抽出して追加キーを登録
            paren_match = _re.search(r'[（(](.+?)[）)]', name)
            if paren_match:
                disambig_char = paren_match.group(1)
                name_to_aka[f"{name_short}\x00{disambig_char[0]}"] = aka
            if aka not in aka_to_display_name:
                aka_to_display_name[aka] = name
        # 同姓が1人だけの場合のみ姓単独マッピングを登録
        for surname, members_list in surname_to_akas.items():
            if len(members_list) == 1:
                name_to_aka[surname] = members_list[0][1]
    
    # 日付行（行3）を取得
    date_row = df.iloc[3]
    
    # スタッフ行の開始位置を特定（行6から）
    staff_start_row = 6
    
    # kintaiデータを構築
    kintai_rows = []
    duty_rows = []
    end_time_constraints: Dict[str, Dict[str, int]] = {}
    no_close_duty_constraints: Dict[str, Set[str]] = {}
    
    consecutive_empty = 0
    for i in range(staff_start_row, len(df)):
        name_cell = df.iloc[i, 1]
        if pd.isna(name_cell) or str(name_cell).strip() == '':
            consecutive_empty += 1
            if consecutive_empty >= 2:
                break  # 2行連続空行でスタッフリスト終了
            continue
        consecutive_empty = 0
        
        name_full = str(name_cell).strip()
        # 姓のみ抽出
        name_short = name_full.split()[0] if ' ' in name_full else name_full.split('　')[0] if '　' in name_full else name_full
        
        # akaを取得（存在しない場合はスキップ）
        aka = name_to_aka.get(name_full)
        if not aka and len(surname_to_akas.get(name_short, [])) <= 1:
            aka = name_to_aka.get(name_short)
        if not aka:
            # 同姓メンバーの区別: 入力名の名部分の先頭文字で括弧付き設定名とマッチ
            given_name = ''
            if '　' in name_full:
                given_name = name_full.split('　', 1)[1]
            elif ' ' in name_full:
                given_name = name_full.split(' ', 1)[1]
            if given_name:
                aka = name_to_aka.get(f"{name_short}\x00{given_name[0]}")
        if not aka:
            logging.warning(f"Unknown member in kinmu sheet: {name_full}")
            continue
        
        # 表示名を取得（jobAのnameを使用、なければname_full）
        display_name = aka_to_display_name.get(aka, name_full)
        
        # kintai行を初期化
        kintai_row = {'name': display_name, 'aka': aka}
        
        # 各日のAM/PM値を処理
        for day_idx in range(31):  # 最大31日分
            col_am = 2 + day_idx * 2
            col_pm = 3 + day_idx * 2
            
            if col_pm >= len(df.columns):
                break
            
            date_val = date_row.iloc[col_am]
            if not isinstance(date_val, dt.datetime):
                continue
            
            date_obj = date_val.date() if hasattr(date_val, 'date') else date_val
            day_num = date_obj.day
            date_iso = date_obj.strftime('%Y-%m-%d')
            
            am_val = df.iloc[i, col_am]
            pm_val = df.iloc[i, col_pm]
            
            # kintai値を抽出
            kintai_val, end_time, no_close_duty = extract_kintai_value_from_kinmu(am_val, pm_val)
            
            if kintai_val:
                kintai_row[day_num] = kintai_val
            
            # 終了時間制約を記録
            if end_time:
                if date_iso not in end_time_constraints:
                    end_time_constraints[date_iso] = {}
                end_time_constraints[date_iso][aka] = end_time
            
            # 締め業務禁止制約を記録（当直入り）
            if no_close_duty and aka:
                if date_iso not in no_close_duty_constraints:
                    no_close_duty_constraints[date_iso] = set()
                no_close_duty_constraints[date_iso].add(aka)
            
            # duty値を抽出（時間パターンでない場合のみ）
            am_norm = normalize_kinmu_value(am_val)
            pm_norm = normalize_kinmu_value(pm_val)
            if not ('時' in am_norm and '分' in pm_norm):
                # AM値からduty抽出
                duty_info = extract_duty_from_kinmu_cell(am_val, duty_times, 'AM')
                if duty_info:
                    duty_rows.append({
                        'duty': duty_info[0],
                        'from': day_num,
                        'till': day_num,
                        'start': duty_info[1],
                        'end': duty_info[2],
                        'name': display_name
                    })
                
                # PM値からduty抽出
                duty_info = extract_duty_from_kinmu_cell(pm_val, duty_times, 'PM')
                if duty_info:
                    duty_rows.append({
                        'duty': duty_info[0],
                        'from': day_num,
                        'till': day_num,
                        'start': duty_info[1],
                        'end': duty_info[2],
                        'name': display_name
                    })
        
        kintai_rows.append(kintai_row)
    
    # DataFrameを作成
    kintai_df = pd.DataFrame(kintai_rows)
    duty_df = pd.DataFrame(duty_rows)
    
    # kintai_dfの列順を整理（name, aka, 21, 22, ..., 31, 1, 2, ..., 20）
    date_cols = [c for c in kintai_df.columns if isinstance(c, int)]
    # 21〜31, 1〜20の順に並べ替え
    date_cols_sorted = sorted([c for c in date_cols if c >= 21]) + sorted([c for c in date_cols if c < 21])
    kintai_df = kintai_df[['name', 'aka'] + date_cols_sorted]
    
    # akaの数値順でソート（A1, A2, ..., A10, A11, ...）
    def aka_sort_key(aka_val):
        """akaから数値部分を抽出してソートキーを生成"""
        aka_str = str(aka_val)
        # 先頭のアルファベット部分と数値部分を分離
        prefix = ''
        num_str = ''
        for c in aka_str:
            if c.isdigit():
                num_str += c
            else:
                if not num_str:
                    prefix += c
        num = int(num_str) if num_str else 0
        return (prefix, num)
    
    if 'aka' in kintai_df.columns and len(kintai_df) > 0:
        kintai_df['_sort_key'] = kintai_df['aka'].apply(aka_sort_key)
        kintai_df = kintai_df.sort_values('_sort_key').drop(columns=['_sort_key']).reset_index(drop=True)
    
    return kintai_df, duty_df, end_time_constraints, no_close_duty_constraints


class Setting:
    def __init__(self, path: Path):
        xl = pd.read_excel(path, sheet_name=None)
        self.penalties = xl.get("penalties")
        self.exception = xl.get("exception")
        self.config = xl.get("config")

        self.jobA_df = xl.get("jobA")
        self.orderA_df = xl.get("orderA")
        self.dictA_df = xl.get("dictA")
        self.jobB_df = xl.get("jobB")
        self.orderB_df = xl.get("orderB")
        self.dictB_df = xl.get("dictB")
        self.weekly_work_df = xl.get("weekly_work")
        self.dict_work_df = xl.get("dict_work")

        self.penalty_map = {}
        if self.penalties is not None:
            for _, r in self.penalties.iterrows():
                name = str(r.get("name")).strip()
                thres = int(r.get("thres")) if not pd.isna(r.get("thres")) else 0
                value = float(r.get("value")) if not pd.isna(r.get("value")) else 0.0
                typ = str(r.get("type")).strip() if not pd.isna(r.get("type")) else ""
                self.penalty_map[name] = {"thres": thres, "value": value, "type": typ}

        self.config_map = {
            "workday_end": 1800,
            "week_start": "Mon",
            "min_interval": 0,
            "rounds_minutes": 10,
            "random_seed": 42,
            "noon": 1230,
        }
        if self.config is not None:
            for _, r in self.config.iterrows():
                k = str(r.get("key")).strip()
                v = r.get("value")
                if k == "workday_end":
                    self.config_map[k] = int(v)
                elif k in ("min_interval", "rounds_minutes", "random_seed", "noon"):
                    self.config_map[k] = int(v)
                elif k == "week_start":
                    self.config_map[k] = str(v)

        self.exceptions = []
        if self.exception is not None:
            for _, r in self.exception.iterrows():
                self.exceptions.append(
                    {
                        "penalty": str(r.get("penalty")).strip(),
                        "target": str(r.get("target")).strip(),
                        "vector": str(r.get("vector")).strip(),
                        "scope": str(r.get("scope")).strip() if not pd.isna(r.get("scope")) else "*",
                    }
                )
        
        self.duty_priority_A = []
        if self.orderA_df is not None and "duty" in self.orderA_df.columns:
            for val in self.orderA_df["duty"].dropna():
                duty_name = str(val).strip()
                if duty_name:
                    self.duty_priority_A.append(duty_name)
        
        self.duty_priority_B = []
        if self.orderB_df is not None and "duty" in self.orderB_df.columns:
            for val in self.orderB_df["duty"].dropna():
                duty_name = str(val).strip()
                if duty_name:
                    self.duty_priority_B.append(duty_name)

    def get_reserve_names(self) -> Set[str]:
        """Return display names of all reserve (dummy) staff from jobA/jobB."""
        reserve_names: Set[str] = set()
        for job_df, dict_df in [(self.jobA_df, self.dictA_df), (self.jobB_df, self.dictB_df)]:
            if job_df is None or 'load' not in job_df.columns:
                continue
            reserve_akas = set()
            for _, r in job_df.iterrows():
                if str(r.get('load', '')).strip().lower() == 'reserve':
                    reserve_akas.add(str(r['aka']).strip())
            if dict_df is not None and 'aka' in dict_df.columns and 'name' in dict_df.columns:
                for _, r in dict_df.iterrows():
                    aka = str(r['aka']).strip()
                    if aka in reserve_akas:
                        name = str(r['name']).strip()
                        if name and name.lower() != 'nan':
                            reserve_names.add(name)
        return reserve_names


class JobData:
    def __init__(self, path: Path, role_prefix: str, target_year: int, target_month: int, kintai_path: Optional[Path] = None, output_dir: Optional[Path] = None):
        xl = pd.read_excel(path, sheet_name=None)
        self.members = xl["members"].copy()
        self.dict = xl["dict"].copy()
        
        shift_generated = False
        self.leave_type_by_date: Dict[str, Dict[str, str]] = {}
        if kintai_path and kintai_path.exists():
            logging.info(f"勤怠ファイルから勤務可能日を生成: {kintai_path}")
            kintai_xl = pd.read_excel(kintai_path, sheet_name=None)
            kintai_df = kintai_xl[list(kintai_xl.keys())[0]].copy()
            
            self.shift, self.leave_type_by_date = derive_shift_from_kintai(kintai_df, self.members, target_year, target_month, cutoff_day=PERIOD_CUTOFF_DAY)
            shift_generated = True
            
            if output_dir:
                debug_shift_path = output_dir / f"generated_shift_{role_prefix}.xlsx"
                self.shift.to_excel(debug_shift_path, index=False)
                logging.info(f"生成されたshiftデータを保存: {debug_shift_path}")
        elif "shift" in xl:
            self.shift = xl["shift"].copy()
        else:
            raise ValueError(f"E101: {path}に'shift'シートが見つかりません。kintaiファイルを指定するか、shiftシートを追加してください.")
        
        if shift_generated and "shift" not in xl:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(path)
                
                from openpyxl.utils.dataframe import dataframe_to_rows
                ws = wb.create_sheet('shift')
                
                for r_idx, row in enumerate(dataframe_to_rows(self.shift, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                
                wb.save(path)
                logging.info(f"生成されたshiftシートを{path}に保存しました")
            except Exception as e:
                logging.warning(f"shiftシートを{path}に保存できませんでした: {e}")

        self.members["prefer_tokens"] = self.members["prefer"].apply(parse_tokens) if "prefer" in self.members else set()
        self.members["avoid_tokens"] = self.members["avoid"].apply(parse_tokens) if "avoid" in self.members else set()
        self.members["aka"] = self.members["aka"].astype(str)

        self.role_prefix = role_prefix  # "A" or "B"
        self.aka_to_name = {str(r["aka"]): str(r["name"]) for _, r in self.dict.iterrows()}

        self.shift["date_iso"] = self.shift["date"].apply(lambda t: parse_date_token(t, target_year, target_month))
        if "day" in self.shift:
            self.shift["day_norm"] = self.shift["day"].apply(normalize_weekday_token)
        else:
            self.shift["day_norm"] = ""

        self.avail_by_date: Dict[str, Set[str]] = {}
        for _, r in self.shift.iterrows():
            date = str(r["date_iso"])
            staffs = parse_tokens(r["staff"])
            self.avail_by_date.setdefault(date, set()).update(staffs)

        self.senior_set = set(self.members[self.members.get("senior", 0) == 1]["aka"].astype(str))
        self.over_flag = {str(r["aka"]): int(r.get("over", 0)) for _, r in self.members.iterrows()}
        self.load = {str(r["aka"]): str(r.get("load", "normal")) for _, r in self.members.iterrows()}

        self.order_priority: Dict[str, Dict[str, float]] = {}
        if "order" in xl:
            order_df = xl["order"]
            import hashlib
            for col in order_df.columns:
                dept = str(col)
                non_null_values = order_df[col].dropna().tolist()
                total_members = len(non_null_values)
                if total_members == 0:
                    continue
                self.order_priority[dept] = {}
                for rank, aka in enumerate(non_null_values):
                    aka_str = str(aka)
                    base_score = (rank + 1) / total_members
                    h = hashlib.md5(f"{dept}:{aka_str}".encode("utf-8")).hexdigest()
                    jitter = int(h[:8], 16) / (2**32) * 1e-6
                    self.order_priority[dept][aka_str] = base_score + jitter

    def get_order_score(self, dept: str, aka: str) -> Optional[float]:
        """Get the order priority score for a staff member in a department"""
        dept_map = self.order_priority.get(dept)
        if not dept_map:
            return None
        return dept_map.get(aka)

    @classmethod
    def from_setting(cls, setting: 'Setting', role_prefix: str, target_year: int, target_month: int, 
                     kintai_path: Optional[Path] = None, output_dir: Optional[Path] = None) -> Optional['JobData']:
        """Create JobData from Setting object instead of separate file"""
        if role_prefix == "A":
            members_df = setting.jobA_df
            order_df = setting.orderA_df
            dict_df = setting.dictA_df
        else:
            members_df = setting.jobB_df
            order_df = setting.orderB_df
            dict_df = setting.dictB_df
        
        if members_df is None or members_df.empty:
            return None
        if dict_df is None or dict_df.empty:
            return None
            
        instance = cls.__new__(cls)
        instance.members = members_df.copy()
        instance.dict = dict_df.copy()
        
        shift_generated = False
        instance.leave_type_by_date = {}
        
        if kintai_path and kintai_path.exists():
            logging.info(f"勤怠ファイルから勤務可能日を生成: {kintai_path}")
            kintai_xl = pd.read_excel(kintai_path, sheet_name=None)
            kintai_df = kintai_xl[list(kintai_xl.keys())[0]].copy()
            
            instance.shift, instance.leave_type_by_date = derive_shift_from_kintai(
                kintai_df, instance.members, target_year, target_month, cutoff_day=PERIOD_CUTOFF_DAY
            )
            shift_generated = True
            
            if output_dir:
                debug_shift_path = output_dir / f"generated_shift_{role_prefix}.xlsx"
                instance.shift.to_excel(debug_shift_path, index=False)
                logging.info(f"生成されたshiftデータを保存: {debug_shift_path}")
        else:
            raise ValueError(f"E101: kintaiファイルが指定されていないか、存在しません。kintai{role_prefix}ファイルを指定してください。")
        
        instance.members["prefer_tokens"] = instance.members["prefer"].apply(parse_tokens) if "prefer" in instance.members else instance.members.apply(lambda _: set(), axis=1)
        instance.members["avoid_tokens"] = instance.members["avoid"].apply(parse_tokens) if "avoid" in instance.members else instance.members.apply(lambda _: set(), axis=1)
        instance.members["aka"] = instance.members["aka"].astype(str)
        
        instance.role_prefix = role_prefix
        instance.aka_to_name = {str(r["aka"]): str(r["name"]) for _, r in instance.dict.iterrows()}
        
        instance.shift["date_iso"] = instance.shift["date"].apply(lambda t: parse_date_token(t, target_year, target_month))
        if "day" in instance.shift:
            instance.shift["day_norm"] = instance.shift["day"].apply(normalize_weekday_token)
        else:
            instance.shift["day_norm"] = ""
        
        instance.avail_by_date = {}
        for _, r in instance.shift.iterrows():
            date = str(r["date_iso"])
            staffs = parse_tokens(r["staff"])
            instance.avail_by_date.setdefault(date, set()).update(staffs)
        
        instance.senior_set = set(instance.members[instance.members.get("senior", 0) == 1]["aka"].astype(str))
        instance.over_flag = {str(r["aka"]): int(r.get("over", 0)) for _, r in instance.members.iterrows()}
        instance.load = {str(r["aka"]): str(r.get("load", "normal")) for _, r in instance.members.iterrows()}

        # 全日不可のメンバーを特定（どの日にも出勤可能でないメンバー）
        all_available = set()
        for avail_set in instance.avail_by_date.values():
            all_available.update(avail_set)
        all_member_akas = set(instance.members["aka"].astype(str))
        never_available = all_member_akas - all_available
        if never_available:
            reserve_only = {a for a in never_available if instance.load.get(a, "normal") == "reserve"}
            real_absent = never_available - reserve_only
            if real_absent:
                logging.info(f"全日不可の実在メンバー（order{role_prefix}の優先順位から除外）: {', '.join(sorted(real_absent))}")

        instance.order_priority = {}
        if order_df is not None and not order_df.empty:
            import hashlib
            for col in order_df.columns:
                dept = str(col)
                # 全日不可の実在メンバーをorderから除外
                non_null_values = [v for v in order_df[col].dropna().tolist() if str(v) not in never_available or instance.load.get(str(v), "normal") == "reserve"]
                total_members = len(non_null_values)
                if total_members == 0:
                    continue
                instance.order_priority[dept] = {}
                for rank, aka in enumerate(non_null_values):
                    aka_str = str(aka)
                    base_score = (rank + 1) / total_members
                    h = hashlib.md5(f"{dept}:{aka_str}".encode("utf-8")).hexdigest()
                    jitter = int(h[:8], 16) / (2**32) * 1e-6
                    instance.order_priority[dept][aka_str] = base_score + jitter
        
        return instance


class WorkData:
    def __init__(self, path: Path, target_year: int, target_month: int):
        self.target_year = target_year
        self.target_month = target_month

        xl = pd.read_excel(path, sheet_name=None)
        self.work = xl["work"].copy()
        self.dict = xl["dict"].copy()

        self.work["date"] = self.work["date"].apply(lambda t: parse_date_token(t, target_year, target_month))
        if "day" in self.work:
            self.work["day_original"] = self.work["day"].copy()
            self.work["day"] = self.work["day"].apply(normalize_weekday_token)

        self.work["start_min"] = self.work["start"].apply(hhmm_to_minutes)
        self.work["end_min"] = (self.work["start_min"] + (self.work["duration"] * 60).astype(int))
        self.work["id"] = self.work["id"].astype(int)
        self.work["avoid_tokens"] = self.work["avoid"].apply(parse_tokens) if "avoid" in self.work else self.work.apply(lambda _: set(), axis=1)
        self.room_name = {str(r["aka"]): str(r["name"]) for _, r in self.dict.iterrows()}
        self.dept_name = {str(r["aka"]): str(r["name"]) for _, r in self.dict.iterrows()}
        
        self.work = self.work[~self.work["dept"].isin(["D16", "D17", "D18"])].copy()

    @classmethod
    def from_setting(cls, setting: 'Setting', target_year: int, target_month: int) -> 'WorkData':
        """Create WorkData from Setting object by expanding weekly_work template to actual dates"""
        import jpholiday
        
        if setting.weekly_work_df is None or setting.weekly_work_df.empty:
            raise ValueError("E102: setting.xlsxにweekly_workシートが見つかりません。")
        if setting.dict_work_df is None or setting.dict_work_df.empty:
            raise ValueError("E103: setting.xlsxにdict_workシートが見つかりません。")
        
        instance = cls.__new__(cls)
        instance.target_year = target_year
        instance.target_month = target_month
        instance.dict = setting.dict_work_df.copy()
        
        period_start = dt.date(target_year, target_month, PERIOD_CUTOFF_DAY)
        if target_month == 12:
            period_end = dt.date(target_year + 1, 1, PERIOD_CUTOFF_DAY - 1)
        else:
            period_end = dt.date(target_year, target_month + 1, PERIOD_CUTOFF_DAY - 1)
        
        japanese_weekday_map = {
            0: "月曜日",
            1: "火曜日",
            2: "水曜日",
            3: "木曜日",
            4: "金曜日",
            5: "土曜日",
            6: "日曜日",
        }
        
        dates_with_weekday = []
        cur = period_start
        while cur <= period_end:
            is_holiday = jpholiday.is_holiday(cur)
            is_sunday = cur.weekday() == 6
            
            if is_holiday or is_sunday:
                cur += dt.timedelta(days=1)
                continue
            
            day_jp = japanese_weekday_map[cur.weekday()]
            dates_with_weekday.append((cur, day_jp))
            cur += dt.timedelta(days=1)
        
        records = []
        weekly_work = setting.weekly_work_df
        
        # まず全てのレコードを作成（idなし）
        for _, base in weekly_work.iterrows():
            template_day = str(base.get("day", "")).strip()
            if not template_day:
                continue
                
            for date_obj, day_jp in dates_with_weekday:
                if day_jp == template_day:
                    rec = {
                        "date": date_obj.isoformat(),
                        "day": template_day,
                        "room": base.get("room"),
                        "dept": base.get("dept"),
                        "start": base.get("start"),
                        "avoid": base.get("avoid") if not pd.isna(base.get("avoid")) else "",
                        "duration": base.get("duration"),
                        "need_A": base.get("need_A"),
                        "need_B": base.get("need_B"),
                    }
                    records.append(rec)
        
        if not records:
            raise ValueError("E104: weekly_workテンプレートから業務を生成できませんでした。曜日の形式を確認してください（例: 月曜日, 火曜日）。")
        
        instance.work = pd.DataFrame(records)
        
        instance.work["date"] = instance.work["date"].apply(lambda t: parse_date_token(t, target_year, target_month))
        if "day" in instance.work:
            instance.work["day_original"] = instance.work["day"].copy()
            instance.work["day"] = instance.work["day"].apply(normalize_weekday_token)
        
        instance.work["start_min"] = instance.work["start"].apply(hhmm_to_minutes)
        instance.work["end_min"] = (instance.work["start_min"] + (instance.work["duration"] * 60).astype(int))
        instance.work["avoid_tokens"] = instance.work["avoid"].apply(parse_tokens) if "avoid" in instance.work else instance.work.apply(lambda _: set(), axis=1)
        instance.room_name = {str(r["aka"]): str(r["name"]) for _, r in instance.dict.iterrows()}
        instance.dept_name = {str(r["aka"]): str(r["name"]) for _, r in instance.dict.iterrows()}
        
        # D16/D17/D18をフィルタ
        instance.work = instance.work[~instance.work["dept"].isin(["D16", "D17", "D18"])].copy()
        
        # 日付→dept番号→start_time順にソートしてidを付与
        def dept_sort_key(dept_val):
            dept_str = str(dept_val)
            # "D1", "D10"などから数値部分を抽出
            match = re.match(r'D?(\d+)', dept_str)
            if match:
                return (0, int(match.group(1)))
            try:
                return (0, int(dept_str))
            except (ValueError, TypeError):
                return (1, dept_str)
        
        instance.work['dept_sort'] = instance.work['dept'].apply(dept_sort_key)
        instance.work = instance.work.sort_values(
            by=['date', 'dept_sort', 'start_min'],
            ascending=[True, True, True]
        ).reset_index(drop=True)
        instance.work['id'] = instance.work.index + 1
        instance.work = instance.work.drop(columns=['dept_sort'])
        
        return instance


def generate_intermediate_files(
    setting: 'Setting',
    target_year: int,
    target_month: int,
    output_dir: Path,
    kintai_df: Optional[pd.DataFrame] = None,
    duty_df: Optional[pd.DataFrame] = None
) -> Tuple[Path, Optional[Path], Path, Optional[Path]]:
    """setting.xlsxから中間ファイル（jobA, jobB, work）を生成してtemporaryフォルダに出力
    
    Args:
        setting: Settingオブジェクト
        target_year: 対象年
        target_month: 対象月
        output_dir: 出力先ディレクトリ（temporaryフォルダ）
        kintai_df: 勤務入力表から抽出したkintaiデータ（オプション）
        duty_df: 勤務入力表から抽出したdutyデータ（オプション）
    
    Returns:
        tuple: (jobA_path, jobB_path, work_path, duty_path)
        - jobB_pathはjobBデータがない場合はNone
        - duty_pathはduty_dfがない場合はNone
    """
    import jpholiday
    from openpyxl import Workbook
    
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # jobA.xlsxを生成
    jobA_path = output_dir / "jobA.xlsx"
    try:
        with pd.ExcelWriter(jobA_path, engine='openpyxl') as writer:
            if setting.jobA_df is not None and not setting.jobA_df.empty:
                setting.jobA_df.to_excel(writer, sheet_name='members', index=False)
            if setting.orderA_df is not None and not setting.orderA_df.empty:
                setting.orderA_df.to_excel(writer, sheet_name='order', index=False)
            if setting.dictA_df is not None and not setting.dictA_df.empty:
                setting.dictA_df.to_excel(writer, sheet_name='dict', index=False)
            
            # kintaiからshiftを生成
            if kintai_df is not None and not kintai_df.empty:
                shift_df, _ = derive_shift_from_kintai(
                    kintai_df, setting.jobA_df, target_year, target_month, cutoff_day=PERIOD_CUTOFF_DAY
                )
                shift_df.to_excel(writer, sheet_name='shift', index=False)
        logging.info(f"中間ファイルを生成: {jobA_path}")
    except PermissionError:
        logging.warning(f"ファイルがロックされています: {jobA_path}")
        raise ValueError(f"E301: ファイルがロックされています。閉じてから再実行してください: {jobA_path}")
    
    # jobB.xlsxを生成（データがある場合のみ）
    jobB_path = None
    if setting.jobB_df is not None and not setting.jobB_df.empty:
        jobB_path = output_dir / "jobB.xlsx"
        try:
            with pd.ExcelWriter(jobB_path, engine='openpyxl') as writer:
                setting.jobB_df.to_excel(writer, sheet_name='members', index=False)
                if setting.orderB_df is not None and not setting.orderB_df.empty:
                    setting.orderB_df.to_excel(writer, sheet_name='order', index=False)
                if setting.dictB_df is not None and not setting.dictB_df.empty:
                    setting.dictB_df.to_excel(writer, sheet_name='dict', index=False)
            logging.info(f"中間ファイルを生成: {jobB_path}")
        except PermissionError:
            logging.warning(f"ファイルがロックされています: {jobB_path}")
            raise ValueError(f"E301: ファイルがロックされています。閉じてから再実行してください: {jobB_path}")
    
    # work.xlsxを生成
    work_path = output_dir / "work.xlsx"
    try:
        # weekly_workを展開
        period_start = dt.date(target_year, target_month, PERIOD_CUTOFF_DAY)
        if target_month == 12:
            period_end = dt.date(target_year + 1, 1, PERIOD_CUTOFF_DAY - 1)
        else:
            period_end = dt.date(target_year, target_month + 1, PERIOD_CUTOFF_DAY - 1)
        
        japanese_weekday_map = {
            0: "月曜日",
            1: "火曜日",
            2: "水曜日",
            3: "木曜日",
            4: "金曜日",
            5: "土曜日",
            6: "日曜日",
        }
        
        dates_with_weekday = []
        cur = period_start
        while cur <= period_end:
            is_holiday = jpholiday.is_holiday(cur)
            is_sunday = cur.weekday() == 6
            
            if is_holiday or is_sunday:
                cur += dt.timedelta(days=1)
                continue
            
            day_jp = japanese_weekday_map[cur.weekday()]
            dates_with_weekday.append((cur, day_jp))
            cur += dt.timedelta(days=1)
        
        records = []
        next_id = 1
        weekly_work = setting.weekly_work_df
        
        if weekly_work is not None and not weekly_work.empty:
            for _, base in weekly_work.iterrows():
                template_day = str(base.get("day", "")).strip()
                if not template_day:
                    continue
                    
                for date_obj, day_jp in dates_with_weekday:
                    if day_jp == template_day:
                        rec = {
                            "id": next_id,
                            "date": date_obj.day,
                            "day": template_day,
                            "room": base.get("room"),
                            "dept": base.get("dept"),
                            "start": base.get("start"),
                            "avoid": base.get("avoid") if not pd.isna(base.get("avoid")) else "",
                            "duration": base.get("duration"),
                            "need_A": base.get("need_A"),
                            "need_B": base.get("need_B"),
                        }
                        records.append(rec)
                        next_id += 1
        
        work_df = pd.DataFrame(records)
        
        with pd.ExcelWriter(work_path, engine='openpyxl') as writer:
            work_df.to_excel(writer, sheet_name='work', index=False)
            if setting.dict_work_df is not None and not setting.dict_work_df.empty:
                setting.dict_work_df.to_excel(writer, sheet_name='dict', index=False)
        logging.info(f"中間ファイルを生成: {work_path}")
    except PermissionError:
        logging.warning(f"ファイルがロックされています: {work_path}")
        raise ValueError(f"E301: ファイルがロックされています。閉じてから再実行してください: {work_path}")
    
    # duty.xlsxを生成（duty_dfがある場合のみ）
    duty_path = None
    if duty_df is not None and not duty_df.empty:
        duty_path = output_dir / "duty.xlsx"
        try:
            with pd.ExcelWriter(duty_path, engine='openpyxl') as writer:
                duty_df.to_excel(writer, sheet_name='fix', index=False)
            logging.info(f"中間ファイルを生成: {duty_path}")
        except PermissionError:
            logging.warning(f"ファイルがロックされています: {duty_path}")
            raise ValueError(f"E301: ファイルがロックされています。閉じてから再実行してください: {duty_path}")
    
    return jobA_path, jobB_path, work_path, duty_path


class DutyTemplate:
    def __init__(self, duty_template_path: str, target_year: int, target_month: int):
        self.duty_template_path = duty_template_path
        self.target_year = target_year
        self.target_month = target_month
        self.assignments = []

        if duty_template_path and Path(duty_template_path).exists():
            self._load_duty_template()
    
    @classmethod
    def from_calendar_sheet(cls, calendar_df: pd.DataFrame, target_year: int, target_month: int) -> 'DutyTemplate':
        """Create DutyTemplate from calendar sheet (カレンダー２) data.
        
        The calendar sheet has format:
        - Row 0: Year/month header
        - Row 1: Column headers (日付, 曜日, then work names with AM/PM)
        - Row 2+: Data rows with date, weekday, and assignments
        
        The 個人業務 column contains entries like "メンバー名(業務名)" for AM and PM.
        """
        instance = cls(None, target_year, target_month)
        
        if calendar_df is None or calendar_df.empty:
            return instance
        
        # Find the 個人業務 columns (AM and PM)
        header_row = calendar_df.iloc[1] if len(calendar_df) > 1 else None
        if header_row is None:
            return instance
        
        duty_am_col = None
        duty_pm_col = None
        
        # Search for 個人業務 in header row
        for col_idx, val in enumerate(header_row):
            if str(val).strip() == '個人業務':
                # Check if next column is AM/PM pattern
                if duty_am_col is None:
                    duty_am_col = col_idx
                elif duty_pm_col is None:
                    duty_pm_col = col_idx
        
        # If we found 個人業務 header, the actual AM/PM are in the row below or same row
        # Based on the calendar structure: header row has work names, and below that AM/PM
        # Actually looking at generate_work_oriented_calendar_view, row1 has work names spanning 2 cols
        # row2 has 日付, 曜日, then AM, PM for each work
        
        # Find columns by looking at row 0 for work names
        header_row0 = calendar_df.iloc[0] if len(calendar_df) > 0 else None
        if header_row0 is not None:
            for col_idx, val in enumerate(header_row0):
                if str(val).strip() == '個人業務':
                    duty_am_col = col_idx
                    duty_pm_col = col_idx + 1
                    break
        
        if duty_am_col is None:
            logging.info("個人業務列が見つかりませんでした")
            return instance
        
        # Parse data rows (starting from row 2)
        import re
        for row_idx in range(2, len(calendar_df)):
            row = calendar_df.iloc[row_idx]
            date_str = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
            
            # Parse date like "12/21" or "1/5"
            date_match = re.match(r'(\d+)/(\d+)', date_str)
            if not date_match:
                continue
            
            month = int(date_match.group(1))
            day = int(date_match.group(2))
            
            # Determine year based on month and target period
            if month >= target_month:
                year = target_year
            else:
                year = target_year + 1
            
            try:
                date_iso = f"{year:04d}-{month:02d}-{day:02d}"
            except:
                continue
            
            # Parse AM column
            am_val = str(row.iloc[duty_am_col]).strip() if duty_am_col < len(row) and pd.notna(row.iloc[duty_am_col]) else ''
            pm_val = str(row.iloc[duty_pm_col]).strip() if duty_pm_col < len(row) and pd.notna(row.iloc[duty_pm_col]) else ''
            
            # Parse entries like "メンバー名(業務名), メンバー名2(業務名2)"
            def parse_duty_entries(text: str, is_am: bool):
                if not text or text == 'nan':
                    return
                
                # Split by comma
                entries = text.split(',')
                for entry in entries:
                    entry = entry.strip()
                    if not entry:
                        continue
                    
                    # Parse "メンバー名(業務名)" format
                    match = re.match(r'(.+?)\((.+?)\)', entry)
                    if match:
                        member_name = match.group(1).strip()
                        duty_name = match.group(2).strip()
                    else:
                        # Just member name without duty specification
                        member_name = entry
                        duty_name = '個人業務'
                    
                    # Set time based on AM/PM
                    if is_am:
                        start_time = 830
                        end_time = 1230
                    else:
                        start_time = 1230
                        end_time = 1700
                    
                    instance.assignments.append({
                        'date': date_iso,
                        'start_time': start_time,
                        'end_time': end_time,
                        'duty': duty_name,
                        'name': member_name,
                        'day': day
                    })
            
            parse_duty_entries(am_val, is_am=True)
            parse_duty_entries(pm_val, is_am=False)
        
        logging.info(f"カレンダーシートから{len(instance.assignments)}件の個人業務を抽出しました")
        return instance

    def _load_duty_template(self):
        """Load and process duty template into individual time slot assignments"""
        df = pd.read_excel(self.duty_template_path)

        for _, row in df.iterrows():
            duty = str(row['duty'])
            from_day = int(row['from'])
            till_day = int(row['till'])
            start_time = int(row['start'])
            end_time = int(row['end'])
            name = str(row['name'])

            for day in range(from_day, till_day + 1):
                try:
                    date_iso = parse_date_token(str(day), self.target_year, self.target_month)
                    date_obj = dt.datetime.strptime(date_iso, '%Y-%m-%d')
                    weekday = date_obj.weekday()
                    is_holiday = jpholiday.is_holiday(date_obj.date())
                    
                    # Skip Sunday assignments
                    if weekday == 6:
                        continue
                    
                    # Skip holiday assignments
                    if is_holiday:
                        continue
                    
                    if weekday == 5 and start_time >= 1200:
                        continue
                    
                    self.assignments.append({
                        'date': date_iso,
                        'start_time': start_time,
                        'end_time': end_time,
                        'duty': duty,
                        'name': name,
                        'day': day
                    })
                except ValueError:
                    continue

    def get_assignments_for_date(self, date_iso: str):
        """Get all duty assignments for a specific date"""
        return [a for a in self.assignments if a['date'] == date_iso]

    def get_assigned_members_for_date_time(self, date_iso: str, start_min: int, end_min: int):
        """Get members already assigned during a specific time period"""
        assigned_members = set()
        for assignment in self.get_assignments_for_date(date_iso):
            duty_start = hhmm_to_minutes(assignment['start_time'])
            duty_end = hhmm_to_minutes(assignment['end_time'])

            if not (end_min <= duty_start or duty_end <= start_min):
                assigned_members.add(assignment['name'])

        return assigned_members


class Optimizer:
    # PuLP status codes
    OPTIMAL = pulp.LpStatusOptimal
    FEASIBLE = pulp.LpStatusOptimal  # PuLP doesn't distinguish FEASIBLE from OPTIMAL
    INFEASIBLE = pulp.LpStatusInfeasible
    
    def __init__(self, setting: Setting, work: WorkData, jobA: Optional[JobData], jobB: Optional[JobData], duty_template: Optional[DutyTemplate] = None):
        self.setting = setting
        self.work = work
        self.jobA = jobA
        self.jobB = jobB
        self.duty_template = duty_template
        self.target_year = work.target_year
        self.target_month = work.target_month
        self.model = None  # Will be created in build_and_solve
        self.solution_values = {}  # Store solution values after solving
        np.random.seed(setting.config_map.get("random_seed", 42))

        self.round10 = True
        self.min_interval = setting.config_map.get("min_interval", 0)
        self.workday_end_hhmm = setting.config_map.get("workday_end", 1800)

        self.x_vars: Dict[Tuple[int, str], pulp.LpVariable] = {}
        self.y_vars: Dict[Tuple[int, str], pulp.LpVariable] = {}

        self.infeasible_reasons: List[str] = []

        self.violations: List[str] = []

    def _work_date_to_iso(self, work_date) -> str:
        """Convert work data date to ISO format for avail_by_date access"""
        return parse_date_token(str(work_date), self.target_year, self.target_month)

    def build_and_solve(self, max_iterations=10):
        """Build and solve the optimization model multiple times to find the best solution for workload leveling.
        
        First attempts to solve without reserve staff. If infeasible, includes reserve staff and retries.
        """
        result = self._build_and_solve_internal(max_iterations, include_reserve=False)
        
        if result == Optimizer.INFEASIBLE:
            logging.info("通常スタッフのみでは不能解。ダミースタッフを含めて再試行します。")
            self.infeasible_reasons = []
            self.violations = []
            result = self._build_and_solve_internal(max_iterations, include_reserve=True)
            if result == Optimizer.OPTIMAL:
                self.violations.append("注意：ダミースタッフが割当されました")
                self.dummy_staff_used = True
            else:
                self.dummy_staff_used = False
        else:
            self.dummy_staff_used = False
        
        return result

    def _build_and_solve_internal(self, max_iterations, include_reserve=False):
        """Internal method to build and solve the optimization model.
        
        Args:
            max_iterations: Number of iterations to try
            include_reserve: If False, exclude staff with load='reserve' from optimization
        """
        best_objective = float('inf')
        best_solution = None

        for iteration in range(max_iterations):
            self.model = pulp.LpProblem(f"ShiftOptimization_{iteration}", pulp.LpMinimize)
            self.violations = []
            self.solution_values = {}

            W = list(self.work.work["id"].astype(int))
            w_rows = {int(r["id"]): r for _, r in self.work.work.iterrows()}

            A_list_full = list(self.jobA.members["aka"]) if self.jobA is not None else []
            B_list_full = list(self.jobB.members["aka"]) if self.jobB is not None else []
            
            if include_reserve:
                A_list = A_list_full
                B_list = B_list_full
            else:
                A_list = [a for a in A_list_full if self.jobA is None or self.jobA.load.get(a, "normal") != "reserve"]
                B_list = [b for b in B_list_full if self.jobB is None or self.jobB.load.get(b, "normal") != "reserve"]

            name_to_aka_A = {self.jobA.aka_to_name.get(a, a): a for a in A_list} if self.jobA else {}
            name_to_aka_B = {self.jobB.aka_to_name.get(b, b): b for b in B_list} if self.jobB else {}

            self.x_vars = {}
            self.y_vars = {}

            for w in W:
                wr = w_rows[w]
                date_iso = self._work_date_to_iso(wr["date"])
                start_min = int(wr["start_min"])
                end_min = int(wr["end_min"])

                assigned_members = set()
                if self.duty_template:
                    assigned_members = self.duty_template.get_assigned_members_for_date_time(date_iso, start_min, end_min)

                for a in A_list:
                    if self.jobA is not None:
                        member_name = self.jobA.aka_to_name.get(a, a)
                        if member_name not in assigned_members:
                            self.x_vars[(w, a)] = pulp.LpVariable(f"x_w{w}_a{a}", cat=pulp.LpBinary)

                for b in B_list:
                    if self.jobB is not None:
                        member_name = self.jobB.aka_to_name.get(b, b)
                        if member_name not in assigned_members:
                            self.y_vars[(w, b)] = pulp.LpVariable(f"y_w{w}_b{b}", cat=pulp.LpBinary)

            objective_terms = []
            
            pen_number = self.setting.penalty_map.get("number", {"type": "hard"})
            relax_number = any(e["penalty"] == "number" and e["vector"] == "relax" for e in self.setting.exceptions)
            number_weight = int(self.setting.penalty_map.get("number", {"value": 1}).get("value", 1))

            for w in W:
                need_A = int(w_rows[w].get("need_A", 0))
                need_B = int(w_rows[w].get("need_B", 0))
                if self.jobA is not None:
                    available_vars_A = [self.x_vars[(w, a)] for a in A_list if (w, a) in self.x_vars]
                    sumA = pulp.lpSum(available_vars_A)
                    if pen_number.get("type", "hard") == "hard" and not relax_number:
                        self.model += sumA == need_A, f"need_A_w{w}"
                    else:
                        diffA = pulp.LpVariable(f"diffA_w{w}", lowBound=0, upBound=len(A_list), cat=pulp.LpInteger)
                        self.model += sumA - need_A <= diffA, f"diffA_upper_w{w}"
                        self.model += need_A - sumA <= diffA, f"diffA_lower_w{w}"
                        objective_terms.append(diffA * number_weight)
                if self.jobB is not None:
                    available_vars_B = [self.y_vars[(w, b)] for b in B_list if (w, b) in self.y_vars]
                    sumB = pulp.lpSum(available_vars_B)
                    if pen_number.get("type", "hard") == "hard" and not relax_number:
                        self.model += sumB == need_B, f"need_B_w{w}"
                    else:
                        diffB = pulp.LpVariable(f"diffB_w{w}", lowBound=0, upBound=len(B_list), cat=pulp.LpInteger)
                        self.model += sumB - need_B <= diffB, f"diffB_upper_w{w}"
                        self.model += need_B - sumB <= diffB, f"diffB_lower_w{w}"
                        objective_terms.append(diffB * number_weight)
            date_to_need_A: Dict[str, int] = {}
            date_to_need_B: Dict[str, int] = {}
            for w in W:
                d = self._work_date_to_iso(w_rows[w]["date"])
                date_to_need_A[d] = date_to_need_A.get(d, 0) + int(w_rows[w].get("need_A", 0))
                date_to_need_B[d] = date_to_need_B.get(d, 0) + int(w_rows[w].get("need_B", 0))


            for w in W:
                date = self._work_date_to_iso(w_rows[w]["date"])
                if self.jobA is not None:
                    availA = self.jobA.avail_by_date.get(date, set())
                    for a in A_list:
                        if a not in availA and (w, a) in self.x_vars:
                            self.model += self.x_vars[(w, a)] == 0, f"unavail_A_w{w}_a{a}"
                if self.jobB is not None:
                    availB = self.jobB.avail_by_date.get(date, set())
                    for b in B_list:
                        if b not in availB and (w, b) in self.y_vars:
                            self.model += self.y_vars[(w, b)] == 0, f"unavail_B_w{w}_b{b}"

            noon_minutes = hhmm_to_minutes(self.setting.config_map.get("noon", 1230))
            
            for w in W:
                wr = w_rows[w]
                date = self._work_date_to_iso(wr["date"])
                start_min = int(wr["start_min"])
                end_min = int(wr["end_min"])
                
                if self.jobA is not None and self.jobA.leave_type_by_date:
                    leave_types_A = self.jobA.leave_type_by_date.get(date, {})
                    for a in A_list:
                        if (w, a) not in self.x_vars:
                            continue
                        leave_type = leave_types_A.get(a, '')
                        if leave_type == 'A' and start_min < noon_minutes:
                            self.model += self.x_vars[(w, a)] == 0, f"leave_A_am_w{w}_a{a}"
                        elif leave_type == 'P' and end_min > noon_minutes:
                            self.model += self.x_vars[(w, a)] == 0, f"leave_A_pm_w{w}_a{a}"
                
                if self.jobB is not None and self.jobB.leave_type_by_date:
                    leave_types_B = self.jobB.leave_type_by_date.get(date, {})
                    for b in B_list:
                        if (w, b) not in self.y_vars:
                            continue
                        leave_type = leave_types_B.get(b, '')
                        if leave_type == 'A' and start_min < noon_minutes:
                            self.model += self.y_vars[(w, b)] == 0, f"leave_B_am_w{w}_b{b}"
                        elif leave_type == 'P' and end_min > noon_minutes:
                            self.model += self.y_vars[(w, b)] == 0, f"leave_B_pm_w{w}_b{b}"

            for w in W:
                date = self._work_date_to_iso(w_rows[w]["date"])
                if self.jobA is not None:
                    availA = self.jobA.avail_by_date.get(date, set())
                    if int(w_rows[w].get("need_A", 0)) > len(availA):
                        self.infeasible_reasons.append(f"{date}: Aの必要数{int(w_rows[w].get('need_A',0))} > 当日利用可能{len(availA)}")
                if self.jobB is not None:
                    availB = self.jobB.avail_by_date.get(date, set())
                    if int(w_rows[w].get("need_B", 0)) > len(availB):
                        self.infeasible_reasons.append(f"{date}: Bの必要数{int(w_rows[w].get('need_B',0))} > 当日利用可能{len(availB)}")

            def overlap(wi, wj):
                si = int(w_rows[wi]["start_min"])
                ei = int(w_rows[wi]["end_min"])
                sj = int(w_rows[wj]["start_min"])
                ej = int(w_rows[wj]["end_min"])
                return not (ei + self.min_interval <= sj or ej + self.min_interval <= si)

            if self.jobA is not None:
                for a in A_list:
                    for i in range(len(W)):
                        for j in range(i + 1, len(W)):
                            wi, wj = W[i], W[j]
                            if self._work_date_to_iso(w_rows[wi]["date"]) != self._work_date_to_iso(w_rows[wj]["date"]):
                                continue
                            if overlap(wi, wj) and (wi, a) in self.x_vars and (wj, a) in self.x_vars:
                                self.model += self.x_vars[(wi, a)] + self.x_vars[(wj, a)] <= 1, f"overlap_A_{a}_w{wi}_w{wj}"
            if self.jobB is not None:
                for b in B_list:
                    for i in range(len(W)):
                        for j in range(i + 1, len(W)):
                            wi, wj = W[i], W[j]
                            if self._work_date_to_iso(w_rows[wi]["date"]) != self._work_date_to_iso(w_rows[wj]["date"]):
                                continue
                            if overlap(wi, wj) and (wi, b) in self.y_vars and (wj, b) in self.y_vars:
                                self.model += self.y_vars[(wi, b)] + self.y_vars[(wj, b)] <= 1, f"overlap_B_{b}_w{wi}_w{wj}"

            pen_senior = self.setting.penalty_map.get("senior", {"type": "hard"})
            ignore_senior_A = any(e["penalty"] == "senior" and e["vector"] == "ignore" and e.get("scope", "*") in ("A", "*") for e in self.setting.exceptions)
            ignore_senior_B = any(e["penalty"] == "senior" and e["vector"] == "ignore" and e.get("scope", "*") in ("B", "*") for e in self.setting.exceptions)

            if self.jobA is not None and not ignore_senior_A:
                seniorA = set(self.jobA.senior_set)
                for w in W:
                    need_A = int(w_rows[w].get("need_A", 0))
                    if need_A > 0 and seniorA:
                        available_senior_vars_A = [self.x_vars[(w, a)] for a in A_list if a in seniorA and (w, a) in self.x_vars]
                        if available_senior_vars_A:
                            sumSenior = pulp.lpSum(available_senior_vars_A)
                            if pen_senior.get("type", "hard") == "hard":
                                self.model += sumSenior >= 1, f"senior_A_w{w}"
                            else:
                                lack = pulp.LpVariable(f"lack_seniorA_w{w}", lowBound=0, upBound=1, cat=pulp.LpInteger)
                                self.model += sumSenior + lack >= 1, f"senior_A_soft_w{w}"
                                val = int(self.setting.penalty_map.get("senior", {"value": 5}).get("value", 5))
                                objective_terms.append(lack * val)

            if self.jobB is not None and not ignore_senior_B:
                seniorB = set(self.jobB.senior_set)
                for w in W:
                    need_B = int(w_rows[w].get("need_B", 0))
                    if need_B > 0 and seniorB:
                        available_senior_vars_B = [self.y_vars[(w, b)] for b in B_list if b in seniorB and (w, b) in self.y_vars]
                        if available_senior_vars_B:
                            sumSenior = pulp.lpSum(available_senior_vars_B)
                            if pen_senior.get("type", "hard") == "hard":
                                self.model += sumSenior >= 1, f"senior_B_w{w}"
                            else:
                                lack = pulp.LpVariable(f"lack_seniorB_w{w}", lowBound=0, upBound=1, cat=pulp.LpInteger)
                                self.model += sumSenior + lack >= 1, f"senior_B_soft_w{w}"
                                val = int(self.setting.penalty_map.get("senior", {"value": 5}).get("value", 5))
                                objective_terms.append(lack * val)

            if self.jobA is not None:
                for d, need in date_to_need_A.items():
                    if need > len(self.jobA.avail_by_date.get(d, set())):
                        self.infeasible_reasons.append(f"{d}: Aの総必要数{need} > 当日出勤可能{len(self.jobA.avail_by_date.get(d, set()))}")
            if self.jobB is not None:
                for d, need in date_to_need_B.items():
                    if need > len(self.jobB.avail_by_date.get(d, set())):
                        self.infeasible_reasons.append(f"{d}: Bの総必要数{need} > 当日出勤可能{len(self.jobB.avail_by_date.get(d, set()))}")

            prefer_val = float(self.setting.penalty_map.get("prefer", {"value": 1}).get("value", 1.0))
            for w in W:
                wr = w_rows[w]
                room = str(wr.get("room"))
                dept = str(wr.get("dept"))

                if self.jobA is not None:
                    for a in A_list:
                        row = self.jobA.members[self.jobA.members["aka"] == a].iloc[0]
                        m_pref = row["prefer_tokens"] if "prefer_tokens" in row else set()
                        reward = 0
                        if room in m_pref or dept in m_pref or f"W{w}" in m_pref:
                            reward += prefer_val
                        if reward > 0 and (w, a) in self.x_vars:
                            objective_terms.append(-reward * self.x_vars[(w, a)])
                if self.jobB is not None:
                    for b in B_list:
                        row = self.jobB.members[self.jobB.members["aka"] == b].iloc[0]
                        m_pref = row["prefer_tokens"] if "prefer_tokens" in row else set()
                        reward = 0
                        if room in m_pref or dept in m_pref or f"W{w}" in m_pref:
                            reward += prefer_val
                        if reward > 0 and (w, b) in self.y_vars:
                            objective_terms.append(-reward * self.y_vars[(w, b)])

            pen_avoid = self.setting.penalty_map.get("avoid", {"type": "hard"})
            if pen_avoid.get("type", "hard") == "hard":
                for w in W:
                    wr = w_rows[w]
                    room = str(wr.get("room"))
                    dept = str(wr.get("dept"))

                    if self.jobA is not None:
                        for a in A_list:
                            row = self.jobA.members[self.jobA.members["aka"] == a].iloc[0]
                            m_avoid = row["avoid_tokens"] if "avoid_tokens" in row else set()
                            if (room in m_avoid or dept in m_avoid or f"W{w}" in m_avoid) and (w, a) in self.x_vars:
                                self.model += self.x_vars[(w, a)] == 0, f"avoid_A_w{w}_a{a}"
                    
                    if self.jobB is not None:
                        for b in B_list:
                            row = self.jobB.members[self.jobB.members["aka"] == b].iloc[0]
                            m_avoid = row["avoid_tokens"] if "avoid_tokens" in row else set()
                            if (room in m_avoid or dept in m_avoid or f"W{w}" in m_avoid) and (w, b) in self.y_vars:
                                self.model += self.y_vars[(w, b)] == 0, f"avoid_B_w{w}_b{b}"

            for w in W:
                wr = w_rows[w]
                room = str(wr.get("room"))
                work_avoid = wr.get("avoid_tokens", set())
                if room in work_avoid:
                    if self.jobA is not None:
                        for a in A_list:
                            if (w, a) in self.x_vars:
                                self.model += self.x_vars[(w, a)] == 0, f"work_avoid_A_w{w}_a{a}"
                    if self.jobB is not None:
                        for b in B_list:
                            if (w, b) in self.y_vars:
                                self.model += self.y_vars[(w, b)] == 0, f"work_avoid_B_w{w}_b{b}"

            objective_terms.extend(self._add_workload_leveling_constraints(W, w_rows, A_list, B_list))

            order_penalty = float(self.setting.penalty_map.get("order", {"value": 100}).get("value", 100.0))
            if order_penalty > 0:
                for w in W:
                    wr = w_rows[w]
                    dept = str(wr.get("dept"))

                    if self.jobA is not None:
                        for a in A_list:
                            if (w, a) not in self.x_vars:
                                continue
                            score = self.jobA.get_order_score(dept, a)
                            if score is None:
                                continue
                            coeff = int(score * order_penalty)
                            if coeff > 0:
                                objective_terms.append(self.x_vars[(w, a)] * coeff)

                    if self.jobB is not None:
                        for b in B_list:
                            if (w, b) not in self.y_vars:
                                continue
                            score = self.jobB.get_order_score(dept, b)
                            if score is None:
                                continue
                            coeff = int(score * order_penalty)
                            if coeff > 0:
                                objective_terms.append(self.y_vars[(w, b)] * coeff)

            after_duty_penalty = float(self.setting.penalty_map.get("after_duty", {"value": 10}).get("value", 10.0))
            if after_duty_penalty > 0:
                for w in W:
                    wr = w_rows[w]
                    date = self._work_date_to_iso(wr["date"])
                    
                    if self.jobA is not None and self.jobA.leave_type_by_date:
                        leave_types_A = self.jobA.leave_type_by_date.get(date, {})
                        for a in A_list:
                            if (w, a) not in self.x_vars:
                                continue
                            leave_type = leave_types_A.get(a, '')
                            if leave_type == 'D':
                                coeff = int(after_duty_penalty)
                                objective_terms.append(self.x_vars[(w, a)] * coeff)
                    
                    if self.jobB is not None and self.jobB.leave_type_by_date:
                        leave_types_B = self.jobB.leave_type_by_date.get(date, {})
                        for b in B_list:
                            if (w, b) not in self.y_vars:
                                continue
                            leave_type = leave_types_B.get(b, '')
                            if leave_type == 'D':
                                coeff = int(after_duty_penalty)
                                objective_terms.append(self.y_vars[(w, b)] * coeff)

            duty_assignment_priority_penalty = float(self.setting.penalty_map.get("duty_priority", {"value": 50}).get("value", 50.0))
            if duty_assignment_priority_penalty > 0 and self.duty_template:
                def has_pm_duty(member_name: str, date_iso: str) -> bool:
                    """Check if member has a PM personal duty on the given date"""
                    for assignment in self.duty_template.assignments:
                        if assignment['name'] == member_name and assignment['date'] == date_iso:
                            if assignment['start_time'] >= self.setting.config_map.get('noon', 1230):
                                return True
                    return False

                def has_am_duty(member_name: str, date_iso: str) -> bool:
                    """Check if member has an AM personal duty on the given date"""
                    for assignment in self.duty_template.assignments:
                        if assignment['name'] == member_name and assignment['date'] == date_iso:
                            if assignment['end_time'] <= self.setting.config_map.get('noon', 1230):
                                return True
                    return False

                def get_duty_priority_score(member_name: str, date_iso: str, duty_priority_list: List[str]) -> float:
                    """Get priority ordering score for a member based on their personal duties.
                    
                    Returns:
                        0.0 if no personal duty
                        0.1 to 1.0 based on duty priority list position (lower = higher priority)
                        1.0 if duty not in priority list
                    """
                    member_duties = []
                    for assignment in self.duty_template.assignments:
                        if assignment['name'] == member_name and assignment['date'] == date_iso:
                            member_duties.append(assignment['duty'])
                    
                    if not member_duties:
                        return 0.0
                    
                    best_score = 1.0
                    for duty in member_duties:
                        duty_str = str(duty).strip()
                        duty_str_normalized = duty_str.replace('\n', '').replace('\r', '')
                        
                        for idx, priority_duty in enumerate(duty_priority_list):
                            priority_duty_normalized = priority_duty.replace('\n', '').replace('\r', '')
                            
                            if priority_duty_normalized == "委員会":
                                if "委員会" in duty_str_normalized:
                                    score = (idx + 1) / (len(duty_priority_list) + 1)
                                    best_score = min(best_score, score)
                                    break
                            elif duty_str_normalized == priority_duty_normalized or duty_str == priority_duty:
                                score = (idx + 1) / (len(duty_priority_list) + 1)
                                best_score = min(best_score, score)
                                break
                    
                    return best_score
                
                for w in W:
                    wr = w_rows[w]
                    date = self._work_date_to_iso(wr["date"])
                    start_min = int(wr["start_min"])
                    end_min = int(wr["end_min"])
                    
                    is_am_work = start_min < noon_minutes
                    work_ends_by_noon = end_min <= noon_minutes
                    
                    if self.jobA is not None:
                        leave_types_A = self.jobA.leave_type_by_date.get(date, {}) if self.jobA.leave_type_by_date else {}
                        duty_priority_list_A = self.setting.duty_priority_A if hasattr(self.setting, 'duty_priority_A') else []
                        
                        for a in A_list:
                            if (w, a) not in self.x_vars:
                                continue
                            
                            member_name = self.jobA.aka_to_name.get(a, a)
                            leave_type = leave_types_A.get(a, '')
                            member_has_pm_duty = has_pm_duty(member_name, date)
                            member_has_am_duty = has_am_duty(member_name, date)
                            
                            if work_ends_by_noon and member_has_pm_duty:
                                # REWARD: PM-duty members should be assigned to AM work
                                duty_score = get_duty_priority_score(member_name, date, duty_priority_list_A) if duty_priority_list_A else 0.5
                                reward = (1.0 - duty_score * 0.5) * duty_assignment_priority_penalty
                                coeff = int(reward)
                                if coeff > 0:
                                    objective_terms.append(self.x_vars[(w, a)] * (-coeff))
                            elif not is_am_work and member_has_am_duty:
                                # REWARD: AM-duty members should be assigned to PM work
                                reward = 0.5 * duty_assignment_priority_penalty
                                coeff = int(reward)
                                if coeff > 0:
                                    objective_terms.append(self.x_vars[(w, a)] * (-coeff))
                            else:
                                base_penalty = 0.0
                                if leave_type == 'P' and is_am_work:
                                    base_penalty = 0.1
                                elif leave_type == 'A' and not is_am_work:
                                    base_penalty = 0.1
                                elif leave_type not in ('', 'D'):
                                    base_penalty = 0.2
                                
                                if base_penalty > 0:
                                    coeff = int(base_penalty * duty_assignment_priority_penalty)
                                    if coeff > 0:
                                        objective_terms.append(self.x_vars[(w, a)] * coeff)
                    
                    if self.jobB is not None:
                        leave_types_B = self.jobB.leave_type_by_date.get(date, {}) if self.jobB.leave_type_by_date else {}
                        duty_priority_list_B = self.setting.duty_priority_B if hasattr(self.setting, 'duty_priority_B') else []
                        
                        for b in B_list:
                            if (w, b) not in self.y_vars:
                                continue
                            
                            member_name = self.jobB.aka_to_name.get(b, b)
                            leave_type = leave_types_B.get(b, '')
                            member_has_pm_duty = has_pm_duty(member_name, date)
                            member_has_am_duty = has_am_duty(member_name, date)
                            
                            if work_ends_by_noon and member_has_pm_duty:
                                duty_score = get_duty_priority_score(member_name, date, duty_priority_list_B) if duty_priority_list_B else 0.5
                                reward = (1.0 - duty_score * 0.5) * duty_assignment_priority_penalty
                                coeff = int(reward)
                                if coeff > 0:
                                    objective_terms.append(self.y_vars[(w, b)] * (-coeff))
                            elif not is_am_work and member_has_am_duty:
                                reward = 0.5 * duty_assignment_priority_penalty
                                coeff = int(reward)
                                if coeff > 0:
                                    objective_terms.append(self.y_vars[(w, b)] * (-coeff))
                            else:
                                base_penalty = 0.0
                                if leave_type == 'P' and is_am_work:
                                    base_penalty = 0.1
                                elif leave_type == 'A' and not is_am_work:
                                    base_penalty = 0.1
                                elif leave_type not in ('', 'D'):
                                    base_penalty = 0.2
                                
                                if base_penalty > 0:
                                    coeff = int(base_penalty * duty_assignment_priority_penalty)
                                    if coeff > 0:
                                        objective_terms.append(self.y_vars[(w, b)] * coeff)

            dummy_penalty_value = float(self.setting.penalty_map.get("dummy", {"value": 10}).get("value", 10.0))
            if include_reserve and dummy_penalty_value > 0:
                reserve_akas_A = set()
                reserve_akas_B = set()
                if self.jobA is not None:
                    reserve_akas_A = {a for a in A_list if self.jobA.load.get(a, "normal") == "reserve"}
                if self.jobB is not None:
                    reserve_akas_B = {b for b in B_list if self.jobB.load.get(b, "normal") == "reserve"}

                real_akas_A = set(A_list) - reserve_akas_A
                real_akas_B = set(B_list) - reserve_akas_B

                # Build avoid lookup per member
                avoid_tokens_A = {}
                if self.jobA is not None:
                    for _, row in self.jobA.members.iterrows():
                        aka = str(row["aka"])
                        avoid_tokens_A[aka] = row.get("avoid_tokens", set()) if "avoid_tokens" in row.index else set()
                avoid_tokens_B = {}
                if self.jobB is not None:
                    for _, row in self.jobB.members.iterrows():
                        aka = str(row["aka"])
                        avoid_tokens_B[aka] = row.get("avoid_tokens", set()) if "avoid_tokens" in row.index else set()

                coeff_base = int(dummy_penalty_value)
                # Much higher penalty when real non-avoid candidates exist
                coeff_strong = coeff_base * 100
                for w in W:
                    wr = w_rows[w]
                    room = str(wr.get("room", ""))
                    dept = str(wr.get("dept", ""))

                    # Count real candidates (has var and no avoid)
                    has_real_candidate_A = False
                    for a in real_akas_A:
                        if (w, a) in self.x_vars:
                            m_avoid = avoid_tokens_A.get(a, set())
                            if room not in m_avoid and dept not in m_avoid and f"W{w}" not in m_avoid:
                                has_real_candidate_A = True
                                break

                    has_real_candidate_B = False
                    for b in real_akas_B:
                        if (w, b) in self.y_vars:
                            m_avoid = avoid_tokens_B.get(b, set())
                            if room not in m_avoid and dept not in m_avoid and f"W{w}" not in m_avoid:
                                has_real_candidate_B = True
                                break

                    # Apply strong penalty when real candidates exist, base penalty otherwise
                    for a in reserve_akas_A:
                        if (w, a) in self.x_vars:
                            c = coeff_strong if has_real_candidate_A else coeff_base
                            if c > 0:
                                objective_terms.append(self.x_vars[(w, a)] * c)
                    for b in reserve_akas_B:
                        if (w, b) in self.y_vars:
                            c = coeff_strong if has_real_candidate_B else coeff_base
                            if c > 0:
                                objective_terms.append(self.y_vars[(w, b)] * c)

            if objective_terms:
                self.model += pulp.lpSum(objective_terms), "Objective"

            self.model.solve(pulp.PULP_CBC_CMD(msg=0, timeLimit=15))
            res = self.model.status

            if res == pulp.LpStatusOptimal:
                current_objective = pulp.value(self.model.objective) if self.model.objective else 0
                if current_objective < best_objective:
                    best_objective = current_objective
                    # Store solution values
                    solution_values = {}
                    for key, var in self.x_vars.items():
                        solution_values[('x', key)] = pulp.value(var)
                    for key, var in self.y_vars.items():
                        solution_values[('y', key)] = pulp.value(var)
                    best_solution = (res, self.x_vars.copy(), self.y_vars.copy(), self.violations.copy(), solution_values)
            elif self.infeasible_reasons:
                self.violations.extend(self.infeasible_reasons)

        if best_solution:
            res, self.x_vars, self.y_vars, self.violations, self.solution_values = best_solution
            return Optimizer.OPTIMAL
        else:
            return Optimizer.INFEASIBLE

    def _add_workload_leveling_constraints(self, W, w_rows, A_list, B_list):
        """Add constraints to level workload within same job type and load groups.

        Uses penalties from setting:
          - diff: intra-group leveling (thres=allowed % difference, value=penalty weight)
          - more: cross-group constraint for load='more' staff (+thres% vs normal avg)
          - less: cross-group constraint for load='less' staff (-thres% vs normal avg)

        Reserve staff are excluded from workload leveling.
        """
        diff_cfg = self.setting.penalty_map.get("diff", {"value": 1, "thres": 0})
        diff_penalty = float(diff_cfg.get("value", 1.0))
        diff_thres_pct = int(diff_cfg.get("thres", 0))

        more_cfg = self.setting.penalty_map.get("more", {"value": 1, "thres": 20})
        more_penalty = float(more_cfg.get("value", 1.0))
        more_thres_pct = int(more_cfg.get("thres", 20))

        less_cfg = self.setting.penalty_map.get("less", {"value": 1, "thres": 20})
        less_penalty = float(less_cfg.get("value", 1.0))
        less_thres_pct = int(less_cfg.get("thres", 20))

        penalty_terms = []

        def _build_load_groups(members_list, job):
            groups = {}
            for m in members_list:
                lt = job.load.get(m, "normal")
                if lt == "reserve":
                    continue
                groups.setdefault(lt, []).append(m)
            return groups

        def _work_expr(var_dict, member, work_ids):
            return pulp.lpSum([var_dict[(w, member)] for w in work_ids if (w, member) in var_dict])

        def _add_intra_group(load_groups, var_dict, role_prefix, total_demand):
            """A: intra-group leveling with diff thres."""
            for load_type, members in load_groups.items():
                if len(members) <= 1:
                    continue
                avg_per_member = total_demand / len(members) if len(members) > 0 else 1
                threshold_count = max(1, int(avg_per_member * diff_thres_pct / 100)) if diff_thres_pct > 0 else 0
                for i, m1 in enumerate(members):
                    for m2 in members[i+1:]:
                        w1 = _work_expr(var_dict, m1, W)
                        w2 = _work_expr(var_dict, m2, W)
                        dv = pulp.LpVariable(f"diff_{role_prefix}_{m1}_{m2}", lowBound=0, upBound=len(W), cat=pulp.LpInteger)
                        self.model += w1 - w2 - threshold_count <= dv, f"diff_{role_prefix}_{m1}_{m2}_upper"
                        self.model += w2 - w1 - threshold_count <= dv, f"diff_{role_prefix}_{m1}_{m2}_lower"
                        penalty_terms.append(dv * diff_penalty)

        def _add_cross_group(load_groups, var_dict, role_prefix):
            """B: cross-group more/less constraints vs normal average."""
            normal_members = load_groups.get("normal", [])
            if not normal_members:
                return
            count_normal = len(normal_members)
            sum_normal = pulp.lpSum([_work_expr(var_dict, n, W) for n in normal_members])

            more_members = load_groups.get("more", [])
            if more_members and more_penalty > 0:
                ratio = 1.0 + more_thres_pct / 100.0
                for m in more_members:
                    work_m = _work_expr(var_dict, m, W)
                    # target = sum_normal / count_normal * ratio
                    # penalize shortfall: target - work_m > 0
                    sv = pulp.LpVariable(f"more_{role_prefix}_{m}", lowBound=0, upBound=len(W), cat=pulp.LpInteger)
                    self.model += sum_normal * ratio / count_normal - work_m <= sv, f"more_{role_prefix}_{m}_short"
                    penalty_terms.append(sv * more_penalty)

            less_members = load_groups.get("less", [])
            if less_members and less_penalty > 0:
                ratio = 1.0 - less_thres_pct / 100.0
                for m in less_members:
                    work_m = _work_expr(var_dict, m, W)
                    # target = sum_normal / count_normal * ratio
                    # penalize excess: work_m - target > 0
                    sv = pulp.LpVariable(f"less_{role_prefix}_{m}", lowBound=0, upBound=len(W), cat=pulp.LpInteger)
                    self.model += work_m - sum_normal * ratio / count_normal <= sv, f"less_{role_prefix}_{m}_over"
                    penalty_terms.append(sv * less_penalty)

        if self.jobA is not None:
            load_groups_A = _build_load_groups(A_list, self.jobA)
            total_demand_A = sum(int(w_rows[w].get("need_A", 0)) for w in W)
            _add_intra_group(load_groups_A, self.x_vars, "A", total_demand_A)
            _add_cross_group(load_groups_A, self.x_vars, "A")

        if self.jobB is not None:
            load_groups_B = _build_load_groups(B_list, self.jobB)
            total_demand_B = sum(int(w_rows[w].get("need_B", 0)) for w in W)
            _add_intra_group(load_groups_B, self.y_vars, "B", total_demand_B)
            _add_cross_group(load_groups_B, self.y_vars, "B")

        dept_penalty = float(self.setting.penalty_map.get("diff", {"value": 1}).get("value", 1.0))
        target_depts = ["D16", "D17"]

        for target_dept in target_depts:
            dept_work_ids = [w for w in W if w_rows[w].get("dept") == target_dept]
            if not dept_work_ids:
                continue

            dept_members_A = []
            dept_members_B = []

            if self.jobA is not None:
                for a in A_list:
                    available_for_dept = any((w, a) in self.x_vars for w in dept_work_ids)
                    if available_for_dept:
                        dept_members_A.append(a)

            if self.jobB is not None:
                for b in B_list:
                    available_for_dept = any((w, b) in self.y_vars for w in dept_work_ids)
                    if available_for_dept:
                        dept_members_B.append(b)

            if len(dept_members_A) > 1:
                for i, a1 in enumerate(dept_members_A):
                    for a2 in dept_members_A[i+1:]:
                        work1_dept = pulp.lpSum([
                            self.x_vars[(w, a1)] for w in dept_work_ids if (w, a1) in self.x_vars
                        ])
                        work2_dept = pulp.lpSum([
                            self.x_vars[(w, a2)] for w in dept_work_ids if (w, a2) in self.x_vars
                        ])

                        member1_name = self.jobA.aka_to_name.get(a1, a1) if self.jobA else a1
                        member2_name = self.jobA.aka_to_name.get(a2, a2) if self.jobA else a2
                        duty1_count = count_duty_assignments(member1_name)
                        duty2_count = count_duty_assignments(member2_name)

                        work1_total = work1_dept + duty1_count
                        work2_total = work2_dept + duty2_count

                        diff_var = pulp.LpVariable(f"diff_dept_{target_dept}_A_{a1}_{a2}", lowBound=0, upBound=len(dept_work_ids) + max(duty1_count, duty2_count), cat=pulp.LpInteger)
                        self.model += work1_total - work2_total <= diff_var, f"diff_dept_{target_dept}_A_{a1}_{a2}_upper"
                        self.model += work2_total - work1_total <= diff_var, f"diff_dept_{target_dept}_A_{a1}_{a2}_lower"
                        penalty_terms.append(diff_var * dept_penalty)

            if len(dept_members_B) > 1:
                for i, b1 in enumerate(dept_members_B):
                    for b2 in dept_members_B[i+1:]:
                        work1_dept = pulp.lpSum([
                            self.y_vars[(w, b1)] for w in dept_work_ids if (w, b1) in self.y_vars
                        ])
                        work2_dept = pulp.lpSum([
                            self.y_vars[(w, b2)] for w in dept_work_ids if (w, b2) in self.y_vars
                        ])

                        member1_name = self.jobB.aka_to_name.get(b1, b1) if self.jobB else b1
                        member2_name = self.jobB.aka_to_name.get(b2, b2) if self.jobB else b2
                        duty1_count = count_duty_assignments(member1_name)
                        duty2_count = count_duty_assignments(member2_name)

                        work1_total = work1_dept + duty1_count
                        work2_total = work2_dept + duty2_count

                        diff_var = pulp.LpVariable(f"diff_dept_{target_dept}_B_{b1}_{b2}", lowBound=0, upBound=len(dept_work_ids) + max(duty1_count, duty2_count), cat=pulp.LpInteger)
                        self.model += work1_total - work2_total <= diff_var, f"diff_dept_{target_dept}_B_{b1}_{b2}_upper"
                        self.model += work2_total - work1_total <= diff_var, f"diff_dept_{target_dept}_B_{b1}_{b2}_lower"
                        penalty_terms.append(diff_var * dept_penalty)

        return penalty_terms

    # ------------------------------------------------------------------
    # Post-processing: greedily swap dummy assignments to real staff
    # ------------------------------------------------------------------
    def post_process_reduce_dummy(self):
        """After solving, try to replace dummy (reserve) staff with real staff.

        For every work item assigned to a dummy, look for a real staff member
        who is available, has no avoid conflict, no time overlap with their
        current assignments or personal duties, and swap them in.

        Stores overrides in ``self._assignment_overrides`` which
        ``extract_output`` will consult.
        """
        self._assignment_overrides: Dict[tuple, int] = {}  # (w, aka) -> 0 or 1

        if self.jobA is None and self.jobB is None:
            return

        Wdf = self.work.work.copy()
        w_rows = {int(r["id"]): r for _, r in Wdf.iterrows()}
        W = list(Wdf["id"].astype(int))

        min_interval = self.min_interval

        def time_overlaps(s1, e1, s2, e2):
            return not (e1 + min_interval <= s2 or e2 + min_interval <= s1)

        def _get_duty_intervals(member_name, date_iso):
            intervals = []
            if self.duty_template:
                for a in self.duty_template.get_assignments_for_date(date_iso):
                    if a['name'] == member_name:
                        intervals.append((hhmm_to_minutes(a['start_time']),
                                          hhmm_to_minutes(a['end_time'])))
            return intervals

        def _solved_value(var_dict, key):
            """Return effective assignment value considering overrides."""
            if key in self._assignment_overrides:
                return self._assignment_overrides[key]
            var = var_dict.get(key)
            if var is None:
                return 0
            return 1 if pulp.value(var) == 1 else 0

        # Process each job type (A and B) independently
        for job, var_dict, role_label in [
            (self.jobA, self.x_vars, "A"),
            (self.jobB, self.y_vars, "B"),
        ]:
            if job is None:
                continue

            all_akas = list(job.members["aka"])
            reserve_akas = {a for a in all_akas if job.load.get(a, "normal") == "reserve"}
            real_akas = [a for a in all_akas if a not in reserve_akas]

            if not reserve_akas:
                continue

            # Build current assignments per staff: aka -> list of (work_id, start_min, end_min, date_iso)
            staff_assignments: Dict[str, List[tuple]] = {a: [] for a in all_akas}
            for w in W:
                wr = w_rows[w]
                for a in all_akas:
                    if _solved_value(var_dict, (w, a)) == 1:
                        staff_assignments[a].append((
                            w,
                            int(wr["start_min"]),
                            int(wr["end_min"]),
                            self._work_date_to_iso(wr["date"]),
                        ))

            # Collect dummy assignments
            dummy_assignments = []
            for w in W:
                wr = w_rows[w]
                date_iso = self._work_date_to_iso(wr["date"])
                for d_aka in reserve_akas:
                    if _solved_value(var_dict, (w, d_aka)) == 1:
                        dummy_assignments.append((w, d_aka, date_iso))

            if not dummy_assignments:
                continue

            # Build avoid tokens per member
            avoid_map = {}
            for _, row in job.members.iterrows():
                aka = str(row["aka"])
                avoid_map[aka] = row.get("avoid_tokens", set()) if "avoid_tokens" in row.index else set()

            swapped = 0
            for w, d_aka, date_iso in dummy_assignments:
                wr = w_rows[w]
                w_start = int(wr["start_min"])
                w_end = int(wr["end_min"])
                w_room = str(wr["room"])
                w_dept = str(wr["dept"])

                best_candidate = None
                best_load = float('inf')

                for r_aka in real_akas:
                    # 1. Available on this date?
                    if r_aka not in job.avail_by_date.get(date_iso, set()):
                        continue

                    # 2. Leave type check
                    leave = job.leave_type_by_date.get(date_iso, {}).get(r_aka, '')
                    if leave == '1':
                        continue
                    noon = hhmm_to_minutes(self.setting.config_map.get("noon", 1230))
                    if leave == 'A' and w_start < noon:
                        continue
                    if leave == 'P' and w_end > noon:
                        continue

                    # 3. Avoid constraint?
                    m_avoid = avoid_map.get(r_aka, set())
                    if w_room in m_avoid or w_dept in m_avoid or f"W{w}" in m_avoid:
                        continue

                    # 4. Time overlap with current assignments?
                    conflict = False
                    for (aw, a_start, a_end, a_date) in staff_assignments[r_aka]:
                        if a_date == date_iso and time_overlaps(w_start, w_end, a_start, a_end):
                            conflict = True
                            break
                    if conflict:
                        continue

                    # 5. Duty conflict?
                    member_name = job.aka_to_name.get(r_aka, r_aka)
                    duty_intervals = _get_duty_intervals(member_name, date_iso)
                    for d_start, d_end in duty_intervals:
                        if time_overlaps(w_start, w_end, d_start, d_end):
                            conflict = True
                            break
                    if conflict:
                        continue

                    # Pick candidate with fewest current assignments (balance load)
                    load = len(staff_assignments[r_aka])
                    if load < best_load:
                        best_load = load
                        best_candidate = r_aka

                if best_candidate is not None:
                    # Swap: remove dummy, add real
                    self._assignment_overrides[(w, d_aka)] = 0
                    self._assignment_overrides[(w, best_candidate)] = 1
                    # Update tracking
                    staff_assignments[d_aka] = [
                        x for x in staff_assignments[d_aka] if x[0] != w
                    ]
                    staff_assignments[best_candidate].append(
                        (w, w_start, w_end, date_iso)
                    )
                    swapped += 1

            if swapped > 0:
                logging.info(f"ポスト処理: {role_label}のダミー割当を{swapped}件、実スタッフに置換しました")

    def extract_output(self):
        rows_assign = []
        rows_available = []
        rows_stats = []
        rows_violation = self._extract_violations()

        overrides = getattr(self, '_assignment_overrides', {})

        Wdf = self.work.work.copy()
        for _, wr in Wdf.iterrows():
            w = int(wr["id"])
            date_iso = str(wr["date"])
            day = str(wr.get("day", ""))
            room = str(wr["room"])
            dept = str(wr["dept"])
            start_h = int(wr["start"])
            end_h = minutes_to_hhmm(int(wr["end_min"]), round10=True)
            start_min = int(wr["start_min"])
            end_min = int(wr["end_min"])

            assA_names: List[str] = []
            assB_names: List[str] = []

            if self.jobA is not None:
                for a in list(self.jobA.members["aka"]):
                    key = (w, a)
                    if key in overrides:
                        if overrides[key] == 1:
                            assA_names.append(self.jobA.aka_to_name.get(a, a))
                    else:
                        var = self.x_vars.get(key)
                        if var is not None and pulp.value(var) == 1:
                            assA_names.append(self.jobA.aka_to_name.get(a, a))
            if self.jobB is not None:
                for b in list(self.jobB.members["aka"]):
                    key = (w, b)
                    if key in overrides:
                        if overrides[key] == 1:
                            assB_names.append(self.jobB.aka_to_name.get(b, b))
                    else:
                        var = self.y_vars.get(key)
                        if var is not None and pulp.value(var) == 1:
                            assB_names.append(self.jobB.aka_to_name.get(b, b))

            rows_assign.append(
                {
                    "id": w,
                    "date": to_dd(date_iso),
                    "date_iso": date_iso,
                    "day": str(wr.get("day_original", "")),
                    "room": str(wr["room"]),
                    "dept": str(wr["dept"]),
                    "start": start_h,
                    "end": end_h,
                    "start_min": start_min,
                    "assign_A": ",".join(assA_names),
                    "assign_B": ",".join(assB_names),
                }
            )

        rows_available_A = []
        rows_available_B = []
        dates = sorted(set(self._work_date_to_iso(d) for d in Wdf["date"]))

        noon_hhmm = self.setting.config_map.get("noon", 1230)
        noon_boundary = hhmm_to_minutes(noon_hhmm)

        for d in dates:
            all_available_A = []
            all_available_B = []

            date_work = Wdf[Wdf["date"].apply(lambda x: self._work_date_to_iso(x)) == d]

            if self.jobA is not None:
                for a in list(self.jobA.members["aka"]):
                    member_name = self.jobA.aka_to_name.get(a, a)

                    assigned_work = []
                    for _, wr in date_work.iterrows():
                        key = (int(wr["id"]), a)
                        if key in overrides:
                            val = overrides[key]
                        else:
                            var = self.x_vars.get(key)
                            val = 1 if var is not None and pulp.value(var) == 1 else 0
                        if val == 1:
                            start_min = int(wr["start_min"])
                            end_min = int(wr["end_min"])
                            assigned_work.append((start_min, end_min))

                    if a in self.jobA.avail_by_date.get(d, set()):
                        duty_conflicts = []
                        if self.duty_template:
                            for assignment in self.duty_template.get_assignments_for_date(d):
                                if assignment['name'] == member_name:
                                    duty_start = hhmm_to_minutes(assignment['start_time'])
                                    duty_end = hhmm_to_minutes(assignment['end_time'])
                                    duty_conflicts.append((duty_start, duty_end))

                        if not assigned_work and not duty_conflicts:
                            all_available_A.append((member_name, ""))
                        elif assigned_work or duty_conflicts:
                            # Combine both optimization assignments and duty assignments
                            all_assignments = assigned_work + duty_conflicts
                            all_end_before_1230 = all(end_min <= noon_boundary for _, end_min in all_assignments)
                            all_start_after_1230 = all(start_min >= noon_boundary for start_min, _ in all_assignments)

                            if all_end_before_1230:
                                # Pattern 2: all work ends before 12:30
                                all_available_A.append((member_name, "PM"))
                            elif all_start_after_1230:
                                # Pattern 3: all work starts after 12:30
                                all_available_A.append((member_name, "AM"))

            if self.jobB is not None:
                for b in list(self.jobB.members["aka"]):
                    member_name = self.jobB.aka_to_name.get(b, b)

                    assigned_work = []
                    for _, wr in date_work.iterrows():
                        key = (int(wr["id"]), b)
                        if key in overrides:
                            val = overrides[key]
                        else:
                            var = self.y_vars.get(key)
                            val = 1 if var is not None and pulp.value(var) == 1 else 0
                        if val == 1:
                            start_min = int(wr["start_min"])
                            end_min = int(wr["end_min"])
                            assigned_work.append((start_min, end_min))

                    if b in self.jobB.avail_by_date.get(d, set()):
                        duty_conflicts = []
                        if self.duty_template:
                            for assignment in self.duty_template.get_assignments_for_date(d):
                                if assignment['name'] == member_name:
                                    duty_start = hhmm_to_minutes(assignment['start_time'])
                                    duty_end = hhmm_to_minutes(assignment['end_time'])
                                    duty_conflicts.append((duty_start, duty_end))

                        if not assigned_work and not duty_conflicts:
                            all_available_B.append((member_name, ""))
                        elif assigned_work or duty_conflicts:
                            # Combine both optimization assignments and duty assignments
                            all_assignments = assigned_work + duty_conflicts
                            all_end_before_1230 = all(end_min <= noon_boundary for _, end_min in all_assignments)
                            all_start_after_1230 = all(start_min >= noon_boundary for start_min, _ in all_assignments)

                            if all_end_before_1230:
                                # Pattern 2: all work ends before 12:30
                                all_available_B.append((member_name, "PM"))
                            elif all_start_after_1230:
                                # Pattern 3: all work starts after 12:30
                                all_available_B.append((member_name, "AM"))

            if all_available_A:
                fully_available_A = []
                pm_available_A = []
                am_available_A = []

                for member_name, suffix in all_available_A:
                    aka = None
                    for a in list(self.jobA.members["aka"]):
                        if self.jobA.aka_to_name.get(a, a) == member_name:
                            aka = a
                            break
                    if aka:
                        member_row = self.jobA.members[self.jobA.members["aka"] == aka].iloc[0]
                        prefer_count = count_prefer_entries(member_row.get("prefer_tokens", set()))

                        if suffix == "":
                            fully_available_A.append((prefer_count, member_name))
                        elif suffix == "PM":
                            pm_available_A.append((prefer_count, member_name))
                        elif suffix == "AM":
                            am_available_A.append((prefer_count, member_name))

                if fully_available_A or pm_available_A or am_available_A:
                    fully_available_A.sort(key=lambda x: x[0], reverse=False)
                    pm_available_A.sort(key=lambda x: x[0], reverse=False)
                    am_available_A.sort(key=lambda x: x[0], reverse=False)

                    fully_members_A = ",".join([member for _, member in fully_available_A]) if fully_available_A else ""
                    pm_members_A = ",".join([member for _, member in pm_available_A]) if pm_available_A else ""
                    am_members_A = ",".join([member for _, member in am_available_A]) if am_available_A else ""

                    rows_available_A.append({
                        "date": to_dd(d),
                        "day": get_japanese_weekday(d),
                        "full": fully_members_A,
                        "AM": am_members_A,
                        "PM": pm_members_A
                    })

            if all_available_B:
                fully_available_B = []
                pm_available_B = []
                am_available_B = []

                for member_name, suffix in all_available_B:
                    aka = None
                    for b in list(self.jobB.members["aka"]):
                        if self.jobB.aka_to_name.get(b, b) == member_name:
                            aka = b
                            break
                    if aka:
                        member_row = self.jobB.members[self.jobB.members["aka"] == aka].iloc[0]
                        prefer_count = count_prefer_entries(member_row.get("prefer_tokens", set()))

                        if suffix == "":
                            fully_available_B.append((prefer_count, member_name))
                        elif suffix == "PM":
                            pm_available_B.append((prefer_count, member_name))
                        elif suffix == "AM":
                            am_available_B.append((prefer_count, member_name))

                if fully_available_B or pm_available_B or am_available_B:
                    fully_available_B.sort(key=lambda x: x[0], reverse=False)
                    pm_available_B.sort(key=lambda x: x[0], reverse=False)
                    am_available_B.sort(key=lambda x: x[0], reverse=False)

                    fully_members_B = ",".join([member for _, member in fully_available_B]) if fully_available_B else ""
                    pm_members_B = ",".join([member for _, member in pm_available_B]) if pm_available_B else ""
                    am_members_B = ",".join([member for _, member in am_available_B]) if am_available_B else ""

                    rows_available_B.append({
                        "date": to_dd(d),
                        "day": get_japanese_weekday(d),
                        "full": fully_members_B,
                        "AM": am_members_B,
                        "PM": pm_members_B
                    })

        # assignシートを作成（WorkDataで既に日付→dept→start_time順にソート・id付与済み）
        df_assign = pd.DataFrame(rows_assign)
        if not df_assign.empty:
            df_assign = df_assign.drop(columns=['date_iso', 'start_min'])

        # duty_ids_by_memberを作成（statsシート用）
        duty_ids_by_member = {}
        if self.duty_template:
            for assignment in self.duty_template.assignments:
                member_name = assignment['name']
                duty_id = f"D{assignment['date'].replace('-', '')}_{assignment['duty']}"
                duty_date = assignment['date']
                duty_start = assignment['start_time']
                if member_name not in duty_ids_by_member:
                    duty_ids_by_member[member_name] = []
                duty_ids_by_member[member_name].append((duty_date, duty_start, duty_id))

        # statsシートを作成（WorkDataで既にソート・id付与済みなのでマッピング不要）
        if self.jobA is not None:
            for a in list(self.jobA.members["aka"]):
                member_name = self.jobA.aka_to_name.get(a, a)
                opt_id_list = []
                for _, wr in Wdf.iterrows():
                    key = (int(wr["id"]), a)
                    if key in overrides:
                        val = overrides[key]
                    else:
                        var = self.x_vars.get(key)
                        val = 1 if var is not None and pulp.value(var) == 1 else 0
                    if val == 1:
                        opt_id_list.append((wr["date"], wr["start_min"], str(int(wr["id"]))))
                
                duty_list = duty_ids_by_member.get(member_name, [])
                all_ids = opt_id_list + duty_list
                all_ids_sorted = sorted(all_ids, key=lambda x: (x[0], x[1]))
                sorted_id_strings = [item[2] for item in all_ids_sorted]
                
                duty_count = len(duty_list)
                opt_count = len(opt_id_list)
                total_count = opt_count + duty_count
                rows_stats.append({
                    "member": member_name,
                    "times": total_count,
                    "id": ",".join(sorted_id_strings),
                    "duty_count": duty_count,
                    "opt_count": opt_count
                })
        if self.jobB is not None:
            for b in list(self.jobB.members["aka"]):
                member_name = self.jobB.aka_to_name.get(b, b)
                opt_id_list = []
                for _, wr in Wdf.iterrows():
                    key = (int(wr["id"]), b)
                    if key in overrides:
                        val = overrides[key]
                    else:
                        var = self.y_vars.get(key)
                        val = 1 if var is not None and pulp.value(var) == 1 else 0
                    if val == 1:
                        opt_id_list.append((wr["date"], wr["start_min"], str(int(wr["id"]))))
                
                duty_list = duty_ids_by_member.get(member_name, [])
                all_ids = opt_id_list + duty_list
                all_ids_sorted = sorted(all_ids, key=lambda x: (x[0], x[1]))
                sorted_id_strings = [item[2] for item in all_ids_sorted]
                
                duty_count = len(duty_list)
                opt_count = len(opt_id_list)
                total_count = opt_count + duty_count
                rows_stats.append({
                    "member": member_name,
                    "times": total_count,
                    "id": ",".join(sorted_id_strings),
                    "duty_count": duty_count,
                    "opt_count": opt_count
                })
        
        return (
            df_assign,
            pd.DataFrame(rows_available_A),
            pd.DataFrame(rows_available_B),
            pd.DataFrame(rows_stats),
            pd.DataFrame(rows_violation),
        )

    def _extract_violations(self):
        """Extract penalty violations for the violation sheet with detailed breakdown"""
        rows_violation = []

        for reason in self.violations:
            rows_violation.append({
                "id": "",
                "penalty": "infeasible",
                "score": 0,
                "description": reason
            })

        if hasattr(self, 'solution_values') and self.solution_values:
            W = list(self.work.work["id"].astype(int))
            w_rows = {int(r["id"]): r for _, r in self.work.work.iterrows()}

            def get_var_value(var):
                """Helper to get variable value, handling None"""
                if var is None:
                    return 0
                return pulp.value(var) or 0

            prefer_penalty = float(self.setting.penalty_map.get("prefer", {"value": 1}).get("value", 1.0))
            if prefer_penalty > 0:
                if self.jobA is not None:
                    A_list = list(self.jobA.members["aka"])
                    for a in A_list:
                        row = self.jobA.members[self.jobA.members["aka"] == a].iloc[0]
                        prefer_tokens = row.get("prefer_tokens", set())

                        for w in W:
                            var = self.x_vars.get((w, a))
                            if get_var_value(var) == 1:
                                work_row = w_rows[w]
                                work_tokens = set()
                                if pd.notna(work_row.get("room")):
                                    work_tokens.add(str(work_row["room"]))
                                if pd.notna(work_row.get("dept")):
                                    work_tokens.add(str(work_row["dept"]))

                                matched_prefs = prefer_tokens.intersection(work_tokens)
                                if matched_prefs:
                                    member_name = self.jobA.aka_to_name.get(a, a)
                                    reward_score = -len(matched_prefs) * prefer_penalty  # Negative because it's a reward
                                    rows_violation.append({
                                        "id": str(w),
                                        "penalty": "prefer",
                                        "score": reward_score,
                                        "description": f"{member_name} 希望マッチ: {','.join(matched_prefs)}"
                                    })

                if self.jobB is not None:
                    B_list = list(self.jobB.members["aka"])
                    for b in B_list:
                        row = self.jobB.members[self.jobB.members["aka"] == b].iloc[0]
                        prefer_tokens = row.get("prefer_tokens", set())

                        for w in W:
                            var = self.y_vars.get((w, b))
                            if get_var_value(var) == 1:
                                work_row = w_rows[w]
                                work_tokens = set()
                                if pd.notna(work_row.get("room")):
                                    work_tokens.add(str(work_row["room"]))
                                if pd.notna(work_row.get("dept")):
                                    work_tokens.add(str(work_row["dept"]))

                                matched_prefs = prefer_tokens.intersection(work_tokens)
                                if matched_prefs:
                                    member_name = self.jobB.aka_to_name.get(b, b)
                                    reward_score = -len(matched_prefs) * prefer_penalty  # Negative because it's a reward
                                    rows_violation.append({
                                        "id": str(w),
                                        "penalty": "prefer",
                                        "score": reward_score,
                                        "description": f"{member_name} 希望マッチ: {','.join(matched_prefs)}"
                                    })

            order_penalty = float(self.setting.penalty_map.get("order", {"value": 100}).get("value", 100.0))
            if order_penalty > 0:
                if self.jobA is not None:
                    A_list = list(self.jobA.members["aka"])
                    for a in A_list:
                        for w in W:
                            var = self.x_vars.get((w, a))
                            if get_var_value(var) == 1:
                                work_row = w_rows[w]
                                dept = str(work_row.get("dept"))
                                score = self.jobA.get_order_score(dept, a)
                                if score is not None:
                                    member_name = self.jobA.aka_to_name.get(a, a)
                                    penalty_score = int(score * order_penalty)
                                    rows_violation.append({
                                        "id": str(w),
                                        "penalty": "order",
                                        "score": penalty_score,
                                        "description": f"{member_name} 優先度: {score:.3f} (業務: {dept})"
                                    })

                if self.jobB is not None:
                    B_list = list(self.jobB.members["aka"])
                    for b in B_list:
                        for w in W:
                            var = self.y_vars.get((w, b))
                            if get_var_value(var) == 1:
                                work_row = w_rows[w]
                                dept = str(work_row.get("dept"))
                                score = self.jobB.get_order_score(dept, b)
                                if score is not None:
                                    member_name = self.jobB.aka_to_name.get(b, b)
                                    penalty_score = int(score * order_penalty)
                                    rows_violation.append({
                                        "id": str(w),
                                        "penalty": "order",
                                        "score": penalty_score,
                                        "description": f"{member_name} 優先度: {score:.3f} (業務: {dept})"
                                    })

            diff_cfg = self.setting.penalty_map.get("diff", {"value": 1, "thres": 0})
            diff_report_penalty = float(diff_cfg.get("value", 1.0))
            diff_report_thres_pct = int(diff_cfg.get("thres", 0))
            if diff_report_penalty > 0:
                duty_counts = {}
                if self.duty_template:
                    for assignment in self.duty_template.assignments:
                        member_name = assignment['name']
                        duty_counts[member_name] = duty_counts.get(member_name, 0) + 1
                
                def _report_load_group(job, member_list, var_dict, role_label):
                    load_groups = {}
                    for m in member_list:
                        load_type = job.load.get(m, "normal")
                        if load_type == "reserve":
                            continue
                        load_groups.setdefault(load_type, []).append(m)

                    for load_type, members in load_groups.items():
                        if len(members) <= 1:
                            continue
                        assignments = {}
                        for m in members:
                            member_name = job.aka_to_name.get(m, m)
                            opt_count = sum(1 for w in W if get_var_value(var_dict.get((w, m))) == 1)
                            duty_count = duty_counts.get(member_name, 0)
                            assignments[m] = opt_count + duty_count

                        assignment_counts = list(assignments.values())
                        max_a = max(assignment_counts)
                        min_a = min(assignment_counts)
                        workload_diff = max_a - min_a

                        avg_per_member = sum(assignment_counts) / len(assignment_counts)
                        threshold_count = max(1, int(avg_per_member * diff_report_thres_pct / 100)) if diff_report_thres_pct > 0 else 0
                        excess = max(0, workload_diff - threshold_count)

                        if excess > 0:
                            penalty_score = excess * diff_report_penalty
                            rows_violation.append({
                                "id": "",
                                "penalty": f"diff_{load_type}",
                                "score": penalty_score,
                                "description": f"負荷平準化違反 ({load_type}): 最大{max_a}件 - 最小{min_a}件 = {workload_diff}件差 (許容{threshold_count}件)"
                            })

                if self.jobA is not None:
                    A_list = list(self.jobA.members["aka"])
                    _report_load_group(self.jobA, A_list, self.x_vars, "A")

                if self.jobB is not None:
                    B_list = list(self.jobB.members["aka"])
                    _report_load_group(self.jobB, B_list, self.y_vars, "B")

            after_duty_penalty = float(self.setting.penalty_map.get("after_duty", {"value": 10}).get("value", 10.0))
            if after_duty_penalty > 0:
                if self.jobA is not None and self.jobA.leave_type_by_date:
                    A_list = list(self.jobA.members["aka"])
                    for a in A_list:
                        for w in W:
                            var = self.x_vars.get((w, a))
                            if get_var_value(var) == 1:
                                work_row = w_rows[w]
                                date = self._work_date_to_iso(work_row["date"])
                                leave_type = self.jobA.leave_type_by_date.get(date, {}).get(a, '')
                                if leave_type == 'D':
                                    member_name = self.jobA.aka_to_name.get(a, a)
                                    penalty_score = int(after_duty_penalty)
                                    rows_violation.append({
                                        "id": str(w),
                                        "penalty": "after_duty",
                                        "score": penalty_score,
                                        "description": f"{member_name} 当直明け割当 ({date})"
                                    })

                if self.jobB is not None and self.jobB.leave_type_by_date:
                    B_list = list(self.jobB.members["aka"])
                    for b in B_list:
                        for w in W:
                            var = self.y_vars.get((w, b))
                            if get_var_value(var) == 1:
                                work_row = w_rows[w]
                                date = self._work_date_to_iso(work_row["date"])
                                leave_type = self.jobB.leave_type_by_date.get(date, {}).get(b, '')
                                if leave_type == 'D':
                                    member_name = self.jobB.aka_to_name.get(b, b)
                                    penalty_score = int(after_duty_penalty)
                                    rows_violation.append({
                                        "id": str(w),
                                        "penalty": "after_duty",
                                        "score": penalty_score,
                                        "description": f"{member_name} 当直明け割当 ({date})"
                                    })

        return rows_violation


CLOSE_DUTY_TYPES = ["out", "teller", "sat", "card"]
CLOSE_DUTY_SYMBOLS = {"out": "+", "teller": "*", "sat": "$", "card": "c"}
CLOSE_DUTY_NAMES = {"out": "外来締め", "teller": "精算機締め", "sat": "土曜締め", "card": "カード締め"}
CLOSE_DUTY_FLAGS = {"out": "out", "teller": "teller", "sat": "close", "card": "card"}


def assign_close_duties(jobA: Optional['JobData'], target_year: int, target_month: int, assign_df: Optional[pd.DataFrame] = None, duty_template: Optional['DutyTemplate'] = None) -> Dict[str, Dict[int, str]]:
    """Assign close duties to eligible members.
    
    Duty types:
    - out (外来締め, +): Every day, members with out=1
    - teller (精算機締め, *): Every day, members with teller=1
    - sat (土曜締め, $): Every Saturday, members with close=1
    - card (カード締め, c): On 15/16 and 31/1 (2 days), members with card=1
    
    Members with leave_type 'D' on a given day are excluded from close duty assignments for that day.
    Leveling is done across all close duty types for members with any close flag.
    """
    import calendar as cal
    import datetime as dt
    
    if jobA is None:
        return {}
    
    eligible_members_by_type = {duty_type: [] for duty_type in CLOSE_DUTY_TYPES}
    all_close_members = []
    close_members_aka = {}
    
    for _, row in jobA.members.iterrows():
        aka = str(row["aka"])
        member_name = jobA.aka_to_name.get(aka, aka)
        
        if is_bantane_name(member_name):
            continue
        
        has_any_close_flag = False
        
        if row.get("out", 0) == 1:
            eligible_members_by_type["out"].append(member_name)
            has_any_close_flag = True
        if row.get("teller", 0) == 1:
            eligible_members_by_type["teller"].append(member_name)
            has_any_close_flag = True
        if row.get("close", 0) == 1:
            eligible_members_by_type["sat"].append(member_name)
            has_any_close_flag = True
        if row.get("card", 0) == 1:
            eligible_members_by_type["card"].append(member_name)
            has_any_close_flag = True
        
        if has_any_close_flag:
            all_close_members.append(member_name)
            close_members_aka[member_name] = aka
    
    if not all_close_members:
        return {}
    
    last_day_of_month = cal.monthrange(target_year, target_month)[1]
    next_month = target_month + 1 if target_month < 12 else 1
    next_year = target_year if target_month < 12 else target_year + 1
    
    assignments = {duty_type: {} for duty_type in CLOSE_DUTY_TYPES}
    total_close_count = {member: 0 for member in all_close_members}
    
    def get_members_excluded_by_duty_leave(date_iso: str) -> set:
        excluded = set()
        if jobA.leave_type_by_date:
            leave_types = jobA.leave_type_by_date.get(date_iso, {})
            for member_name, aka in close_members_aka.items():
                if leave_types.get(aka, '') == 'D':
                    excluded.add(member_name)
        return excluded
    
    def get_members_with_work_on_day(day: int) -> set:
        members_with_work = set()
        date_str = parse_date_token(str(day), target_year, target_month, use_period=True, cutoff=PERIOD_CUTOFF_DAY)
        
        if assign_df is not None and not assign_df.empty:
            day_str = f"{day:02d}"
            day_assignments = assign_df[assign_df['date'] == day_str]
            for _, row in day_assignments.iterrows():
                assign_a = str(row.get('assign_A', ''))
                assign_b = str(row.get('assign_B', ''))
                if assign_a and assign_a != 'nan':
                    members_with_work.update([m.strip() for m in assign_a.split(',') if m.strip()])
                if assign_b and assign_b != 'nan':
                    members_with_work.update([m.strip() for m in assign_b.split(',') if m.strip()])
        
        if duty_template is not None:
            duty_assignments = duty_template.get_assignments_for_date(date_str)
            for assignment in duty_assignments:
                members_with_work.add(assignment['name'])
        
        return members_with_work
    
    def assign_duty(duty_type: str, day: int, date_iso: str, assigned_today: set) -> Optional[str]:
        eligible = eligible_members_by_type[duty_type]
        if not eligible:
            return None
        
        excluded_by_duty_leave = get_members_excluded_by_duty_leave(date_iso)
        members_with_work = get_members_with_work_on_day(day)
        
        available = [m for m in eligible if m in members_with_work and m not in excluded_by_duty_leave and m not in assigned_today]
        if not available:
            available = [m for m in eligible if m not in excluded_by_duty_leave and m not in assigned_today]
        
        if not available:
            return None
        
        candidates = sorted(available, key=lambda m: (total_close_count.get(m, 0), m))
        selected = candidates[0]
        assignments[duty_type][day] = selected
        total_close_count[selected] = total_close_count.get(selected, 0) + 1
        return selected
    
    all_days = list(range(PERIOD_CUTOFF_DAY, last_day_of_month + 1)) + list(range(1, PERIOD_CUTOFF_DAY))
    
    for day in all_days:
        if day >= PERIOD_CUTOFF_DAY:
            day_year, day_month = target_year, target_month
        else:
            day_year, day_month = next_year, next_month
        
        date_obj = dt.date(day_year, day_month, day)
        date_iso = date_obj.strftime("%Y-%m-%d")
        
        is_sunday = date_obj.weekday() == 6
        is_saturday = date_obj.weekday() == 5
        is_holiday = jpholiday.is_holiday(date_obj)
        
        if is_sunday or is_holiday:
            continue
        
        assigned_today = set()
        
        selected = assign_duty("out", day, date_iso, assigned_today)
        if selected:
            assigned_today.add(selected)
        
        selected = assign_duty("teller", day, date_iso, assigned_today)
        if selected:
            assigned_today.add(selected)
        
        if is_saturday:
            assign_duty("sat", day, date_iso, assigned_today)
    
    card_target_days = [(15, 16), (last_day_of_month, 1)]
    
    for primary_day, fallback_day in card_target_days:
        if primary_day >= PERIOD_CUTOFF_DAY:
            day_year, day_month = target_year, target_month
        else:
            day_year, day_month = next_year, next_month
        
        if primary_day > last_day_of_month and day_month == target_month:
            continue
        
        try:
            date_obj = dt.date(day_year, day_month, primary_day)
        except ValueError:
            continue
        
        if jpholiday.is_holiday(date_obj) or date_obj.weekday() == 6:
            if fallback_day >= PERIOD_CUTOFF_DAY:
                fallback_year, fallback_month = target_year, target_month
            else:
                fallback_year, fallback_month = next_year, next_month
            
            try:
                date_obj = dt.date(fallback_year, fallback_month, fallback_day)
            except ValueError:
                continue
        
        if date_obj.day not in all_days:
            continue
        
        actual_day = date_obj.day
        date_iso = date_obj.strftime("%Y-%m-%d")
        
        assigned_on_day = set()
        for duty_type in ["out", "teller", "sat"]:
            if actual_day in assignments[duty_type]:
                assigned_on_day.add(assignments[duty_type][actual_day])
        
        assign_duty("card", actual_day, date_iso, assigned_on_day)
    
    return assignments


def generate_calendar_view(assign_df: pd.DataFrame, target_year: int, target_month: int, work_data: 'WorkData', 
                           jobA: Optional['JobData'] = None, jobB: Optional['JobData'] = None,
                           duty_template: Optional['DutyTemplate'] = None, close_assignments: Optional[Dict] = None):
    """Generate calendar-style view showing staff assignments by date and time period (AM/PM) with formatting"""
    import calendar
    import datetime as dt
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    
    last_day_of_month = calendar.monthrange(target_year, target_month)[1]
    next_month = target_month + 1 if target_month < 12 else 1
    next_year = target_year if target_month < 12 else target_year + 1
    
    all_dates = []
    for day in range(PERIOD_CUTOFF_DAY, last_day_of_month + 1):
        all_dates.append(f"{target_year:04d}-{target_month:02d}-{day:02d}")
    for day in range(1, PERIOD_CUTOFF_DAY):
        all_dates.append(f"{next_year:04d}-{next_month:02d}-{day:02d}")
    
    dept_name_map = {}
    if work_data and hasattr(work_data, 'dept_name'):
        dept_name_map = work_data.dept_name
    
    all_members = []
    member_to_job = {}
    
    if jobA is not None:
        for _, row in jobA.members.iterrows():
            member_name = jobA.aka_to_name.get(row["aka"], row["aka"])
            all_members.append(member_name)
            member_to_job[member_name] = 'A'
    
    if jobB is not None:
        for _, row in jobB.members.iterrows():
            member_name = jobB.aka_to_name.get(row["aka"], row["aka"])
            all_members.append(member_name)
            member_to_job[member_name] = 'B'
    
    # Fallback: if jobA and jobB are None, extract members from assign_df
    if not all_members and assign_df is not None and not assign_df.empty:
        members_from_assign = set()
        if 'assign_A' in assign_df.columns:
            for val in assign_df['assign_A'].dropna():
                for member in str(val).split(','):
                    member = member.strip()
                    if member and not is_bantane_name(member):
                        members_from_assign.add(member)
                        member_to_job[member] = 'A'
        if 'assign_B' in assign_df.columns:
            for val in assign_df['assign_B'].dropna():
                for member in str(val).split(','):
                    member = member.strip()
                    if member and not is_bantane_name(member):
                        members_from_assign.add(member)
                        member_to_job[member] = 'B'
        all_members = sorted(members_from_assign)
    
    assignments = {member: {date: {'AM': [], 'PM': []} for date in all_dates} for member in all_members}
    
    available_by_date = {}
    if jobA is not None:
        for date_str in all_dates:
            available_by_date[date_str] = jobA.avail_by_date.get(date_str, set())
    
    noon_hhmm = 1230
    noon_boundary = hhmm_to_minutes(noon_hhmm)
    
    if assign_df is not None and not assign_df.empty:
        for _, row in assign_df.iterrows():
            date_str = parse_date_token(str(int(row['date'])), target_year, target_month, use_period=True, cutoff=PERIOD_CUTOFF_DAY)
            dept_aka = str(row['dept'])
            dept_name = dept_name_map.get(dept_aka, dept_aka)
            start_time = hhmm_to_minutes(int(row['start']))
            end_time = hhmm_to_minutes(int(row['end']))
            
            if 'assign_A' in row and pd.notna(row['assign_A']) and str(row['assign_A']).strip():
                for member in str(row['assign_A']).split(','):
                    member = member.strip()
                    if member and member in assignments:
                        if end_time <= noon_boundary:
                            assignments[member][date_str]['AM'].append(dept_name)
                        elif start_time >= noon_boundary:
                            assignments[member][date_str]['PM'].append(dept_name)
                        else:
                            assignments[member][date_str]['AM'].append(dept_name)
                            assignments[member][date_str]['PM'].append(dept_name)
            
            if 'assign_B' in row and pd.notna(row['assign_B']) and str(row['assign_B']).strip():
                for member in str(row['assign_B']).split(','):
                    member = member.strip()
                    if member and member in assignments:
                        if end_time <= noon_boundary:
                            assignments[member][date_str]['AM'].append(dept_name)
                        elif start_time >= noon_boundary:
                            assignments[member][date_str]['PM'].append(dept_name)
                        else:
                            assignments[member][date_str]['AM'].append(dept_name)
                            assignments[member][date_str]['PM'].append(dept_name)
    
    if duty_template is not None:
        for assignment in duty_template.assignments:
            member_name = assignment['name']
            date_str = assignment['date']
            duty_name = assignment['duty']
            start_time = hhmm_to_minutes(assignment['start_time'])
            end_time = hhmm_to_minutes(assignment['end_time'])
            
            if member_name in assignments and date_str in assignments[member_name]:
                if end_time <= noon_boundary:
                    assignments[member_name][date_str]['AM'].append(duty_name)
                elif start_time >= noon_boundary:
                    assignments[member_name][date_str]['PM'].append(duty_name)
                else:
                    assignments[member_name][date_str]['AM'].append(duty_name)
                    assignments[member_name][date_str]['PM'].append(duty_name)
    
    header_rows = []
    
    row1 = [f"{target_year}年{target_month}月"]
    for _ in range(len(all_dates) * 2):
        row1.append('')
    header_rows.append(row1)
    
    row2 = ['氏名']
    for date_str in all_dates:
        date_obj = dt.datetime.strptime(date_str, '%Y-%m-%d')
        day_str = f"{date_obj.month}/{date_obj.day}"
        row2.extend([day_str, ''])
    header_rows.append(row2)
    
    row3 = ['']
    weekday_names = ['月', '火', '水', '木', '金', '土', '日']
    for date_str in all_dates:
        date_obj = dt.datetime.strptime(date_str, '%Y-%m-%d')
        weekday = weekday_names[date_obj.weekday()]
        row3.extend([weekday, ''])
    header_rows.append(row3)
    
    row4 = ['']
    for _ in all_dates:
        row4.extend(['AM', 'PM'])
    header_rows.append(row4)
    
    data_rows = []
    for member in all_members:
        row = [member]
        for date_str in all_dates:
            date_obj = dt.datetime.strptime(date_str, '%Y-%m-%d')
            day = date_obj.day
            
            am_tasks = assignments[member][date_str]['AM']
            pm_tasks = assignments[member][date_str]['PM']
            
            am_text = '\n'.join(am_tasks) if am_tasks else ''
            pm_text = '\n'.join(pm_tasks) if pm_tasks else ''
            
            if close_assignments:
                symbols = []
                for duty_type in CLOSE_DUTY_TYPES:
                    if close_assignments.get(duty_type, {}).get(day) == member:
                        symbols.append(CLOSE_DUTY_SYMBOLS[duty_type])
                
                if symbols:
                    if pm_text:
                        pm_text += '\n' + ''.join(symbols)
                    else:
                        pm_text = ''.join(symbols)
            
            row.extend([am_text, pm_text])
        data_rows.append(row)
    
    all_rows = header_rows + data_rows
    
    max_cols = max(len(row) for row in all_rows)
    for row in all_rows:
        while len(row) < max_cols:
            row.append('')
    
    df = pd.DataFrame(all_rows)
    
    unique_depts = set()
    for member in all_members:
        for date_str in all_dates:
            unique_depts.update(assignments[member][date_str]['AM'])
            unique_depts.update(assignments[member][date_str]['PM'])
    
    colors = [
        "FFE6E6", "E6F3FF", "E6FFE6", "FFF9E6", "FFE6F9", "E6FFFF",
        "F9E6FF", "FFE6CC", "E6F9FF", "F3FFE6", "FFDAB9", "E6E6FF",
        "FFB6C1", "B0E0E6", "98FB98", "FFFFE0", "DDA0DD", "AFEEEE"
    ]
    
    dept_colors = {}
    sorted_depts = sorted(unique_depts)
    for idx, dept in enumerate(sorted_depts):
        dept_colors[dept] = colors[idx % len(colors)]
    
    return df, all_dates, dept_colors


def generate_availability_from_assign(assign_df: pd.DataFrame, target_year: int, target_month: int) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Generate フリー(A)/フリー(B) data from 割当情報 when those sheets are missing.
    
    Logic: For each date, find members who appear in 割当情報 but are not assigned to any task in that time slot.
    Free members = Members appearing in assign columns - Members assigned in that time slot
    
    Args:
        assign_df: DataFrame with assignment data (columns: 日付/date, 開始時刻/start, 終了時刻/end, 割当(A)/assign_A, 割当(B)/assign_B)
        target_year: Target year for date parsing
        target_month: Target month for date parsing
    
    Returns:
        Tuple of (available_A_df, available_B_df) with columns: date, day, full, AM, PM
    """
    import datetime as dt
    
    if assign_df.empty:
        return pd.DataFrame(columns=["date", "day", "full", "AM", "PM"]), pd.DataFrame(columns=["date", "day", "full", "AM", "PM"])
    
    # Normalize column names (support both Japanese and English)
    df = assign_df.copy()
    column_mapping = {
        "日付": "date",
        "開始時刻": "start",
        "終了時刻": "end",
        "割当(A)": "assign_A",
        "割当(B)": "assign_B",
    }
    df.rename(columns=column_mapping, inplace=True)
    
    # Check required columns exist
    required_cols = ["date", "start", "end"]
    for col in required_cols:
        if col not in df.columns:
            logging.warning(f"generate_availability_from_assign: Missing required column '{col}'")
            return pd.DataFrame(columns=["date", "day", "full", "AM", "PM"]), pd.DataFrame(columns=["date", "day", "full", "AM", "PM"])
    
    # Extract all members from assign columns
    def extract_members(series: pd.Series) -> set:
        members = set()
        for v in series.dropna():
            for name in str(v).split(","):
                name = name.strip()
                if name and not is_bantane_name(name):
                    members.add(name)
        return members
    
    all_A_members = extract_members(df.get("assign_A", pd.Series(dtype=str)))
    all_B_members = extract_members(df.get("assign_B", pd.Series(dtype=str)))
    
    if not all_A_members and not all_B_members:
        logging.warning("generate_availability_from_assign: No members found in assign columns")
        return pd.DataFrame(columns=["date", "day", "full", "AM", "PM"]), pd.DataFrame(columns=["date", "day", "full", "AM", "PM"])
    
    # Parse dates
    PERIOD_CUTOFF_DAY = 21
    def safe_parse_date(val):
        try:
            return parse_date_token(str(int(val)), target_year, target_month, use_period=True, cutoff=PERIOD_CUTOFF_DAY)
        except:
            return None
    
    df["date_iso"] = df["date"].apply(safe_parse_date)
    df = df[df["date_iso"].notna()]
    
    if df.empty:
        return pd.DataFrame(columns=["date", "day", "full", "AM", "PM"]), pd.DataFrame(columns=["date", "day", "full", "AM", "PM"])
    
    # Parse times
    def safe_parse_time(val):
        try:
            return hhmm_to_minutes(int(val))
        except:
            return None
    
    df["start_min"] = df["start"].apply(safe_parse_time)
    df["end_min"] = df["end"].apply(safe_parse_time)
    df = df[df["start_min"].notna() & df["end_min"].notna()]
    
    noon_boundary = hhmm_to_minutes(1230)
    
    rows_A = []
    rows_B = []
    
    for date_iso in sorted(df["date_iso"].unique()):
        day_df = df[df["date_iso"] == date_iso]
        
        assigned_A_AM = set()
        assigned_A_PM = set()
        assigned_B_AM = set()
        assigned_B_PM = set()
        
        for _, row in day_df.iterrows():
            s = row["start_min"]
            e = row["end_min"]
            spans_AM = s < noon_boundary
            spans_PM = e > noon_boundary
            
            members_A = [m.strip() for m in str(row.get("assign_A", "")).split(",") if m.strip() and not is_bantane_name(m.strip())]
            members_B = [m.strip() for m in str(row.get("assign_B", "")).split(",") if m.strip() and not is_bantane_name(m.strip())]
            
            if spans_AM:
                assigned_A_AM.update(members_A)
                assigned_B_AM.update(members_B)
            if spans_PM:
                assigned_A_PM.update(members_A)
                assigned_B_PM.update(members_B)
        
        # Calculate free members (complement)
        full_A = sorted([m for m in all_A_members if m not in assigned_A_AM and m not in assigned_A_PM])
        am_A = sorted([m for m in all_A_members if m not in assigned_A_AM and m in assigned_A_PM])
        pm_A = sorted([m for m in all_A_members if m in assigned_A_AM and m not in assigned_A_PM])
        
        full_B = sorted([m for m in all_B_members if m not in assigned_B_AM and m not in assigned_B_PM])
        am_B = sorted([m for m in all_B_members if m not in assigned_B_AM and m in assigned_B_PM])
        pm_B = sorted([m for m in all_B_members if m in assigned_B_AM and m not in assigned_B_PM])
        
        dd = to_dd(date_iso)
        weekday_jp = get_japanese_weekday(date_iso)
        
        rows_A.append({
            "date": dd,
            "day": weekday_jp,
            "full": ", ".join(full_A),
            "AM": ", ".join(am_A),
            "PM": ", ".join(pm_A),
        })
        rows_B.append({
            "date": dd,
            "day": weekday_jp,
            "full": ", ".join(full_B),
            "AM": ", ".join(am_B),
            "PM": ", ".join(pm_B),
        })
    
    available_A_df = pd.DataFrame(rows_A, columns=["date", "day", "full", "AM", "PM"])
    available_B_df = pd.DataFrame(rows_B, columns=["date", "day", "full", "AM", "PM"])
    
    return available_A_df, available_B_df


def generate_work_oriented_calendar_view(assign_df: pd.DataFrame, target_year: int, target_month: int, work_data: 'WorkData',
                                         jobA: Optional['JobData'] = None, jobB: Optional['JobData'] = None,
                                         duty_template: Optional['DutyTemplate'] = None, close_assignments: Optional[Dict] = None,
                                         available_A_df: Optional[pd.DataFrame] = None, available_B_df: Optional[pd.DataFrame] = None):
    """Generate work-oriented calendar view showing which members are assigned to each work on each date/time"""
    import calendar
    import datetime as dt
    
    last_day_of_month = calendar.monthrange(target_year, target_month)[1]
    next_month = target_month + 1 if target_month < 12 else 1
    next_year = target_year if target_month < 12 else target_year + 1
    
    all_dates = []
    for day in range(PERIOD_CUTOFF_DAY, last_day_of_month + 1):
        all_dates.append(f"{target_year:04d}-{target_month:02d}-{day:02d}")
    for day in range(1, PERIOD_CUTOFF_DAY):
        all_dates.append(f"{next_year:04d}-{next_month:02d}-{day:02d}")
    
    all_members = []
    if jobA is not None:
        for _, row in jobA.members.iterrows():
            member_name = jobA.aka_to_name.get(row["aka"], row["aka"])
            all_members.append(member_name)
    if jobB is not None:
        for _, row in jobB.members.iterrows():
            member_name = jobB.aka_to_name.get(row["aka"], row["aka"])
            all_members.append(member_name)
    
    dept_name_map = {}
    if work_data and hasattr(work_data, 'dept_name'):
        dept_name_map = work_data.dept_name
    
    work_assignments = {}
    
    noon_hhmm = 1230
    noon_boundary = hhmm_to_minutes(noon_hhmm)
    
    if assign_df is not None and not assign_df.empty:
        for _, row in assign_df.iterrows():
            date_str = parse_date_token(str(int(row['date'])), target_year, target_month, use_period=True, cutoff=PERIOD_CUTOFF_DAY)
            dept_aka = str(row['dept'])
            dept_name = dept_name_map.get(dept_aka, dept_aka)
            start_time = hhmm_to_minutes(int(row['start']))
            end_time = hhmm_to_minutes(int(row['end']))
            
            if dept_name not in work_assignments:
                work_assignments[dept_name] = {date: {'AM': [], 'PM': []} for date in all_dates}
            
            if 'assign_A' in row and pd.notna(row['assign_A']) and str(row['assign_A']).strip():
                for member in str(row['assign_A']).split(','):
                    member = member.strip()
                    if member:
                        if end_time <= noon_boundary:
                            work_assignments[dept_name][date_str]['AM'].append(member)
                        elif start_time >= noon_boundary:
                            work_assignments[dept_name][date_str]['PM'].append(member)
                        else:
                            work_assignments[dept_name][date_str]['AM'].append(member)
                            work_assignments[dept_name][date_str]['PM'].append(member)
            
            if 'assign_B' in row and pd.notna(row['assign_B']) and str(row['assign_B']).strip():
                for member in str(row['assign_B']).split(','):
                    member = member.strip()
                    if member:
                        if end_time <= noon_boundary:
                            work_assignments[dept_name][date_str]['AM'].append(member)
                        elif start_time >= noon_boundary:
                            work_assignments[dept_name][date_str]['PM'].append(member)
                        else:
                            work_assignments[dept_name][date_str]['AM'].append(member)
                            work_assignments[dept_name][date_str]['PM'].append(member)
    
    duty_busy_by_date = {date: {'AM': set(), 'PM': set()} for date in all_dates}
    
    if duty_template is not None:
        if '個人業務' not in work_assignments:
            work_assignments['個人業務'] = {date: {'AM': [], 'PM': []} for date in all_dates}
        
        for assignment in duty_template.assignments:
            member_name = assignment['name']
            date_str = assignment['date']
            duty_name = assignment['duty']
            start_time = hhmm_to_minutes(assignment['start_time'])
            end_time = hhmm_to_minutes(assignment['end_time'])
            
            if date_str in all_dates:
                if duty_name in work_assignments:
                    target_key = duty_name
                    info_text = member_name
                else:
                    target_key = '個人業務'
                    info_text = f"{member_name}({duty_name})"
                
                if end_time <= noon_boundary:
                    work_assignments[target_key][date_str]['AM'].append(info_text)
                    duty_busy_by_date[date_str]['AM'].add(member_name)
                elif start_time >= noon_boundary:
                    work_assignments[target_key][date_str]['PM'].append(info_text)
                    duty_busy_by_date[date_str]['PM'].add(member_name)
                else:
                    work_assignments[target_key][date_str]['AM'].append(info_text)
                    work_assignments[target_key][date_str]['PM'].append(info_text)
                    duty_busy_by_date[date_str]['AM'].add(member_name)
                    duty_busy_by_date[date_str]['PM'].add(member_name)
    
    if close_assignments:
        for duty_type in CLOSE_DUTY_TYPES:
            work_name = CLOSE_DUTY_NAMES[duty_type]
            if close_assignments.get(duty_type):
                if work_name not in work_assignments:
                    work_assignments[work_name] = {date: {'AM': [], 'PM': []} for date in all_dates}
                
                for day, member in close_assignments[duty_type].items():
                    date_str = parse_date_token(str(day), target_year, target_month, use_period=True, cutoff=PERIOD_CUTOFF_DAY)
                    if date_str in work_assignments[work_name]:
                        work_assignments[work_name][date_str]['PM'].append(member)
    
    header_rows = []
    
    row1 = [f"{target_year}年{target_month}月"]
    for _ in range(len(all_dates) * 2):
        row1.append('')
    header_rows.append(row1)
    
    row2 = ['']
    for date_str in all_dates:
        date_obj = dt.datetime.strptime(date_str, '%Y-%m-%d')
        day_str = f"{date_obj.month}/{date_obj.day}"
        row2.extend([day_str, ''])
    header_rows.append(row2)
    
    row3 = ['']
    weekday_names = ['月', '火', '水', '木', '金', '土', '日']
    for date_str in all_dates:
        date_obj = dt.datetime.strptime(date_str, '%Y-%m-%d')
        weekday = weekday_names[date_obj.weekday()]
        row3.extend([weekday, ''])
    header_rows.append(row3)
    
    row4 = ['時間帯']
    for _ in all_dates:
        row4.extend(['AM', 'PM'])
    header_rows.append(row4)
    
    rest_row_data = {}
    available_row_data = {}
    
    for date_str in all_dates:
        date_obj = dt.datetime.strptime(date_str, '%Y-%m-%d')
        day = date_obj.day
        day_str = to_dd(date_str)
        
        available_am = []
        available_pm = []
        
        if available_A_df is not None and not available_A_df.empty and 'date' in available_A_df.columns:
            date_rows = available_A_df[available_A_df['date'].astype(str) == day_str]
            for _, row in date_rows.iterrows():
                full_members = str(row.get('full', '')).split(',') if pd.notna(row.get('full')) else []
                am_members = str(row.get('AM', '')).split(',') if pd.notna(row.get('AM')) else []
                pm_members = str(row.get('PM', '')).split(',') if pd.notna(row.get('PM')) else []
                
                for m in full_members:
                    m = m.strip()
                    if m:
                        available_am.append(m)
                        available_pm.append(m)
                
                for m in am_members:
                    m = m.strip()
                    if m:
                        available_am.append(m)
                
                for m in pm_members:
                    m = m.strip()
                    if m:
                        available_pm.append(m)
        
        if available_B_df is not None and not available_B_df.empty and 'date' in available_B_df.columns:
            date_rows = available_B_df[available_B_df['date'].astype(str) == day_str]
            for _, row in date_rows.iterrows():
                full_members = str(row.get('full', '')).split(',') if pd.notna(row.get('full')) else []
                am_members = str(row.get('AM', '')).split(',') if pd.notna(row.get('AM')) else []
                pm_members = str(row.get('PM', '')).split(',') if pd.notna(row.get('PM')) else []
                
                for m in full_members:
                    m = m.strip()
                    if m:
                        available_am.append(m)
                        available_pm.append(m)
                
                for m in am_members:
                    m = m.strip()
                    if m:
                        available_am.append(m)
                
                for m in pm_members:
                    m = m.strip()
                    if m:
                        available_pm.append(m)
        
        def base_member_name(s: str) -> str:
            s = s.strip()
            if s.endswith(' am') or s.endswith(' pm'):
                return s.rsplit(' ', 1)[0]
            return s
        
        busy_am = duty_busy_by_date[date_str]['AM']
        busy_pm = duty_busy_by_date[date_str]['PM']
        available_am = [m for m in available_am if base_member_name(m) not in busy_am]
        available_pm = [m for m in available_pm if base_member_name(m) not in busy_pm]
        
        available_am = [m for m in available_am if not is_bantane_name(base_member_name(m))]
        available_pm = [m for m in available_pm if not is_bantane_name(base_member_name(m))]
        
        available_row_data[date_str] = {'AM': available_am, 'PM': available_pm}
        
        is_sunday = date_obj.weekday() == 6
        
        if is_sunday:
            rest_row_data[date_str] = {'AM': [], 'PM': []}
        else:
            staff_on_date = set()
            if jobA is not None and hasattr(jobA, 'avail_by_date'):
                staff_akas = jobA.avail_by_date.get(date_str, set())
                for aka in staff_akas:
                    member_name = jobA.aka_to_name.get(aka, aka)
                    staff_on_date.add(member_name)
            
            rest_members = [m for m in all_members if m not in staff_on_date]
            
            rest_row_data[date_str] = {'AM': rest_members, 'PM': rest_members}
    
    merge_groups = {
        '会計': ['会計１', '会計２', '会計３', '会計４'],
        '外来受付': ['外来受付１', '外来受付２', '外来受付３'],
        '文書': ['文書１', '文書２'],
        '紹介受付': ['紹介受付１', '紹介受付２', '紹介受付３']
    }
    
    merged_assignments = {}
    for merged_name, source_names in merge_groups.items():
        merged_assignments[merged_name] = {date: {'AM': [], 'PM': []} for date in all_dates}
        for source_name in source_names:
            if source_name in work_assignments:
                for date_str in all_dates:
                    merged_assignments[merged_name][date_str]['AM'].extend(work_assignments[source_name][date_str]['AM'])
                    merged_assignments[merged_name][date_str]['PM'].extend(work_assignments[source_name][date_str]['PM'])
    
    for merged_name in merge_groups.keys():
        for source_name in merge_groups[merged_name]:
            if source_name in work_assignments:
                del work_assignments[source_name]
    
    work_assignments.update(merged_assignments)
    
    desired_order = ['健診', '文書', '会計', '紹介受付', '外来受付', '登録', '引継ぎ', '外来締め', '精算機締め', '土曜締め', 'カード締め', '個人業務', 'フリー']
    
    ordered_works = [w for w in desired_order if w in work_assignments or w == 'フリー']
    remaining_works = [w for w in sorted(work_assignments.keys()) if w not in desired_order]
    
    all_work_names = ordered_works + remaining_works
    
    transposed_rows = []
    
    row1 = [f"{target_year}年{target_month}月", '']
    for work_name in all_work_names:
        row1.extend([work_name, ''])
    transposed_rows.append(row1)
    
    row2 = ['日付', '曜日']
    for _ in all_work_names:
        row2.extend(['AM', 'PM'])
    transposed_rows.append(row2)
    
    weekday_names = ['月', '火', '水', '木', '金', '土', '日']
    
    for date_str in all_dates:
        date_obj = dt.datetime.strptime(date_str, '%Y-%m-%d')
        day_str = f"{date_obj.month}/{date_obj.day}"
        weekday = weekday_names[date_obj.weekday()]
        
        row = [day_str, weekday]
        
        for work_name in all_work_names:
            if work_name == 'フリー':
                am_text = ', '.join(available_row_data[date_str]['AM']) if available_row_data[date_str]['AM'] else ''
                pm_text = ', '.join(available_row_data[date_str]['PM']) if available_row_data[date_str]['PM'] else ''
            else:
                am_members = work_assignments[work_name][date_str]['AM']
                pm_members = work_assignments[work_name][date_str]['PM']
                am_text = ', '.join(am_members) if am_members else ''
                pm_text = ', '.join(pm_members) if pm_members else ''
            
            row.extend([am_text, pm_text])
        
        transposed_rows.append(row)
    
    max_cols = max(len(row) for row in transposed_rows)
    for row in transposed_rows:
        while len(row) < max_cols:
            row.append('')
    
    df = pd.DataFrame(transposed_rows)
    
    return df, all_dates, all_work_names


def apply_work_oriented_calendar_formatting(ws, all_dates, all_work_names, target_year: int, target_month: int):
    """Apply formatting to work-oriented calendar view worksheet (transposed layout)"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import datetime as dt
    
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
    
    for work_idx, work_name in enumerate(all_work_names):
        col_am = 3 + work_idx * 2
        col_pm = col_am + 1
        
        ws.merge_cells(start_row=1, start_column=col_am, end_row=1, end_column=col_pm)
        ws.cell(row=1, column=col_am).alignment = Alignment(horizontal='center', vertical='center')
    
    for date_idx, date_str in enumerate(all_dates):
        date_obj = dt.datetime.strptime(date_str, '%Y-%m-%d')
        is_sunday = date_obj.weekday() == 6
        is_holiday = jpholiday.is_holiday(date_obj)
        
        row_idx = 3 + date_idx
        
        for work_idx in range(len(all_work_names)):
            col_am = 3 + work_idx * 2
            col_pm = col_am + 1
            
            cell_am = ws.cell(row=row_idx, column=col_am)
            cell_pm = ws.cell(row=row_idx, column=col_pm)
            
            if is_sunday or is_holiday:
                cell_am.fill = gray_fill
                cell_pm.fill = gray_fill
            else:
                cell_am.fill = blue_fill
                cell_pm.fill = red_fill
    
    base_width = 12
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = base_width
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 5
    
    header_row = 1
    for col_idx in range(3, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        if cell.value:
            val = str(cell.value).strip()
            if val == '個人業務':
                ws.column_dimensions[get_column_letter(col_idx)].width = base_width * 4
                if col_idx + 1 <= ws.max_column:
                    ws.column_dimensions[get_column_letter(col_idx + 1)].width = base_width * 4
            elif val == 'フリー':
                ws.column_dimensions[get_column_letter(col_idx)].width = base_width * 4
                if col_idx + 1 <= ws.max_column:
                    ws.column_dimensions[get_column_letter(col_idx + 1)].width = base_width * 4
    
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 45


def apply_calendar_formatting(ws, all_dates, target_year: int, target_month: int, dept_colors: dict):
    """Apply formatting to calendar view worksheet with department-based colors"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import datetime as dt
    
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
    
    for date_idx, date_str in enumerate(all_dates):
        date_obj = dt.datetime.strptime(date_str, '%Y-%m-%d')
        is_sunday = date_obj.weekday() == 6
        
        col_am = 2 + date_idx * 2
        col_pm = col_am + 1
        
        for row_idx in range(5, ws.max_row + 1):
            cell_am = ws.cell(row=row_idx, column=col_am)
            cell_pm = ws.cell(row=row_idx, column=col_pm)
            
            if is_sunday:
                cell_am.fill = gray_fill
                cell_pm.fill = gray_fill
            else:
                am_text = str(cell_am.value) if cell_am.value else ""
                pm_text = str(cell_pm.value) if cell_pm.value else ""
                
                am_depts = [d.strip() for d in am_text.replace('+', '').replace('*', '').replace('$', '').replace('c', '').split('\n') if d.strip()]
                pm_depts = [d.strip() for d in pm_text.replace('+', '').replace('*', '').replace('$', '').replace('c', '').split('\n') if d.strip()]
                
                if am_depts and am_depts[0] in dept_colors:
                    cell_am.fill = PatternFill(start_color=dept_colors[am_depts[0]], end_color=dept_colors[am_depts[0]], fill_type="solid")
                else:
                    cell_am.fill = white_fill
                
                if pm_depts and pm_depts[0] in dept_colors:
                    cell_pm.fill = PatternFill(start_color=dept_colors[pm_depts[0]], end_color=dept_colors[pm_depts[0]], fill_type="solid")
                else:
                    cell_pm.fill = white_fill
    
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
    
    ws.column_dimensions['A'].width = 15
    
    for row_idx in range(1, ws.max_row + 1):
        if row_idx == 1:
            ws.row_dimensions[row_idx].height = 60
        elif row_idx == 2:
            ws.row_dimensions[row_idx].height = 30
        else:
            ws.row_dimensions[row_idx].height = 40
    
    for date_idx in range(len(all_dates)):
        col_date = 2 + date_idx * 2
        col_date_next = col_date + 1
        
        ws.merge_cells(start_row=2, start_column=col_date, end_row=2, end_column=col_date_next)
        ws.cell(row=2, column=col_date).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        ws.merge_cells(start_row=3, start_column=col_date, end_row=3, end_column=col_date_next)
        ws.cell(row=3, column=col_date).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    legend_col = ws.max_column + 2
    ws.cell(row=1, column=legend_col).value = "【凡例】"
    ws.cell(row=1, column=legend_col).font = Font(bold=True)
    row_idx = 2
    for duty_type in CLOSE_DUTY_TYPES:
        symbol = CLOSE_DUTY_SYMBOLS[duty_type]
        name = CLOSE_DUTY_NAMES[duty_type]
        ws.cell(row=row_idx, column=legend_col).value = f"{symbol} {name}"
        row_idx += 1
    ws.column_dimensions[get_column_letter(legend_col)].width = 20
    
    cell_a1 = ws.cell(row=1, column=1)
    if cell_a1.value and str(cell_a1.value).strip() == '氏名':
        cell_a1.font = Font(color="FFFFFF")


_RESERVE_NAMES: Set[str] = set()
_UNREGISTERED_STAFF: List[str] = []

WORD_JOINER = '\u2060'

def protect_name_for_excel(name: str) -> str:
    """Insert WORD JOINER between each character to prevent line breaks within the name"""
    s = str(name)
    if len(s) <= 1:
        return s
    return WORD_JOINER.join(list(s))

def protect_text_for_excel(text: str, separator: str = ', ') -> str:
    """Protect all names in a comma-separated text by inserting WORD JOINER within each name"""
    if not text:
        return text
    names = [n.strip() for n in text.split(',') if n.strip()]
    protected_names = [protect_name_for_excel(n) for n in names]
    return separator.join(protected_names)


def set_reserve_names(names: Set[str]):
    """Set the global reserve (dummy) staff names from setting load='reserve'."""
    global _RESERVE_NAMES
    _RESERVE_NAMES = names


def is_bantane_name(name: str) -> bool:
    """Check if a name is a reserve (dummy) staff name."""
    return name.strip() in _RESERVE_NAMES


def filter_bantane_from_text(text: str, separator: str = ",") -> str:
    """Filter out bantane names from a comma-separated or newline-separated text."""
    if not text or pd.isna(text):
        return ""
    parts = [p.strip() for p in str(text).split(separator)]
    filtered = [p for p in parts if p and not is_bantane_name(p.split()[0] if ' ' in p else p)]
    return separator.join(filtered)


def auto_adjust_column_width(ws, max_width: int = 50, min_width: int = 8):
    """Auto-adjust column widths based on content with text wrapping for long content.
    
    Args:
        ws: openpyxl worksheet
        max_width: Maximum column width (default 50). Content exceeding this will wrap.
        min_width: Minimum column width (default 8).
    """
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell_value = str(cell.value)
                lines = cell_value.split('\n')
                line_max = max(len(line) for line in lines) if lines else 0
                
                if line_max > max_length:
                    max_length = line_max
                
                cell.alignment = Alignment(
                    horizontal='left',
                    vertical='center',
                    wrap_text=True
                )
        
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def format_assign_sheet(ws):
    """Format assign sheet: adjust column widths, center all content."""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    
    header_row = 1
    assign_a_col = None
    assign_b_col = None
    start_col = None
    end_col = None
    base_width = 12
    
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        val = str(cell.value).strip() if cell.value else ""
        if val == "割当(A)":
            assign_a_col = col_idx
        elif val == "割当(B)":
            assign_b_col = col_idx
        elif val == "開始時刻":
            start_col = col_idx
        elif val == "終了時刻":
            end_col = col_idx
    
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        if col_idx == assign_a_col or col_idx == assign_b_col:
            ws.column_dimensions[column_letter].width = base_width * 3
        elif col_idx == start_col or col_idx == end_col:
            ws.column_dimensions[column_letter].width = base_width * 1.5
        else:
            max_length = 0
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max(max_length + 2, 8)
    
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def format_available_sheet(ws):
    """Format available_A/B sheet: center date/day/full columns, center all headers."""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    
    header_row = 1
    date_col = day_col = full_col = am_col = pm_col = None
    
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        val = str(cell.value).strip() if cell.value else ""
        if val == "日付":
            date_col = col_idx
        elif val == "曜日":
            day_col = col_idx
        elif val == "終日フリー":
            full_col = col_idx
        elif val == "AMフリー":
            am_col = col_idx
        elif val == "PMフリー":
            pm_col = col_idx
    
    for col_idx in [date_col, day_col, full_col, am_col, pm_col]:
        if col_idx:
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for row_idx in range(1, ws.max_row + 1):
        if row_idx == 1:
            ws.row_dimensions[row_idx].height = 15
        else:
            ws.row_dimensions[row_idx].height = 30


def format_stats_sheet(ws):
    """Format stats sheet: center all content, double width for specific columns."""
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
    
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for row_idx in range(1, ws.max_row + 1):
        if row_idx == 1:
            ws.row_dimensions[row_idx].height = 15
        else:
            ws.row_dimensions[row_idx].height = 45
    
    cell_a1 = ws.cell(row=1, column=1)
    if cell_a1.value and str(cell_a1.value).strip() == 'member':
        cell_a1.font = Font(color="FFFFFF")
    
    header_row = 1
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        val = str(cell.value).strip() if cell.value else ""
        if val in ["合計回数", "個人業務", "最適化割当"]:
            column_letter = get_column_letter(col_idx)
            current_width = ws.column_dimensions[column_letter].width or 12
            ws.column_dimensions[column_letter].width = current_width * 2


def format_violation_sheet(ws):
    """Format violation sheet: double description width, center id/penalty/score columns."""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    
    header_row = 1
    id_col = penalty_col = score_col = description_col = None
    
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        val = str(cell.value).strip() if cell.value else ""
        if val == "業務ID":
            id_col = col_idx
        elif val == "制約":
            penalty_col = col_idx
        elif val == "点数":
            score_col = col_idx
        elif val == "詳細":
            description_col = col_idx
    
    if description_col:
        column_letter = get_column_letter(description_col)
        current_width = ws.column_dimensions[column_letter].width or 12
        ws.column_dimensions[column_letter].width = current_width * 2
    
    for col_idx in [id_col, penalty_col, score_col]:
        if col_idx:
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def format_duty_sheet(ws):
    """Format duty sheet: center all content and double the width of '担当' column."""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    
    # Find the column index for '担当' (or 'name' for backward compatibility)
    name_col_idx = None
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value in ('担当', 'name'):
            name_col_idx = col_idx
            break
    
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Double the width of '担当' column
    if name_col_idx:
        col_letter = get_column_letter(name_col_idx)
        current_width = ws.column_dimensions[col_letter].width or 10
        ws.column_dimensions[col_letter].width = min(current_width * 2, 255)


def generate_weekly_gantt_images(assign_df: pd.DataFrame, output_path: Path, target_year: int, target_month: int, work_data: 'WorkData', duty_template: Optional['DutyTemplate'] = None, available_A_df: Optional[pd.DataFrame] = None, available_B_df: Optional[pd.DataFrame] = None, close_assignments: Optional[Dict] = None):
    """Generate daily gantt chart images for each week with Windows font compatibility"""
    if (assign_df is None or assign_df.empty) and duty_template is None:
        return

    import datetime as dt
    from io import BytesIO
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from matplotlib.patches import Rectangle
    import numpy as np
    import matplotlib.font_manager as fm
    from openpyxl import load_workbook, Workbook
    from openpyxl.drawing.image import Image

    configure_matplotlib_fonts()

    if assign_df is not None and not assign_df.empty:
        assign_df = assign_df.copy()
        assign_df['date_int'] = assign_df['date'].astype(int)

        datetime_list = []
        for _, row in assign_df.iterrows():
            try:
                date_iso = parse_date_token(str(int(row['date_int'])), target_year, target_month, use_period=True, cutoff=PERIOD_CUTOFF_DAY)
                date_obj = dt.datetime.fromisoformat(date_iso)
                datetime_list.append(date_obj)
            except ValueError:
                datetime_list.append(None)

        assign_df['datetime'] = datetime_list
        assign_df = assign_df.dropna(subset=['datetime'])

        if assign_df.empty and duty_template is None:
            return

        week_start_list = []
        for dt_obj in assign_df['datetime']:
            week_start = dt_obj - dt.timedelta(days=dt_obj.weekday())
            week_start_list.append(week_start)

        assign_df['week_start'] = week_start_list
        unique_weeks = sorted(assign_df['week_start'].unique())
    else:
        assign_df = pd.DataFrame()
        if duty_template and duty_template.assignments:
            unique_dates = set()
            for assignment in duty_template.assignments:
                try:
                    date_obj = dt.datetime.strptime(assignment['date'], "%Y-%m-%d")
                    unique_dates.add(date_obj)
                except ValueError:
                    continue

            if not unique_dates:
                return

            unique_weeks = set()
            for date_obj in unique_dates:
                week_start = date_obj - dt.timedelta(days=date_obj.weekday())
                unique_weeks.add(week_start)
            unique_weeks = sorted(unique_weeks)
        else:
            return

    try:
        wb = load_workbook(output_path)
    except FileNotFoundError:
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

    colors = ['#8B0000', '#FF0000', '#FF4500', '#FFA500', '#FFFF00', '#ADFF2F',
              '#00FF7F', '#00FFFF', '#0080FF', '#0000FF', '#8000FF', '#FF00FF']

    week_list = []
    for week_start_dt in unique_weeks:
        if hasattr(week_start_dt, 'to_pydatetime'):
            week_start_dt = week_start_dt.to_pydatetime()
        elif not isinstance(week_start_dt, dt.datetime):
            continue
        week_list.append(week_start_dt)

    week_list.sort()

    for week_start_dt in week_list:
        week_end = week_start_dt + dt.timedelta(days=6)
        if not assign_df.empty:
            week_data = assign_df[
                (assign_df['datetime'] >= week_start_dt) &
                (assign_df['datetime'] <= week_end)
            ]
        else:
            week_data = pd.DataFrame()

        # Calculate period start (21st of target_month)
        period_start = dt.datetime(target_year, target_month, PERIOD_CUTOFF_DAY)
        
        first_period_monday = period_start
        while first_period_monday.weekday() != 0:  # 0 = Monday
            first_period_monday -= dt.timedelta(days=1)
        
        # Calculate which week this is from the first period Monday
        days_from_first_period_monday = (week_start_dt - first_period_monday).days
        week_number = (days_from_first_period_monday // 7) + 1
        
        week_number = max(1, week_number)
        
        sheet_name = f"第{week_number}週"

        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)

        weekday_names = ['月', '火', '水', '木', '金', '土', '日']

        current_row = 1
        for day_offset in range(7):
            current_date = week_start_dt + dt.timedelta(days=day_offset)
            day_data = week_data[week_data['datetime'].dt.date == current_date.date()]

            if day_data.empty:
                continue

            assigned_rooms_data = day_data[
                (day_data['assign_A'].notna() & (day_data['assign_A'] != '')) |
                (day_data['assign_B'].notna() & (day_data['assign_B'] != ''))
            ]

            if assigned_rooms_data.empty:
                unique_rooms = []
                unique_departments = []
                dept_colors = {}
                y_positions = {}
            else:
                unique_rooms = sorted(assigned_rooms_data['room'].unique(), key=lambda x: int(x) if str(x).isdigit() else float('inf'), reverse=True)
                unique_departments = sorted(assigned_rooms_data['dept'].unique())
                dept_colors = {dept: colors[i % len(colors)] for i, dept in enumerate(unique_departments)}
                y_positions = {room: i for i, room in enumerate(unique_rooms)}

            duty_rows = []

            # Skip creating chart if no rooms have assignments and no duty assignments
            if len(unique_rooms) == 0 and (not duty_template or not duty_template.get_assignments_for_date(current_date.strftime('%Y-%m-%d'))):
                continue

            # Calculate figure size after processing all duty assignments
            total_rows = len(unique_rooms)
            if duty_template:
                date_iso = current_date.strftime('%Y-%m-%d')
                duty_assignments = duty_template.get_assignments_for_date(date_iso)
                if duty_assignments:
                    temp_duty_rows = []
                    for assignment in duty_assignments:
                        try:
                            start_time_int = assignment['start_time']
                            end_time_int = assignment['end_time']

                            start_hour = start_time_int // 100
                            start_min = start_time_int % 100
                            start_time = start_hour + start_min / 60.0

                            end_hour = end_time_int // 100
                            end_min = end_time_int % 100
                            end_time = end_hour + end_min / 60.0

                            if end_time <= start_time:
                                continue

                            assigned_row = None
                            for row_idx, row_assignments in enumerate(temp_duty_rows):
                                conflict = False
                                for existing_start, existing_end in row_assignments:
                                    if not (end_time <= existing_start or existing_end <= start_time):
                                        conflict = True
                                        break
                                if not conflict:
                                    assigned_row = row_idx
                                    row_assignments.append((start_time, end_time))
                                    break

                            if assigned_row is None:
                                temp_duty_rows.append([(start_time, end_time)])
                        except (ValueError, TypeError):
                            continue
                    total_rows += len(temp_duty_rows)

            fig, ax = plt.subplots(figsize=(12, max(4, total_rows * 0.6)))
            if duty_template:
                date_iso = current_date.strftime('%Y-%m-%d')
                duty_assignments = duty_template.get_assignments_for_date(date_iso)

                for assignment in duty_assignments:
                    try:
                        start_time_int = assignment['start_time']
                        end_time_int = assignment['end_time']

                        start_hour = start_time_int // 100
                        start_min = start_time_int % 100
                        start_time = start_hour + start_min / 60.0

                        end_hour = end_time_int // 100
                        end_min = end_time_int % 100
                        end_time = end_hour + end_min / 60.0

                        duration = end_time - start_time

                        if duration <= 0:
                            continue

                        assigned_row = None
                        for row_idx, row_assignments in enumerate(duty_rows):
                            conflict = False
                            for existing_start, existing_end in row_assignments:
                                if not (end_time <= existing_start or existing_end <= start_time):
                                    conflict = True
                                    break
                            if not conflict:
                                assigned_row = row_idx
                                row_assignments.append((start_time, end_time))
                                break

                        if assigned_row is None:
                            duty_rows.append([(start_time, end_time)])
                            assigned_row = len(duty_rows) - 1

                        duty = assignment['duty']
                        name = assignment['name']

                        y_pos = len(unique_rooms) + assigned_row

                        rect = Rectangle((start_time, y_pos - 0.4), duration, 0.8,
                                       facecolor='#FFD700', edgecolor='black', linewidth=0.5, alpha=0.3)
                        ax.add_patch(rect)

                        label = f'{duty}:{name}'
                        ax.text(start_time + duration/2, y_pos, label,
                               ha='center', va='center', fontsize=9, color='black', weight='bold')

                    except (ValueError, TypeError) as e:
                        print(f"Error processing duty assignment: {e}")
                        continue

            for _, r in assigned_rooms_data.iterrows():
                start_time = int(r["start"])
                end_time = int(r["end"])
                room = r['room']

                start_hour = start_time // 100 + (start_time % 100) / 60
                end_hour = end_time // 100 + (end_time % 100) / 60
                duration = end_hour - start_hour

                y_pos = len(unique_rooms) - 1 - y_positions[room]

                rect = Rectangle((start_hour, y_pos - 0.4), duration, 0.8,
                               facecolor=dept_colors[r['dept']], edgecolor='black', linewidth=0.5, alpha=0.3)
                ax.add_patch(rect)

                bar_height = 0.8
                font_size = 9  # Fixed font size

                room_display = work_data.room_name.get(room, room) if work_data else room
                dept_display = work_data.dept_name.get(r['dept'], r['dept']) if work_data else r['dept']

                line1 = f'ID:{r["id"]} 部署:{dept_display}'
                assign_A_str = str(r["assign_A"]).strip() if r["assign_A"] and str(r["assign_A"]).strip() and str(r["assign_A"]) != 'nan' else ''
                assign_B_str = str(r["assign_B"]).strip() if r["assign_B"] and str(r["assign_B"]).strip() and str(r["assign_B"]) != 'nan' else ''

                line_spacing = bar_height / 4
                ax.text(start_hour + duration/2, y_pos + line_spacing, line1,
                       ha='center', va='center', fontsize=font_size, color='black', weight='bold')

                def _render_assign_line(ax, x, y, prefix, assign_str, fontsize):
                    """Render assignment line with bantane names in red."""
                    if not assign_str:
                        return
                    names = [n.strip() for n in assign_str.split(',')]
                    has_bantane = any(is_bantane_name(n) for n in names)
                    if not has_bantane:
                        ax.text(x, y, f'{prefix}{assign_str}',
                               ha='center', va='center', fontsize=fontsize, color='black', weight='bold')
                    else:
                        # Render prefix + non-bantane names in black, bantane names in red
                        parts = []
                        for n in names:
                            parts.append((n, 'red' if is_bantane_name(n) else 'black'))
                        full_text = f'{prefix}' + ','.join(n for n, _ in parts)
                        # Use centered text with colored segments via individual text calls
                        # First render invisible full text to get centering, then overlay colored parts
                        txt_obj = ax.text(x, y, full_text,
                                         ha='center', va='center', fontsize=fontsize, color='black', weight='bold', alpha=0)
                        renderer = ax.figure.canvas.get_renderer()
                        bbox = txt_obj.get_window_extent(renderer=renderer)
                        inv = ax.transData.inverted()
                        left_data, _ = inv.transform((bbox.x0, bbox.y0))
                        # Render from left
                        cursor_x = left_data
                        segments = [(prefix, 'black')]
                        for i, (n, c) in enumerate(parts):
                            segments.append((n, c))
                            if i < len(parts) - 1:
                                segments.append((',', 'black'))
                        for seg_text, seg_color in segments:
                            t = ax.text(cursor_x, y, seg_text,
                                       ha='left', va='center', fontsize=fontsize, color=seg_color, weight='bold')
                            seg_bbox = t.get_window_extent(renderer=renderer)
                            seg_width_data = inv.transform((seg_bbox.x1, 0))[0] - inv.transform((seg_bbox.x0, 0))[0]
                            cursor_x += seg_width_data

                if assign_A_str:
                    _render_assign_line(ax, start_hour + duration/2, y_pos, '職種1:', assign_A_str, font_size)
                if assign_B_str:
                    _render_assign_line(ax, start_hour + duration/2, y_pos - line_spacing, '職種2:', assign_B_str, font_size)

            ax.set_xlim(6, 20)
            ax.set_xticks(range(6, 21))
            ax.set_xticklabels([f"{h:02d}:00" for h in range(6, 21)])
            ax.set_xlabel('時刻')
            ax.set_ylabel('部屋')

            max_y = len(unique_rooms)
            duty_row_count = len(duty_rows)
            if duty_row_count > 0:
                max_y += duty_row_count

            ax.set_ylim(-0.5, max_y - 0.5)

            y_labels = []
            y_ticks = []

            if duty_row_count > 0:
                for i in range(duty_row_count):
                    y_labels.append(f'業務割当{i+1}' if duty_row_count > 1 else '業務割当')
                    y_ticks.append(len(unique_rooms) + duty_row_count - 1 - i)

            for room in unique_rooms:
                y_labels.append(work_data.room_name.get(room, f'部屋 {room}') if work_data else f'部屋 {room}')
                y_ticks.append(len(unique_rooms) - 1 - y_positions[room])

            ax.set_yticks(y_ticks)
            ax.set_yticklabels(y_labels)

            ax.grid(True, alpha=0.3)
            ax.set_axisbelow(True)

            day_title = f"{current_date.month}/{current_date.day}({weekday_names[current_date.weekday()]})"
            ax.set_title(day_title, fontsize=12, weight='bold', pad=15)

            if available_A_df is not None or available_B_df is not None:
                if isinstance(current_date.day, int):
                    current_date_str = f"{current_date.day:02d}"
                else:
                    current_date_str = str(current_date.day).zfill(2)
                availability_text = []

                if available_A_df is not None and not available_A_df.empty:
                    if available_A_df['date'].dtype in ['int64', 'int32']:
                        day_avail_A = available_A_df[available_A_df['date'] == current_date.day]
                    else:
                        day_avail_A = available_A_df[available_A_df['date'] == current_date_str]
                    if not day_avail_A.empty:
                        row_A = day_avail_A.iloc[0]
                        full_val = str(row_A['full']).strip() if pd.notna(row_A['full']) and str(row_A['full']).strip() else ""
                        am_val = str(row_A['AM']).strip() if pd.notna(row_A['AM']) and str(row_A['AM']).strip() else ""
                        pm_val = str(row_A['PM']).strip() if pd.notna(row_A['PM']) and str(row_A['PM']).strip() else ""

                        if full_val and full_val != 'nan':
                            availability_text.append(f"職種1 終日フリー: {full_val}")
                        if am_val and am_val != 'nan':
                            availability_text.append(f"職種1 AMフリー: {am_val}")
                        if pm_val and pm_val != 'nan':
                            availability_text.append(f"職種1 PMフリー: {pm_val}")

                if available_B_df is not None and not available_B_df.empty:
                    if available_B_df['date'].dtype in ['int64', 'int32']:
                        day_avail_B = available_B_df[available_B_df['date'] == current_date.day]
                    else:
                        day_avail_B = available_B_df[available_B_df['date'] == current_date_str]
                    if not day_avail_B.empty:
                        row_B = day_avail_B.iloc[0]
                        full_val = str(row_B['full']).strip() if pd.notna(row_B['full']) and str(row_B['full']).strip() else ""
                        am_val = str(row_B['AM']).strip() if pd.notna(row_B['AM']) and str(row_B['AM']).strip() else ""
                        pm_val = str(row_B['PM']).strip() if pd.notna(row_B['PM']) and str(row_B['PM']).strip() else ""

                        if full_val and full_val != 'nan':
                            availability_text.append(f"職種2 終日フリー: {full_val}")
                        if am_val and am_val != 'nan':
                            availability_text.append(f"職種2 AMフリー: {am_val}")
                        if pm_val and pm_val != 'nan':
                            availability_text.append(f"職種2 PMフリー: {pm_val}")
                
                if close_assignments:
                    day = current_date.day
                    for duty_type in CLOSE_DUTY_TYPES:
                        if close_assignments.get(duty_type, {}).get(day):
                            symbol = CLOSE_DUTY_SYMBOLS[duty_type]
                            name = CLOSE_DUTY_NAMES[duty_type]
                            availability_text.append(f"{name}{symbol}: {close_assignments[duty_type][day]}")

                if availability_text:
                    availability_display = '\n'.join(availability_text)
                    ax.text(20.5, max_y*0.8, availability_display,
                           ha='left', va='center', fontsize=9, color='black',
                           bbox=dict(boxstyle="round,pad=0.3", facecolor='lightgray', alpha=0.7))

            plt.tight_layout()
            plt.subplots_adjust(right=0.65)

            img_buffer = BytesIO()
            fig.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
            img_buffer.seek(0)
            plt.close(fig)

            img = Image(img_buffer)
            original_width = img.width
            original_height = img.height

            target_width = 800
            if original_width > 0:
                aspect_ratio = original_height / original_width
                img.width = target_width
                img.height = int(target_width * aspect_ratio)
            else:
                img.width = 800
                img.height = 400

            cell_position = f'A{current_row}'
            ws.add_image(img, cell_position)

            current_row += int(img.height / 15) + 2

    desired_order = [
        "個人カレンダー",
        "カレンダー",
        "第1週", "第2週", "第3週", "第4週", "第5週", "第6週",
        "フリー(A)",
        "フリー(B)",
        "勤務統計",
        "最適化情報",
        "割当情報"
    ]
    
    order_map = {name: i for i, name in enumerate(desired_order)}
    wb._sheets.sort(key=lambda ws: order_map.get(ws.title, len(order_map)))

    wb.save(output_path)


class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setting_path: Optional[Path] = None
        self.kintai_A_path: Optional[Path] = None
        self.kintai_B_path: Optional[Path] = None
        self.duty_path: Optional[Path] = None
        self.visualization_input_path: Optional[Path] = None
        self.output_path: Optional[Path] = None
        self.kinmu_path: Optional[Path] = None

        self.setAcceptDrops(True)

        self.window_width = 800
        self.window_height = 900
        
        self.base_font_size = max(10, int(self.window_height / 60))
        self.current_font_size = self.base_font_size
        self.small_font_size = max(8, int(self.base_font_size * 0.7))
        self.path_font_size = max(8, int(self.base_font_size * 0.8))

        self._auto_load_files()

        self.log = QTextEdit()
        self.log.setReadOnly(True)

        self.assign_df: Optional[pd.DataFrame] = None
        self.available_df: Optional[pd.DataFrame] = None
        self.stats_df: Optional[pd.DataFrame] = None
        self.violation_df: Optional[pd.DataFrame] = None

        self.tabs = QTabWidget()

        self.input_field_width = 200

        self._build_ui()

    def _auto_load_files(self):
        """Automatically load files from the files subdirectory based on naming convention"""
        self._refresh_paths_from_files_dir(force=True)
    
    def _refresh_paths_from_files_dir(self, force: bool = False):
        """Refresh file paths from the files and temporary subdirectories
        
        Args:
            force: If True, overwrite existing paths. If False, only fill in None paths.
        """
        app_dir = get_app_dir()
        files_dir = app_dir / "files"
        temporary_dir = app_dir / "temporary"
        
        # Files to search in 'files' folder
        files_mappings = {
            'setting': 'setting_path',
            'manual': 'visualization_input_path',
        }
        
        # Files to search in 'temporary' folder (generated intermediate files)
        temporary_mappings = {
            'kintaiA': 'kintai_A_path',
            'kintaiB': 'kintai_B_path',
            'duty': 'duty_path',
            'generated_shift_A': 'generated_shift_A_path',
            'generated_shift_B': 'generated_shift_B_path',
        }
        
        # Search in files folder
        if files_dir.exists():
            for xlsx_file in files_dir.glob("*.xlsx"):
                file_name = xlsx_file.name.lower()
                for prefix, attr_name in files_mappings.items():
                    if prefix.lower() in file_name:
                        current_value = getattr(self, attr_name, None)
                        if force or current_value is None:
                            setattr(self, attr_name, xlsx_file)
                        break
        
        # Search in output folder for output*.xlsx (for visualization)
        output_dir = app_dir / "output"
        if output_dir.exists():
            for xlsx_file in output_dir.glob("output*.xlsx"):
                current_value = getattr(self, 'visualization_input_path', None)
                if force or current_value is None:
                    self.visualization_input_path = xlsx_file
                    break
        
        # Search in temporary folder
        if temporary_dir.exists():
            for xlsx_file in temporary_dir.glob("*.xlsx"):
                file_name = xlsx_file.name.lower()
                for prefix, attr_name in temporary_mappings.items():
                    if prefix.lower() in file_name:
                        current_value = getattr(self, attr_name, None)
                        if force or current_value is None:
                            setattr(self, attr_name, xlsx_file)
                        break

    def _compute_target_period(self) -> Tuple[int, int]:
        """Compute the target year and month based on button state

        If an override period was detected from the input file, use that instead.

        If button is checked (current period mode / 同月修正モード):
            Returns the period that contains today (21st to 20th of next month).
            Example: If today is Dec 13, we're in period 11/21-12/20, so return (2025, 11)
            Example: If today is Dec 25, we're in period 12/21-1/20, so return (2025, 12)
        
        If button is unchecked (next period mode - default):
            Returns the period that starts on the next 21st after today.
            Example: If today is Dec 13, next 21st is Dec 21, so return (2025, 12) for period 12/21-1/20
            Example: If today is Dec 25, next 21st is Jan 21, so return (2026, 1) for period 1/21-2/20
        """
        # 手動指定の期間がある場合は優先
        manual = getattr(self, '_manual_period', None)
        if manual is not None:
            return manual

        # 入力ファイルから検出された期間がある場合はそれを優先
        override = getattr(self, '_override_period', None)
        if override is not None:
            return override

        today = dt.date.today()
        
        # Check if current period mode is enabled (同月修正モード)
        use_current_period = getattr(self, 'current_period_button', None) and self.current_period_button.isChecked()
        
        if use_current_period:
            # Current period mode: return the period containing today
            if today.day >= 21:
                # We're in the period that started on the 21st of this month
                return today.year, today.month
            else:
                # We're in the period that started on the 21st of last month
                if today.month == 1:
                    return today.year - 1, 12
                else:
                    return today.year, today.month - 1
        else:
            # Next period mode (default): return the period starting on the next 21st
            if today.day < 21:
                # Next 21st is this month
                return today.year, today.month
            else:
                # Next 21st is next month
                if today.month == 12:
                    return today.year + 1, 1
                else:
                    return today.year, today.month + 1

    def _get_default_output_path(self, mode: str = "opt") -> Path:
        """Get the default output path in output subdirectory with timestamp filename
        
        Args:
            mode: "opt" for optimization output, "vis" for visualization output
        """
        now = dt.datetime.now()
        ts = now.strftime("%Y%m%d%H%M")
        filename = f"output_{mode}_{ts}.xlsx"
        
        app_dir = get_app_dir()
        output_dir = app_dir / "output"
        output_dir.mkdir(parents=True, exist_ok=True)
        
        return output_dir / filename

    def _set_progress(self, value: int, message: str):
        """Update progress bar and log"""
        self.progress_bar.setValue(value)
        self.progress_label.setText(message)
        self.append_log(message)
        QApplication.processEvents()

    def _solver_status_to_japanese(self, status: int) -> str:
        """Convert PuLP solver status to Japanese description"""
        status_map = {
            pulp.LpStatusOptimal: "最適解が見つかりました",
            pulp.LpStatusInfeasible: "実行不能（制約を満たす解が存在しません）",
            pulp.LpStatusUnbounded: "解が無限大です",
            pulp.LpStatusNotSolved: "解が見つかりませんでした",
            pulp.LpStatusUndefined: "モデルが無効です",
        }
        return status_map.get(status, f"不明なステータス ({status})")

    def _build_ui(self):
        self.setWindowTitle("ばんたね病院外来シフト最適化")
        
        v = QVBoxLayout(self)

        button_style = """
            QPushButton {
                border: 2px solid #8f8f91;
                border-radius: 6px;
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #f6f7fa, stop: 1 #dadbde);
                min-width: 120px;
                padding: 8px;
            }
            QPushButton:hover {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #e7e8eb, stop: 1 #c8c9cc);
            }
            QPushButton:pressed {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #dadbde, stop: 1 #f6f7fa);
            }
        """

        progress_row= QHBoxLayout()
        self.progress_bar = QProgressBar()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFormat("%p% - %v/100")
        progress_row.addWidget(self.progress_bar)
        v.addLayout(progress_row)
        
        self.progress_label = QLabel("")
        self.progress_label.setFont(QFont("", self.current_font_size))
        v.addWidget(self.progress_label)

        self.tabs.addTab(self.log, "ログ")
        tab_font = QFont("", self.current_font_size)
        self.tabs.setFont(tab_font)
        v.addWidget(self.tabs, 1)
        
        # トグルボタン用のスタイル（オン状態）
        toggle_button_style_on = """
            QPushButton {
                border: 2px solid #4a90d9;
                border-radius: 6px;
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #5da5e8, stop: 1 #3a7fc4);
                color: white;
                min-width: 120px;
                padding: 8px;
            }
            QPushButton:hover {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #6db5f8, stop: 1 #4a8fd4);
            }
        """
        
        # ボタンをウィンドウ下端に配置（左：同月修正モード、右：ログ参照）
        bottom_row = QHBoxLayout()
        bottom_row.addStretch(1)
        
        # 同月修正モードトグルボタン
        self.current_period_button = QPushButton("同月修正モード")
        self.current_period_button.setFont(QFont("", self.current_font_size))
        self.current_period_button.setCheckable(True)  # トグルボタンとして動作
        self.current_period_button.setChecked(False)  # デフォルトはオフ（翌月版）
        self.current_period_button.setMinimumSize(int(self.current_font_size * 8), int(self.current_font_size * 2.5))
        self.current_period_button.setStyleSheet(button_style)
        self.current_period_button.toggled.connect(self._on_current_period_toggled)
        bottom_row.addWidget(self.current_period_button)
        
        # スタイルを保存しておく
        self._button_style_normal = button_style
        self._button_style_on = toggle_button_style_on

        # 期間指定ボタン
        self._manual_period = None  # (year, month) or None
        btn_period = QPushButton("期間指定")
        btn_period.setFont(QFont("", max(8, self.current_font_size - 2)))
        btn_period.clicked.connect(self._on_select_period)
        btn_period.setMinimumSize(int(self.current_font_size * 5), int(self.current_font_size * 2))
        btn_period.setStyleSheet(button_style)
        btn_period.setToolTip("対象期間を手動指定します（テスト用）")
        bottom_row.addWidget(btn_period)
        self._btn_period = btn_period

        bottom_row.addStretch(1)
        
        btn_log = QPushButton("ログ参照")
        btn_log.setFont(QFont("", self.current_font_size))
        btn_log.clicked.connect(self.open_log_file)
        btn_log.setMinimumSize(int(self.current_font_size * 6), int(self.current_font_size * 2.5))
        btn_log.setStyleSheet(button_style)
        
        bottom_row.addWidget(btn_log)
        bottom_row.addStretch(1)
        v.addLayout(bottom_row)
    
    def _on_current_period_toggled(self, checked: bool):
        """Handle toggle of current period mode button"""
        if checked:
            self.current_period_button.setStyleSheet(self._button_style_on)
            self.append_log("同月修正モード: オン（現在の期間を対象）")
        else:
            self.current_period_button.setStyleSheet(self._button_style_normal)
            self.append_log("同月修正モード: オフ（翌月の期間を対象）")

    def _on_select_period(self):
        """期間指定ダイアログ表示"""
        today = dt.date.today()
        current_year = today.year

        from PySide6.QtWidgets import QDialog, QFormLayout, QDialogButtonBox

        dlg = QDialog(self)
        dlg.setWindowTitle("対象期間の設定")
        layout = QFormLayout(dlg)

        year_spin = QSpinBox()
        year_spin.setRange(2020, 2040)
        year_spin.setValue(current_year)
        layout.addRow("年", year_spin)

        month_spin = QSpinBox()
        month_spin.setRange(1, 12)
        month_spin.setValue(today.month)
        layout.addRow("月", month_spin)

        note = QLabel("※ 指定した月の21日〜翌月20日が対象期間になります\n※「自動」を押すと入力ファイルからの自動検出に戻ります")
        note.setFont(QFont("", max(8, self.current_font_size - 2)))
        layout.addRow(note)

        buttons = QDialogButtonBox()
        btn_ok = buttons.addButton("設定", QDialogButtonBox.AcceptRole)
        btn_auto = buttons.addButton("自動", QDialogButtonBox.ResetRole)
        btn_cancel = buttons.addButton("キャンセル", QDialogButtonBox.RejectRole)
        layout.addRow(buttons)

        result = {"action": "cancel"}

        def on_ok():
            result["action"] = "set"
            dlg.accept()

        def on_auto():
            result["action"] = "auto"
            dlg.accept()

        btn_ok.clicked.connect(on_ok)
        btn_auto.clicked.connect(on_auto)
        btn_cancel.clicked.connect(dlg.reject)

        if dlg.exec() == QDialog.Accepted:
            if result["action"] == "set":
                self._manual_period = (year_spin.value(), month_spin.value())
                self._btn_period.setText(f"期間: {year_spin.value()}/{month_spin.value()}")
                self._btn_period.setStyleSheet(self._button_style_on)
                self.append_log(f"対象期間を手動設定: {year_spin.value()}年{month_spin.value()}月")
            elif result["action"] == "auto":
                self._manual_period = None
                self._btn_period.setText("期間指定")
                self._btn_period.setStyleSheet(self._button_style_normal)
                self.append_log("対象期間を自動検出に戻しました")

    def _update_font_sizes(self):
        self.base_font_size = max(10, int(self.window_height / 60))
        self.current_font_size = self.base_font_size
        self.small_font_size = max(8, int(self.base_font_size * 0.7))
        self.path_font_size = max(8, int(self.base_font_size * 0.8))

    def append_log(self, s: str):
        self.log.append(s)

    def open_log_file(self):
        log_file = Path("shift_app.log")
        if not log_file.exists():
            QMessageBox.information(self, "情報", "ログファイル (shift_app.log) がまだ作成されていません。\n\nアプリを実行すると、ログファイルが作成されます。")
            return
        
        try:
            import subprocess
            import platform
            import os
            import shutil
            import re
            
            system = platform.system()
            is_linux = (system == "Linux")
            
            in_wsl = False
            if is_linux:
                if "WSL_DISTRO_NAME" in os.environ or "WSL_INTEROP" in os.environ:
                    in_wsl = True
                elif os.path.exists("/proc/version"):
                    try:
                        with open("/proc/version", "r") as f:
                            version_info = f.read().lower()
                            if "microsoft" in version_info or "wsl" in version_info:
                                in_wsl = True
                    except:
                        pass
            
            if in_wsl:
                abs_path = str(log_file.absolute())
                
                if shutil.which("wslview"):
                    try:
                        subprocess.run(["wslview", abs_path], check=False)
                        return
                    except:
                        pass
                
                win_path = None
                
                if shutil.which("wslpath"):
                    try:
                        result = subprocess.run(["wslpath", "-w", abs_path], 
                                              capture_output=True, text=True, check=True)
                        win_path = result.stdout.strip()
                    except:
                        pass
                
                if not win_path:
                    match = re.match(r'^/mnt/([a-zA-Z])/(.*)$', abs_path)
                    if match:
                        drive = match.group(1).upper()
                        rest = match.group(2).replace('/', '\\')
                        win_path = f"{drive}:\\{rest}"
                
                if win_path:
                    if shutil.which("explorer.exe"):
                        try:
                            subprocess.run(["explorer.exe", win_path], check=False)
                            return
                        except:
                            pass
                    
                    if shutil.which("powershell.exe"):
                        try:
                            subprocess.run(["powershell.exe", "-NoProfile", "-Command", 
                                          f'Start-Process -FilePath "{win_path}"'], check=False)
                            return
                        except:
                            pass
                    
                    if shutil.which("notepad.exe"):
                        try:
                            subprocess.run(["notepad.exe", win_path], check=False)
                            return
                        except:
                            pass
                    
                    QMessageBox.information(self, "情報", 
                        f"ログファイルを自動で開けませんでした。\n\n"
                        f"以下のパスのファイルを手動で開いてください：\n\n{win_path}")
                    return
                else:
                    QMessageBox.warning(self, "エラー", 
                        f"Windowsパスへの変換に失敗しました。\n\n"
                        f"WSLパス: {abs_path}\n\n手動でファイルを開いてください。")
                    return
            
            if system == "Windows":
                import os
                os.startfile(str(log_file))
                return
            
            if system == "Darwin":
                subprocess.run(["open", str(log_file)], check=False)
                return
            
            if is_linux:
                if shutil.which("xdg-open"):
                    subprocess.run(["xdg-open", str(log_file)], check=False)
                    return
                else:
                    QMessageBox.information(self, "情報", 
                        f"ファイルを自動で開くコマンドが見つかりませんでした。\n\n"
                        f"以下のパスのファイルを手動で開いてください：\n\n{log_file.absolute()}")
                    return
            
        except Exception as e:
            QMessageBox.warning(self, "エラー", 
                f"ログファイルを開けませんでした。\n\n"
                f"ファイルパス: {log_file.absolute()}\n\n"
                f"エラー: {str(e)}\n\n手動でファイルを開いてください。")

    def run_assign(self):
        """Run assignment mode"""
        try:
            self._set_progress(0, "割当実行を開始します...")
            app_dir = get_app_dir()
            files_dir = app_dir / "files"
            
            missing_files = []
            if not self.setting_path:
                missing_files.append("setting*.xlsx (基本設定)")
            if not self.duty_path:
                missing_files.append("duty*.xlsx (個人業務)")
            
            if missing_files:
                msg = "割り当てモードに必要なファイルが見つかりません。\n\n"
                msg += f"filesフォルダ: {files_dir}\n\n"
                msg += "不足ファイル:\n"
                for f in missing_files:
                    msg += f"  - {f}\n"
                msg += "\nfilesフォルダに上記のファイルを配置してください。"
                msg += "\n\n※ kintaiA*.xlsx, kintaiB*.xlsx はオプションです。"
                QMessageBox.warning(self, "ファイル不足", msg)
                self._set_progress(0, "")
                return

            target_year, target_month = self._compute_target_period()
            self.output_path = self._get_default_output_path("opt")
            
            self._set_progress(10, "入力ファイルを検証中...")
            all_issues = []
            
            errors = [issue for issue in all_issues if issue.severity == "ERROR"]
            warnings = [issue for issue in all_issues if issue.severity == "WARNING"]
            
            if errors:
                error_summary = f"入力ファイルに{len(errors)}個のエラーが見つかりました。\n\n"
                error_summary += "最初の5個のエラー:\n\n"
                for i, issue in enumerate(errors[:5]):
                    error_summary += f"【エラー {i+1}】\n"
                    error_summary += f"ファイル: {issue.file_name}\n"
                    error_summary += f"シート: {issue.sheet}\n"
                    error_summary += f"場所: {issue.location}\n"
                    error_summary += f"問題: {issue.field} - {issue.value}\n"
                    error_summary += f"対処: {issue.hint}\n\n"
                
                if len(errors) > 5:
                    error_summary += f"\n...他{len(errors)-5}個のエラーがあります。\n"
                
                error_summary += "\n入力ファイルを修正してから再度実行してください。"
                
                for issue in all_issues:
                    logging.error(issue.to_string())
                
                QMessageBox.critical(self, "入力エラー", error_summary)
                self._set_progress(0, f"入力検証エラー: {len(errors)}個のエラーが見つかりました。")
                return
            
            if warnings:
                self.append_log(f"警告: {len(warnings)}個の警告が見つかりました。")
                for warning in warnings:
                    logging.warning(warning.to_string())

            self._set_progress(20, "データを読み込み中...")
            setting = Setting(self.setting_path)
            set_reserve_names(setting.get_reserve_names())
            work = WorkData.from_setting(setting, target_year, target_month)
            # generated_shift_A/Bはtemporaryフォルダに出力
            app_dir = get_app_dir()
            temporary_dir = app_dir / "temporary"
            temporary_dir.mkdir(parents=True, exist_ok=True)
            _UNREGISTERED_STAFF.clear()
            jobA = JobData.from_setting(setting, "A", target_year, target_month,
                         kintai_path=self.kintai_A_path if hasattr(self, 'kintai_A_path') else None,
                         output_dir=temporary_dir)
            jobB = JobData.from_setting(setting, "B", target_year, target_month,
                         kintai_path=self.kintai_B_path if hasattr(self, 'kintai_B_path') else None,
                         output_dir=temporary_dir)
            if _UNREGISTERED_STAFF:
                names = '、'.join(dict.fromkeys(_UNREGISTERED_STAFF))
                QMessageBox.warning(
                    self, "未登録スタッフ",
                    f"{names}さんの勤務条件が設定されていません。\n"
                    "setting.xlsxのjobA/jobB・dictA/dictB・orderA/orderBシートを\n"
                    "更新してアプリを再起動してください。"
                )
                _UNREGISTERED_STAFF.clear()
                return
            duty_template = DutyTemplate(self.duty_path, target_year, target_month) if self.duty_path else None

            self._set_progress(30, "最適化モデルを構築中...")
            opt = Optimizer(setting, work, jobA, jobB, duty_template)
            logging.info(f"mode=実行, setting={self.setting_path}, output={self.output_path}")
            
            self._set_progress(40, "最適化計算中（5回繰り返し）...")
            iterations = 5
            res = opt.build_and_solve(max_iterations=iterations)
            status_text = self._solver_status_to_japanese(res)
            self._set_progress(60, f"ソルバー結果: {status_text}")
            logging.info(f"solver status: {res} ({status_text}), objective={pulp.value(opt.model.objective) if res == Optimizer.OPTIMAL else 'N/A'}")

            if res != Optimizer.OPTIMAL:
                violations = getattr(opt, "violations", [])
                
                msg = "【実行不能エラー (E311)】\n\n"
                msg += "割り当てが成立しませんでした。以下の理由が考えられます：\n\n"
                
                if violations:
                    msg += "【主な問題】\n"
                    for i, violation in enumerate(violations[:5], 1):
                        msg += f"{i}. {violation}\n"
                    
                    if len(violations) > 5:
                        msg += f"\n...他{len(violations)-5}件の問題があります。\n"
                    
                    msg += "\n【対処方法】\n"
                    msg += "1. 該当日の必要人数（need_A/need_B）を見直す\n"
                    msg += "   → 業務一覧ファイル(work.xlsx)のneed_A/need_B列を確認\n\n"
                    msg += "2. メンバーのsenior設定を確認する\n"
                    msg += "   → 職種設定ファイル(jobA/jobB.xlsx)のmembersシートでsenior=1を追加\n\n"
                    msg += "3. 該当日に可用なスタッフを追加する\n"
                    msg += "   → 職種設定ファイル(jobA/jobB.xlsx)のshiftシートでstaffを追加\n\n"
                    msg += "4. ペナルティ設定を緩和する\n"
                    msg += "   → 基本設定ファイル(setting.xlsx)のexceptionシートで制約を緩和\n\n"
                else:
                    msg += "制約条件が厳しすぎる可能性があります。\n\n"
                    msg += "【対処方法】\n"
                    msg += "1. 各日の必要人数と可用人数のバランスを確認\n"
                    msg += "2. シニアスタッフの人数が十分か確認\n"
                    msg += "3. 設定ファイルの制約条件を見直す\n\n"
                
                msg += "詳細なログは shift_app.log ファイルに記録されています。"
                
                self._set_progress(0, "実行不能: 割り当てが成立しませんでした。")
                for violation in violations:
                    logging.error(f"Violation: {violation}")
                
                QMessageBox.critical(self, "実行不能エラー", msg)
                return

            self._set_progress(65, "ダミー割当を最適化中...")
            opt.post_process_reduce_dummy()
            self._set_progress(70, "結果を抽出中...")
            assign_df, available_A_df, available_B_df, stats_df, violation_df = opt.extract_output()
            
            for col in ['full', 'AM', 'PM']:
                if col in available_A_df.columns:
                    available_A_df[col] = available_A_df[col].apply(lambda x: filter_bantane_from_text(x, ","))
                if col in available_B_df.columns:
                    available_B_df[col] = available_B_df[col].apply(lambda x: filter_bantane_from_text(x, ","))
            
            if 'id' in stats_df.columns:
                import re
                work_name_map = {}
                if setting.dict_work_df is not None and not setting.dict_work_df.empty:
                    for _, row in setting.dict_work_df.iterrows():
                        aka = str(row.get('aka', '')).strip()
                        name = str(row.get('name', '')).strip()
                        if aka and name:
                            work_name_map[aka] = name
                
                work_id_to_dept = {}
                if work is not None and hasattr(work, 'work') and not work.work.empty:
                    for _, row in work.work.iterrows():
                        work_id = str(int(row['id']))
                        dept = str(row.get('dept', '')).strip()
                        if work_id and dept:
                            work_id_to_dept[work_id] = dept
                
                def convert_id_to_name(x):
                    if pd.isna(x):
                        return x
                    x_str = str(x)
                    converted_parts = []
                    for part in x_str.split(','):
                        part = part.strip()
                        if not part:
                            continue
                        date_prefix_match = re.match(r'D\d{8}_(.+)', part)
                        if date_prefix_match:
                            converted_parts.append(date_prefix_match.group(1))
                        elif part.isdigit():
                            dept = work_id_to_dept.get(part, '')
                            if dept:
                                converted_parts.append(work_name_map.get(dept, dept))
                            else:
                                converted_parts.append(part)
                        else:
                            converted_parts.append(work_name_map.get(part, part))
                    return ','.join(converted_parts)
                
                stats_df['id'] = stats_df['id'].apply(convert_id_to_name)
                
                def get_unique_elements(x):
                    if pd.isna(x):
                        return x
                    parts = [p.strip() for p in str(x).split(',') if p.strip()]
                    seen = set()
                    unique_parts = []
                    for p in parts:
                        if p not in seen:
                            seen.add(p)
                            unique_parts.append(p)
                    return ','.join(unique_parts)
                
                stats_df['id'] = stats_df['id'].apply(get_unique_elements)
            
            if 'score' in violation_df.columns:
                violation_df = violation_df[violation_df['score'] != 0].copy()
            
            self.assign_df = assign_df
            self.available_A_df = available_A_df
            self.available_B_df = available_B_df
            self.stats_df = stats_df
            self.violation_df = violation_df

            self._set_progress(80, "カレンダービューを生成中...")
            close_assignments = assign_close_duties(jobA, target_year, target_month, assign_df, duty_template)
            calendar_df, all_dates, dept_colors = generate_calendar_view(assign_df, target_year, target_month, work, jobA, jobB, duty_template, close_assignments)
            work_calendar_df, work_all_dates, work_all_names = generate_work_oriented_calendar_view(assign_df, target_year, target_month, work, jobA, jobB, duty_template, close_assignments, available_A_df, available_B_df)
            
            self._set_progress(85, "Excelファイルを出力中...")
            
            assign_df = assign_df.rename(columns={
                'id': '業務ID', 'date': '日付', 'day': '曜日', 'room': '部屋', 'dept': '業務',
                'start': '開始時刻', 'end': '終了時刻', 'assign_A': '割当(A)', 'assign_B': '割当(B)'
            })
            available_A_df = available_A_df.rename(columns={
                'date': '日付', 'day': '曜日', 'full': '終日フリー', 'AM': 'AMフリー', 'PM': 'PMフリー'
            })
            available_B_df = available_B_df.rename(columns={
                'date': '日付', 'day': '曜日', 'full': '終日フリー', 'AM': 'AMフリー', 'PM': 'PMフリー'
            })
            stats_df = stats_df.rename(columns={
                'member': 'member', 'times': '合計回数', 'id': '業務内容', 'duty_count': '個人業務', 'opt_count': '最適化割当'
            })
            stats_df = stats_df[['member', '合計回数', '個人業務', '最適化割当', '業務内容']]
            violation_df = violation_df.rename(columns={
                'id': '業務ID', 'penalty': '制約', 'score': '点数', 'description': '詳細'
            })
            
            # Create duty DataFrame from duty_template for output (with Japanese column names)
            duty_df = pd.DataFrame()
            if duty_template is not None and duty_template.assignments:
                duty_records = []
                for assignment in duty_template.assignments:
                    duty_records.append({
                        '個人業務': assignment['duty'],
                        'いつから': assignment['day'],
                        'いつまで': assignment['day'],
                        '開始時刻': assignment['start_time'],
                        '終了時刻': assignment['end_time'],
                        '担当': assignment['name']
                    })
                duty_df = pd.DataFrame(duty_records)
            
            # Read kintaiA data for 勤怠 sheet
            kintai_df_for_output = pd.DataFrame()
            if hasattr(self, 'kintai_A_path') and self.kintai_A_path and Path(self.kintai_A_path).exists():
                try:
                    kintai_xl = pd.read_excel(self.kintai_A_path, sheet_name=None)
                    kintai_df_for_output = kintai_xl[list(kintai_xl.keys())[0]].copy()
                except Exception as e:
                    logging.warning(f"kintaiAファイルの読み込みに失敗しました: {e}")
            
            with pd.ExcelWriter(self.output_path, engine="openpyxl") as w:
                assign_df.to_excel(w, sheet_name="割当情報", index=False)
                available_A_df.to_excel(w, sheet_name="フリー(A)", index=False)
                available_B_df.to_excel(w, sheet_name="フリー(B)", index=False)
                stats_df.to_excel(w, sheet_name="勤務統計", index=False)
                violation_df.to_excel(w, sheet_name="最適化情報", index=False)
                
                auto_adjust_column_width(w.book["割当情報"])
                format_assign_sheet(w.book["割当情報"])
                
                auto_adjust_column_width(w.book["フリー(A)"])
                format_available_sheet(w.book["フリー(A)"])
                
                auto_adjust_column_width(w.book["フリー(B)"])
                format_available_sheet(w.book["フリー(B)"])
                
                auto_adjust_column_width(w.book["勤務統計"])
                format_stats_sheet(w.book["勤務統計"])
                
                auto_adjust_column_width(w.book["最適化情報"])
                format_violation_sheet(w.book["最適化情報"])
                
                calendar_df.to_excel(w, sheet_name="個人カレンダー", index=False, header=False)
                ws = w.book["個人カレンダー"]
                apply_calendar_formatting(ws, all_dates, target_year, target_month, dept_colors)
                work_calendar_df.to_excel(w, sheet_name="カレンダー", index=False, header=False)
                ws2 = w.book["カレンダー"]
                apply_work_oriented_calendar_formatting(ws2, work_all_dates, work_all_names, target_year, target_month)
                
                # Add 個人業務 sheet with duty data
                if not duty_df.empty:
                    duty_df.to_excel(w, sheet_name="個人業務", index=False)
                    auto_adjust_column_width(w.book["個人業務"])
                    format_duty_sheet(w.book["個人業務"])
                
                # Add 勤怠 sheet as the last sheet
                if not kintai_df_for_output.empty:
                    kintai_df_for_output.to_excel(w, sheet_name="勤怠", index=False)
                    auto_adjust_column_width(w.book["勤怠"])

            self._set_progress(90, "ガントチャートを生成中...")
            if not self.assign_df.empty:
                generate_weekly_gantt_images(self.assign_df, self.output_path, target_year, target_month, work, duty_template, self.available_A_df, self.available_B_df, close_assignments)

            self._set_progress(100, f"完了: {self.output_path} に出力しました。")
            
            # Clean up temporary files after successful output generation
            try:
                if temporary_dir.exists():
                    for temp_file in temporary_dir.glob('*'):
                        if temp_file.is_file():
                            temp_file.unlink()
                            logging.info(f"一時ファイルを削除しました: {temp_file}")
            except Exception as e:
                logging.warning(f"一時ファイルの削除に失敗しました: {e}")
            
            completion_msg = f"最適化が完了しました。\n\n出力先: {self.output_path}"
            if getattr(opt, 'dummy_staff_used', False):
                self.append_log("注意：ダミースタッフが割当されました")
                completion_msg += "\n\n注意：ダミースタッフが割当されました"
            
            QMessageBox.information(self, "完了", completion_msg)

        except PermissionError as e:
            msg = "【ファイルアクセスエラー (E401)】\n\n"
            msg += "ファイルへのアクセスが拒否されました。\n\n"
            msg += "【対処方法】\n"
            msg += "1. 出力先ファイルが他のプログラム（Excelなど）で開かれている場合は閉じてください\n"
            msg += "2. ファイルが読み取り専用になっていないか確認してください\n"
            msg += "3. 別の保存先を選択してください\n\n"
            msg += f"エラー詳細: {str(e)}"
            logging.error(f"PermissionError: {str(e)}")
            self._set_progress(0, "エラー発生")
            QMessageBox.critical(self, "ファイルアクセスエラー", msg)
        
        except OSError as e:
            msg = "【ファイル操作エラー (E402)】\n\n"
            msg += "ファイルの読み書き中にエラーが発生しました。\n\n"
            msg += "【対処方法】\n"
            msg += "1. ファイルが他のプログラムで開かれている場合は閉じてください\n"
            msg += "2. ディスクの空き容量を確認してください\n"
            msg += "3. ファイルパスに使用できない文字が含まれていないか確認してください\n"
            msg += "4. 別の保存先を選択してください\n\n"
            msg += f"エラー詳細: {str(e)}"
            logging.error(f"OSError: {str(e)}")
            self._set_progress(0, "エラー発生")
            QMessageBox.critical(self, "ファイル操作エラー", msg)
        
        except ImportError as e:
            msg = "【モジュール不足エラー (E403)】\n\n"
            msg += "必要なPythonモジュールが見つかりません。\n\n"
            msg += "【対処方法】\n"
            msg += "コマンドプロンプトまたはターミナルで以下のコマンドを実行してください：\n\n"
            msg += "pip install -r requirements.txt\n\n"
            msg += "または、不足しているモジュールを個別にインストール：\n"
            msg += "pip install pandas openpyxl ortools PyQt6 matplotlib\n\n"
            msg += f"エラー詳細: {str(e)}"
            logging.error(f"ImportError: {str(e)}")
            self._set_progress(0, "エラー発生")
            QMessageBox.critical(self, "モジュール不足エラー", msg)
        
        except ValueError as e:
            msg = "【データ形式エラー (E201)】\n\n"
            msg += "入力データの形式が正しくありません。\n\n"
            msg += "【対処方法】\n"
            msg += "1. 日付の形式を確認してください（dd または yyyy-mm-dd）\n"
            msg += "2. 数値項目に文字が入っていないか確認してください\n"
            msg += "3. 必須項目が空欄になっていないか確認してください\n\n"
            msg += f"エラー詳細: {str(e)}\n\n"
            msg += "詳細なログは shift_app.log ファイルに記録されています。"
            logging.error(f"ValueError: {str(e)}")
            traceback.print_exc()
            self._set_progress(0, "エラー発生")
            QMessageBox.critical(self, "データ形式エラー", msg)
        
        except KeyError as e:
            msg = "【データ不足エラー (E104)】\n\n"
            msg += "必要な列またはデータが見つかりません。\n\n"
            msg += "【対処方法】\n"
            msg += "1. 入力ファイルに必要な列がすべて存在するか確認してください\n"
            msg += "2. 列名のスペルが正しいか確認してください\n"
            msg += "3. テンプレートファイルと比較して、列が不足していないか確認してください\n\n"
            msg += f"不足している項目: {str(e)}\n\n"
            msg += "詳細なログは shift_app.log ファイルに記録されています。"
            logging.error(f"KeyError: {str(e)}")
            traceback.print_exc()
            self._set_progress(0, "エラー発生")
            QMessageBox.critical(self, "データ不足エラー", msg)
        
        except Exception as e:
            msg = "【予期しないエラー (E999)】\n\n"
            msg += "予期しないエラーが発生しました。\n\n"
            msg += f"エラー詳細: {str(e)}\n\n"
            msg += "【対処方法】\n"
            msg += "1. shift_app.log ファイルで詳細なエラー情報を確認してください\n"
            msg += "2. 入力ファイルの内容を確認してください\n"
            msg += "3. 問題が解決しない場合は、開発者に shift_app.log を共有してください\n"
            logging.error(f"Unexpected error: {str(e)}")
            traceback.print_exc()
            self._set_progress(0, "エラー発生")
            QMessageBox.critical(self, "予期しないエラー", msg)
        finally:
            # 入力ファイル由来の期間オーバーライドを必ずクリア
            self._override_period = None

    def run_visualize(self):
        """Run visualization only mode"""
        try:
            self._set_progress(0, "可視化のみを開始します...")
            app_dir = get_app_dir()
            files_dir = app_dir / "files"
            temporary_dir = app_dir / "temporary"
            
            # Refresh file paths from files and temporary directories (in case files were added after startup)
            self._refresh_paths_from_files_dir(force=False)
            
            # Check required files (duty is no longer required - will be read from output file)
            missing_files = []
            if not self.setting_path or not self.setting_path.exists():
                missing_files.append("setting*.xlsx (基本設定) → filesフォルダ")
            if not self.visualization_input_path or not self.visualization_input_path.exists():
                missing_files.append("manual*.xlsx → filesフォルダ、または output*.xlsx → outputフォルダ")
            
            if missing_files:
                msg = "可視化モードに必要なファイルが見つかりません。\n\n"
                msg += f"filesフォルダ: {files_dir}\n"
                msg += f"temporaryフォルダ: {temporary_dir}\n\n"
                msg += "不足ファイル:\n"
                for f in missing_files:
                    msg += f"  - {f}\n"
                msg += "\n上記のファイルを適切なフォルダに配置してください。"
                QMessageBox.warning(self, "ファイル不足", msg)
                self._set_progress(0, "")
                return

            self.assign_df = pd.DataFrame()
            self.available_A_df = pd.DataFrame()
            self.available_B_df = pd.DataFrame()
            self.stats_df = pd.DataFrame()
            self.violation_df = pd.DataFrame()
            duty_df_from_output = pd.DataFrame()

            target_year, target_month = self._compute_target_period()
            self.output_path = self._get_default_output_path("vis")

            self._set_progress(10, "可視化入力データを読み込み中...")
            kintai_df_from_output = pd.DataFrame()  # For fallback when kintai files don't exist
            if self.visualization_input_path:
                xl = pd.read_excel(self.visualization_input_path, sheet_name=None)
                # Support both English and Japanese sheet names
                self.assign_df = xl.get("割当情報", xl.get("assign", pd.DataFrame()))
                self.available_A_df = xl.get("フリー(A)", xl.get("available_A", pd.DataFrame()))
                self.available_B_df = xl.get("フリー(B)", xl.get("available_B", pd.DataFrame()))
                self.stats_df = xl.get("勤務統計", xl.get("stats", pd.DataFrame()))
                self.violation_df = xl.get("最適化情報", xl.get("violation", pd.DataFrame()))
                # Read 個人業務 sheet from output file (fallback for duty*.xlsx)
                duty_df_from_output = xl.get("個人業務", pd.DataFrame())
                # Read 勤怠 sheet from output file (fallback for kintai files)
                kintai_df_from_output = xl.get("勤怠", pd.DataFrame())
                
                # Normalize Japanese column names back to English for internal processing
                jp_to_en_assign = {
                    '業務ID': 'id', '日付': 'date', '曜日': 'day', '部屋': 'room', '業務': 'dept',
                    '開始時刻': 'start', '終了時刻': 'end', '割当(A)': 'assign_A', '割当(B)': 'assign_B'
                }
                jp_to_en_available = {
                    '日付': 'date', '曜日': 'day', '終日フリー': 'full', 'AMフリー': 'AM', 'PMフリー': 'PM'
                }
                jp_to_en_stats = {
                    '合計回数': 'times', '業務内容': 'id', '個人業務': 'duty_count', '最適化割当': 'opt_count'
                }
                jp_to_en_violation = {
                    '業務ID': 'id', '制約': 'penalty', '点数': 'score', '詳細': 'description'
                }
                
                if not self.assign_df.empty:
                    self.assign_df.rename(columns=jp_to_en_assign, inplace=True)
                if not self.available_A_df.empty:
                    self.available_A_df.rename(columns=jp_to_en_available, inplace=True)
                if not self.available_B_df.empty:
                    self.available_B_df.rename(columns=jp_to_en_available, inplace=True)
                if not self.stats_df.empty:
                    self.stats_df.rename(columns=jp_to_en_stats, inplace=True)
                if not self.violation_df.empty:
                    self.violation_df.rename(columns=jp_to_en_violation, inplace=True)
                
                # Generate フリー(A)/フリー(B) from 割当情報 if those sheets are empty
                if (self.available_A_df.empty or self.available_B_df.empty) and not self.assign_df.empty:
                    logging.info("フリー(A)/フリー(B)シートが空のため、割当情報から生成します")
                    gen_A, gen_B = generate_availability_from_assign(self.assign_df, target_year, target_month)
                    if self.available_A_df.empty:
                        self.available_A_df = gen_A
                    if self.available_B_df.empty:
                        self.available_B_df = gen_B

            work = None
            duty_template = None
            setting = None
            jobA = None
            jobB = None

            self._set_progress(20, "設定データを読み込み中...")
            if self.setting_path:
                setting = Setting(self.setting_path)
                set_reserve_names(setting.get_reserve_names())
                work = WorkData.from_setting(setting, target_year, target_month)
                # generated_shift_A/Bはtemporaryフォルダに出力
                temporary_dir = self.output_path.parent.parent / "temporary" if self.output_path else Path.cwd() / "temporary"
                temporary_dir.mkdir(parents=True, exist_ok=True)
                
                # In visualization mode, JobData is optional (only needed if kintai files exist)
                kintai_A = self.kintai_A_path if hasattr(self, 'kintai_A_path') and self.kintai_A_path and Path(self.kintai_A_path).exists() else None
                kintai_B = self.kintai_B_path if hasattr(self, 'kintai_B_path') and self.kintai_B_path and Path(self.kintai_B_path).exists() else None
                
                # Fallback: if kintai files don't exist but 勤怠 sheet exists in output, use that
                if not kintai_A and not kintai_df_from_output.empty:
                    try:
                        fallback_kintai_path = temporary_dir / "kintaiA_from_output.xlsx"
                        kintai_df_from_output.to_excel(fallback_kintai_path, index=False)
                        kintai_A = fallback_kintai_path
                        logging.info(f"kintaiAファイルがないため、outputの勤怠シートを使用します: {fallback_kintai_path}")
                    except Exception as e:
                        logging.warning(f"勤怠シートからkintaiAファイルの作成に失敗しました: {e}")
                
                if kintai_A:
                    try:
                        jobA = JobData.from_setting(setting, "A", target_year, target_month,
                                     kintai_path=kintai_A,
                                     output_dir=temporary_dir)
                    except ValueError:
                        jobA = None
                        logging.info("kintaiAファイルがないため、JobAは作成されませんでした（可視化モードでは問題ありません）")
                
                if kintai_B:
                    try:
                        jobB = JobData.from_setting(setting, "B", target_year, target_month,
                                     kintai_path=kintai_B,
                                     output_dir=temporary_dir)
                    except ValueError:
                        jobB = None
                        logging.info("kintaiBファイルがないため、JobBは作成されませんでした（可視化モードでは問題ありません）")
            
            # Create DutyTemplate: prefer duty*.xlsx, fallback to 個人業務 sheet from output file
            if self.duty_path and Path(self.duty_path).exists():
                duty_template = DutyTemplate(self.duty_path, target_year, target_month)
                logging.info(f"duty*.xlsxから個人業務を読み込みました: {self.duty_path}")
            elif not duty_df_from_output.empty:
                # Create DutyTemplate from 個人業務 sheet in output file (support both Japanese and English column names)
                duty_template = DutyTemplate(None, target_year, target_month)
                for _, row in duty_df_from_output.iterrows():
                    try:
                        # Support both Japanese and English column names
                        duty = str(row.get('個人業務', row.get('duty', '')))
                        from_day = int(row.get('いつから', row.get('from', 0)))
                        start_time = int(row.get('開始時刻', row.get('start', 0)))
                        end_time = int(row.get('終了時刻', row.get('end', 0)))
                        name = str(row.get('担当', row.get('name', '')))
                        
                        if not duty or not name:
                            continue
                        
                        date_iso = parse_date_token(str(from_day), target_year, target_month)
                        duty_template.assignments.append({
                            'date': date_iso,
                            'start_time': start_time,
                            'end_time': end_time,
                            'duty': duty,
                            'name': name,
                            'day': from_day
                        })
                    except (ValueError, TypeError):
                        continue
                logging.info(f"outputファイルの個人業務シートから{len(duty_template.assignments)}件の個人業務を読み込みました")
            else:
                duty_template = None
                logging.info("個人業務データが見つかりませんでした（duty*.xlsxもoutputの個人業務シートもなし）")

            if not self.assign_df.empty:
                gantt_output_path = self.output_path

                self._set_progress(60, "カレンダービューを生成中...")
                close_assignments = assign_close_duties(jobA, target_year, target_month, self.assign_df, duty_template)
                calendar_df, all_dates, dept_colors = generate_calendar_view(self.assign_df, target_year, target_month, work, jobA, jobB, duty_template, close_assignments)
                work_calendar_df, work_all_dates, work_all_names = generate_work_oriented_calendar_view(self.assign_df, target_year, target_month, work, jobA, jobB, duty_template, close_assignments, self.available_A_df, self.available_B_df)
                
                self._set_progress(80, "ガントチャートを生成中...")
                try:
                    generate_weekly_gantt_images(self.assign_df, gantt_output_path, target_year, target_month, work, duty_template, self.available_A_df, self.available_B_df, close_assignments)
                    
                    with pd.ExcelWriter(gantt_output_path, engine="openpyxl", mode='a', if_sheet_exists='replace') as writer:
                        calendar_df.to_excel(writer, sheet_name="個人カレンダー", index=False, header=False)
                        ws = writer.book["個人カレンダー"]
                        apply_calendar_formatting(ws, all_dates, target_year, target_month, dept_colors)
                        work_calendar_df.to_excel(writer, sheet_name="カレンダー", index=False, header=False)
                        ws2 = writer.book["カレンダー"]
                        apply_work_oriented_calendar_formatting(ws2, work_all_dates, work_all_names, target_year, target_month)
                        
                        # Reorder sheets: 個人カレンダー, カレンダー, 第1週～第6週
                        wb = writer.book
                        desired_order = ["個人カレンダー", "カレンダー"]
                        for i in range(1, 7):
                            sheet_name = f"第{i}週"
                            if sheet_name in wb.sheetnames:
                                desired_order.append(sheet_name)
                        # Add any remaining sheets not in desired_order
                        for sheet_name in wb.sheetnames:
                            if sheet_name not in desired_order:
                                desired_order.append(sheet_name)
                        wb._sheets = [wb[name] for name in desired_order if name in wb.sheetnames]
                    
                except (PermissionError, OSError) as e:
                    gantt_output_path = Path.cwd() / self.output_path.name
                    generate_weekly_gantt_images(self.assign_df, gantt_output_path, target_year, target_month, work, duty_template, self.available_A_df, self.available_B_df, close_assignments)
                    
                    with pd.ExcelWriter(gantt_output_path, engine="openpyxl", mode='a', if_sheet_exists='replace') as writer:
                        calendar_df.to_excel(writer, sheet_name="個人カレンダー", index=False, header=False)
                        ws = writer.book["個人カレンダー"]
                        apply_calendar_formatting(ws, all_dates, target_year, target_month, dept_colors)
                        work_calendar_df.to_excel(writer, sheet_name="カレンダー", index=False, header=False)
                        ws2 = writer.book["カレンダー"]
                        apply_work_oriented_calendar_formatting(ws2, work_all_dates, work_all_names, target_year, target_month)
                        
                        # Reorder sheets: 個人カレンダー, カレンダー, 第1週～第6週
                        wb = writer.book
                        desired_order = ["個人カレンダー", "カレンダー"]
                        for i in range(1, 7):
                            sheet_name = f"第{i}週"
                            if sheet_name in wb.sheetnames:
                                desired_order.append(sheet_name)
                        # Add any remaining sheets not in desired_order
                        for sheet_name in wb.sheetnames:
                            if sheet_name not in desired_order:
                                desired_order.append(sheet_name)
                        wb._sheets = [wb[name] for name in desired_order if name in wb.sheetnames]
                    
                    self.append_log(f"ガントチャートを作業ディレクトリに保存しました: {gantt_output_path}")
                
                self._set_progress(100, f"完了: {gantt_output_path} に出力しました。")
            else:
                self._set_progress(100, "割り当てデータが空のため、ガントチャートは生成されませんでした。")

            # Clean up temporary files after successful output generation
            try:
                if temporary_dir.exists():
                    for temp_file in temporary_dir.glob('*'):
                        if temp_file.is_file():
                            temp_file.unlink()
                            logging.info(f"一時ファイルを削除しました: {temp_file}")
            except Exception as e:
                logging.warning(f"一時ファイルの削除に失敗しました: {e}")

            QMessageBox.information(self, "完了", f"可視化が完了しました。\n\n出力先: {self.output_path}")

        except PermissionError as e:
            msg = "【ファイルアクセスエラー (E401)】\n\n"
            msg += "ファイルへのアクセスが拒否されました。\n\n"
            msg += f"エラー詳細: {str(e)}"
            logging.error(f"PermissionError: {str(e)}")
            self._set_progress(0, "エラー発生")
            QMessageBox.critical(self, "ファイルアクセスエラー", msg)
        
        except Exception as e:
            msg = "【予期しないエラー (E999)】\n\n"
            msg += "予期しないエラーが発生しました。\n\n"
            msg += f"エラー詳細: {str(e)}\n\n"
            logging.error(f"Unexpected error: {str(e)}")
            traceback.print_exc()
            self._set_progress(0, "エラー発生")
            QMessageBox.critical(self, "予期しないエラー", msg)

    def dragEnterEvent(self, event: QDragEnterEvent):
        """Handle drag enter event for file drops"""
        md = event.mimeData()
        if md.hasUrls():
            for url in md.urls():
                if url.isLocalFile():
                    file_path = url.toLocalFile().lower()
                    if file_path.endswith(('.xlsx', '.xls')):
                        event.acceptProposedAction()
                        return
        event.ignore()

    def dropEvent(self, event: QDropEvent):
        """Handle file drop event"""
        urls = event.mimeData().urls()
        if not urls:
            return
        
        path = Path(urls[0].toLocalFile())
        self.handle_dropped_file(path)

    def handle_dropped_file(self, path: Path):
        """Handle a dropped Excel file and determine processing mode based on sheet names
        
        - If '勤務入力表' sheet exists -> optimization mode (extract kintai/duty from kinmu)
        - If 'assign' sheet exists -> visualization mode
        """
        try:
            self.append_log(f"ファイルを読み込み中: {path.name}")
            xl = pd.ExcelFile(path)
            sheet_names = xl.sheet_names
            self.append_log(f"シート: {sheet_names}")
            
            has_kinmu_sheet = any('勤務入力表' in name for name in sheet_names)
            has_assign_sheet = 'assign' in sheet_names or '割当情報' in sheet_names
            
            if has_kinmu_sheet:
                self.append_log("勤務入力表シートを検出 → 最適化モードで処理します")
                self.kinmu_path = path
                
                # 入力ファイルから期間を自動検出、計算期間と異なる場合は入力ファイルの期間を使用
                # ※ validation内で_compute_target_periodを使うため、先に検出・設定する
                detected_period = self._detect_period_from_input(xl)
                computed_period = self._compute_target_period()
                if detected_period and detected_period != computed_period:
                    self._override_period = detected_period
                    self.append_log(
                        f"入力ファイルの期間（{detected_period[0]}年{detected_period[1]}月）を使用します "
                        f"（計算上の期間: {computed_period[0]}年{computed_period[1]}月）"
                    )
                else:
                    self._override_period = None

                if not self._validate_input_structure(path, mode="opt", excel_file=xl):
                    self._override_period = None
                    return
                
                self._prepare_kintai_and_duty_from_kinmu(path, xl)
                self.run_assign()
                
            elif has_assign_sheet:
                self.append_log("割当情報シートを検出 → 可視化モードで処理します")
                self.visualization_input_path = path
                self._override_period = None
                
                if not self._validate_input_structure(path, mode="vis", excel_file=xl):
                    return
                
                self.run_visualize()
            else:
                QMessageBox.warning(
                    self, 
                    "ファイル形式エラー",
                    "ドロップされたファイルには「勤務入力表」シートも「割当情報」シートも含まれていません。\n\n"
                    "最適化を行う場合は「勤務入力表」シートを含むファイルを、\n"
                    "可視化を行う場合は「割当情報」シートを含むファイルをドロップしてください。"
                )
                
        except Exception as e:
            logging.error(f"ファイル読み込みエラー: {str(e)}")
            QMessageBox.critical(
                self,
                "読み込みエラー",
                f"Excelファイルを読み込めませんでした:\n{str(e)}"
            )

    def _detect_period_from_input(self, excel_file: pd.ExcelFile) -> Optional[Tuple[int, int]]:
        """入力ファイルの日付から対象期間を自動検出する

        Returns:
            (year, month) tuple or None if detection fails
        """
        try:
            kinmu_sheet_name = None
            for name in excel_file.sheet_names:
                if '勤務入力表' in name:
                    kinmu_sheet_name = name
                    break
            if kinmu_sheet_name is None:
                return None

            df = pd.read_excel(excel_file, sheet_name=kinmu_sheet_name, header=None)
            date_row = df.iloc[3]

            for col_idx in range(2, len(date_row), 2):
                date_val = date_row.iloc[col_idx]
                if isinstance(date_val, dt.datetime):
                    d = date_val.date()
                    if d.day == 21:
                        return (d.year, d.month)
                    break
        except Exception as e:
            logging.warning(f"入力ファイルからの期間検出失敗: {e}")
        return None

    def _validate_input_structure(self, path: Path, mode: str, excel_file: pd.ExcelFile) -> bool:
        """Validate input file structure before processing
        
        Args:
            path: Path to the input file
            mode: "opt" for optimization, "vis" for visualization
            excel_file: Already opened ExcelFile object
            
        Returns:
            True if validation passes, False otherwise
        """
        errors = []
        
        if not self.setting_path:
            errors.append("settingファイルが見つかりません。filesフォルダにsetting*.xlsxを配置してください。")
            QMessageBox.critical(self, "設定エラー", "\n".join(errors))
            return False
        
        try:
            setting = Setting(self.setting_path)
            set_reserve_names(setting.get_reserve_names())
        except Exception as e:
            errors.append(f"settingファイルの読み込みに失敗しました: {str(e)}")
            QMessageBox.critical(self, "設定エラー", "\n".join(errors))
            return False
        
        if mode == "opt":
            kinmu_sheet_name = None
            for name in excel_file.sheet_names:
                if '勤務入力表' in name:
                    kinmu_sheet_name = name
                    break
            
            if kinmu_sheet_name is None:
                errors.append("勤務入力表シートが見つかりません。")
                QMessageBox.critical(self, "構造エラー", "\n".join(errors))
                return False
            
            df = pd.read_excel(excel_file, sheet_name=kinmu_sheet_name, header=None)
            
            defined_members = set()
            if hasattr(setting, 'jobA_df') and setting.jobA_df is not None:
                if 'name' in setting.jobA_df.columns:
                    for name in setting.jobA_df['name']:
                        name_str = str(name).strip()
                        name_short = name_str.split()[0] if ' ' in name_str else name_str.split('　')[0] if '　' in name_str else name_str
                        defined_members.add(name_str)
                        defined_members.add(name_short)
            if hasattr(setting, 'jobB_df') and setting.jobB_df is not None:
                if 'name' in setting.jobB_df.columns:
                    for name in setting.jobB_df['name']:
                        name_str = str(name).strip()
                        name_short = name_str.split()[0] if ' ' in name_str else name_str.split('　')[0] if '　' in name_str else name_str
                        defined_members.add(name_str)
                        defined_members.add(name_short)
            
            # dictAのaka2列による別名も追加
            dictA = getattr(setting, 'dictA_df', None)
            if dictA is not None and 'aka2' in dictA.columns:
                for aka2 in dictA['aka2']:
                    if pd.notna(aka2):
                        name_str = str(aka2).strip()
                        name_short = name_str.split()[0] if ' ' in name_str else name_str.split('　')[0] if '　' in name_str else name_str
                        defined_members.add(name_str)
                        defined_members.add(name_short)
            
            # dictBのaka2列による別名も追加
            dictB = getattr(setting, 'dictB_df', None)
            if dictB is not None and 'aka2' in dictB.columns:
                for aka2 in dictB['aka2']:
                    if pd.notna(aka2):
                        name_str = str(aka2).strip()
                        name_short = name_str.split()[0] if ' ' in name_str else name_str.split('　')[0] if '　' in name_str else name_str
                        defined_members.add(name_str)
                        defined_members.add(name_short)
            
            unknown_members = []
            staff_start_row = 6
            consecutive_empty = 0
            for i in range(staff_start_row, min(len(df), staff_start_row + 100)):
                name_cell = df.iloc[i, 1]
                if pd.isna(name_cell) or str(name_cell).strip() == '':
                    consecutive_empty += 1
                    if consecutive_empty >= 2:
                        break
                    continue
                consecutive_empty = 0
                name_full = str(name_cell).strip()
                name_short = name_full.split()[0] if ' ' in name_full else name_full.split('　')[0] if '　' in name_full else name_full
                
                if name_full not in defined_members and name_short not in defined_members:
                    unknown_members.append(name_short)
            
            if unknown_members:
                unique_unknown = list(dict.fromkeys(unknown_members))
                names = '、'.join(unique_unknown[:10])
                if len(unique_unknown) > 10:
                    names += f"（他{len(unique_unknown) - 10}名）"
                QMessageBox.warning(
                    self, "未登録スタッフ",
                    f"{names}さんの勤務条件が設定されていません。\n"
                    "setting.xlsxのjobA/jobB・dictA/dictB・orderA/orderBシートを\n"
                    "更新してアプリを再起動してください。"
                )
                return
            
            date_row = df.iloc[3]
            target_year, target_month = self._compute_target_period()
            
            period_start = dt.date(target_year, target_month, 21)
            if target_month == 12:
                period_end = dt.date(target_year + 1, 1, 20)
            else:
                period_end = dt.date(target_year, target_month + 1, 20)
            
            file_dates = []
            for col_idx in range(2, len(date_row), 2):
                date_val = date_row.iloc[col_idx]
                if isinstance(date_val, dt.datetime):
                    file_dates.append(date_val.date())
            
            if file_dates:
                first_date = file_dates[0]
                last_date = file_dates[-1]
                
                if first_date.day != 21:
                    errors.append(f"入力ファイルの開始日が21日ではありません（{first_date}）。入力ファイルを確認してください。")
                
                if last_date.day != 20:
                    errors.append(f"入力ファイルの終了日が20日ではありません（{last_date}）。入力ファイルを確認してください。")
                
                if first_date != period_start:
                    errors.append(f"入力ファイルの期間（{first_date}〜{last_date}）が対象期間（{period_start}〜{period_end}）と一致しません。")
        
        if errors:
            error_msg = "入力ファイルに問題が見つかりました:\n\n"
            for i, err in enumerate(errors, 1):
                error_msg += f"{i}. {err}\n"
            error_msg += "\n入力ファイルを修正してから再度ドロップしてください。"
            
            QMessageBox.warning(self, "構造チェックエラー", error_msg)
            self.append_log(f"構造チェックエラー: {len(errors)}件の問題が見つかりました")
            for err in errors:
                self.append_log(f"  - {err}")
            return False
        
        self.append_log("構造チェック: OK")
        return True

    def _prepare_kintai_and_duty_from_kinmu(self, kinmu_path: Path, excel_file: pd.ExcelFile):
        """Extract kintai and duty data from kinmu file and prepare for optimization
        
        Args:
            kinmu_path: Path to the kinmu input file
            excel_file: Already opened ExcelFile object
        """
        try:
            setting = Setting(self.setting_path)
            set_reserve_names(setting.get_reserve_names())
            target_year, target_month = self._compute_target_period()
            
            members_df_list = []
            if hasattr(setting, 'jobA_df') and setting.jobA_df is not None:
                members_df_list.append(setting.jobA_df)
            if hasattr(setting, 'jobB_df') and setting.jobB_df is not None:
                members_df_list.append(setting.jobB_df)
            
            if members_df_list:
                members_df = pd.concat(members_df_list, ignore_index=True)
            else:
                members_df = pd.DataFrame(columns=['name', 'aka'])
            
            # dictAのaka2列による別名マッピングを追加
            # aka2には勤務入力表で使われている人名が格納されている
            dictA = getattr(setting, 'dictA_df', None)
            if dictA is not None and 'aka' in dictA.columns and 'aka2' in dictA.columns:
                synonym_rows = []
                for _, r in dictA.iterrows():
                    aka = str(r['aka']).strip() if pd.notna(r['aka']) else ''
                    aka2 = r['aka2']
                    if pd.isna(aka2) or not aka:
                        continue
                    name_input = str(aka2).strip()
                    if not name_input:
                        continue
                    synonym_rows.append({'name': name_input, 'aka': aka})
                
                if synonym_rows:
                    synonyms_df = pd.DataFrame(synonym_rows)
                    # members_dfに存在する他の列を埋めておく
                    for col in members_df.columns:
                        if col not in synonyms_df.columns:
                            synonyms_df[col] = pd.NA
                    members_df = pd.concat([members_df, synonyms_df], ignore_index=True)
            
            # dictBのaka2列による別名マッピングも追加（存在する場合）
            dictB = getattr(setting, 'dictB_df', None)
            if dictB is not None and 'aka' in dictB.columns and 'aka2' in dictB.columns:
                synonym_rows = []
                for _, r in dictB.iterrows():
                    aka = str(r['aka']).strip() if pd.notna(r['aka']) else ''
                    aka2 = r['aka2']
                    if pd.isna(aka2) or not aka:
                        continue
                    name_input = str(aka2).strip()
                    if not name_input:
                        continue
                    synonym_rows.append({'name': name_input, 'aka': aka})
                
                if synonym_rows:
                    synonyms_df = pd.DataFrame(synonym_rows)
                    for col in members_df.columns:
                        if col not in synonyms_df.columns:
                            synonyms_df[col] = pd.NA
                    members_df = pd.concat([members_df, synonyms_df], ignore_index=True)
            
            kintai_df, duty_df, end_time_constraints, no_close_duty_constraints = extract_kintai_and_duty_from_kinmu(
                kinmu_path,
                self.duty_path,
                members_df,
                target_year,
                target_month
            )
            
            app_dir = get_app_dir()
            temp_dir = app_dir / "temporary"
            temp_dir.mkdir(exist_ok=True)
            
            kintai_output_path = temp_dir / "kintaiA.xlsx"
            try:
                kintai_df.to_excel(kintai_output_path, index=False)
                self.kintai_A_path = kintai_output_path
                self.append_log(f"kintaiデータを生成: {kintai_output_path}")
            except PermissionError:
                QMessageBox.warning(
                    self,
                    "ファイルアクセスエラー",
                    f"kintaiファイルへのアクセスが拒否されました。\n"
                    f"ファイルが開かれている場合は閉じてください:\n{kintai_output_path}"
                )
                return
            
            if not duty_df.empty:
                duty_output_path = temp_dir / "duty.xlsx"
                try:
                    duty_df.to_excel(duty_output_path, sheet_name='fix', index=False)
                    self.duty_path = duty_output_path
                    self.append_log(f"dutyデータを生成: {duty_output_path}")
                except PermissionError:
                    QMessageBox.warning(
                        self,
                        "ファイルアクセスエラー",
                        f"dutyファイルへのアクセスが拒否されました。\n"
                        f"ファイルが開かれている場合は閉じてください:\n{duty_output_path}"
                    )
                    return
            
            self._end_time_constraints = end_time_constraints
            self._no_close_duty_constraints = no_close_duty_constraints
            
            self.append_log(f"勤務入力表からの抽出完了: kintai={len(kintai_df)}行, duty={len(duty_df)}行")
            
        except Exception as e:
            logging.error(f"勤務入力表からの抽出エラー: {str(e)}")
            traceback.print_exc()
            QMessageBox.critical(
                self,
                "抽出エラー",
                f"勤務入力表からのデータ抽出に失敗しました:\n{str(e)}"
            )

    def closeEvent(self, event):
        super().closeEvent(event)


def _global_excepthook(exc_type, exc_value, exc_traceback):
    err = "".join(traceback.format_exception(exc_type, exc_value, exc_traceback))
    try:
        QMessageBox.critical(None, "未処理例外", err)
    except Exception:
        pass
    logging.error(err)
    print(err, file=sys.stderr)
    sys.exit(1)

def ensure_xlrd_installed():
    """xlrdがインストールされていない場合はインストールする（.xls形式のサポート用）"""
    try:
        import xlrd
    except ImportError:
        import subprocess
        logging.info("xlrdがインストールされていません。インストールを開始します...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "xlrd"])
            logging.info("xlrdのインストールが完了しました。")
        except subprocess.CalledProcessError as e:
            logging.error(f"xlrdのインストールに失敗しました: {e}")
            raise RuntimeError("xlrdのインストールに失敗しました。手動で 'pip install xlrd' を実行してください。")


from license_manager import LicenseManager  # noqa: E402


class LicenseDialog(QDialog):
    """License authentication dialog (fallback when auto-login unavailable)."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("ライセンス認証")
        self.setFixedSize(480, 280)
        self.authenticated = False

        layout = QVBoxLayout(self)

        # Title
        title_label = QLabel("Shift Optimizer - ライセンス認証")
        title_label.setStyleSheet(
            "font-size: 14px; font-weight: bold; margin-bottom: 10px;"
        )
        layout.addWidget(title_label)

        # User ID
        id_layout = QHBoxLayout()
        id_label = QLabel("ユーザーID:")
        id_label.setFixedWidth(100)
        self.id_input = QLineEdit()
        id_layout.addWidget(id_label)
        id_layout.addWidget(self.id_input)
        layout.addLayout(id_layout)

        # Password
        pw_layout = QHBoxLayout()
        pw_label = QLabel("パスワード:")
        pw_label.setFixedWidth(100)
        self.pw_input = QLineEdit()
        self.pw_input.setEchoMode(QLineEdit.EchoMode.Password)
        pw_layout.addWidget(pw_label)
        pw_layout.addWidget(self.pw_input)
        layout.addLayout(pw_layout)

        # Status message (word-wrap for multi-line errors)
        self.status_label = QLabel("")
        self.status_label.setStyleSheet("color: red;")
        self.status_label.setWordWrap(True)
        layout.addWidget(self.status_label)

        # Buttons
        btn_layout = QHBoxLayout()
        self.login_btn = QPushButton("認証")
        self.login_btn.clicked.connect(self.authenticate)
        self.cancel_btn = QPushButton("キャンセル")
        self.cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(self.login_btn)
        btn_layout.addWidget(self.cancel_btn)
        layout.addLayout(btn_layout)

        # Enter key triggers login
        self.pw_input.returnPressed.connect(self.authenticate)

    def authenticate(self):
        """Validate license credentials"""
        user_id = self.id_input.text().strip()
        password = self.pw_input.text()

        if not user_id or not password:
            self.status_label.setText(
                "ユーザーIDとパスワードを入力してください。"
            )
            return

        manager = LicenseManager()
        is_valid, message = manager.validate_license(user_id, password)

        if is_valid:
            self.authenticated = True
            self.status_label.setStyleSheet("color: green;")
            self.status_label.setText(message)
            QTimer.singleShot(1000, self.accept)
        else:
            self.status_label.setStyleSheet("color: red;")
            self.status_label.setText(message)


def main():
    ensure_xlrd_installed()
    sys.excepthook = _global_excepthook
    app = QApplication(sys.argv)
    
    # Ensure standard subdirectories exist next to the executable
    app_dir = get_app_dir()
    for subdir in ("files", "input", "output"):
        (app_dir / subdir).mkdir(parents=True, exist_ok=True)
    
    default_font_name = get_system_japanese_font()
    if default_font_name:
        app.setFont(QFont(default_font_name, 10))
    
    # License authentication — auto-login if valid .license exists in files/
    manager = LicenseManager()
    auto_ok, auto_msg = manager.validate_license_auto()
    if auto_ok:
        logging.info(auto_msg)
    else:
        license_dialog = LicenseDialog()
        if license_dialog.exec() != QDialog.DialogCode.Accepted or not license_dialog.authenticated:
            sys.exit(0)
    
    w = MainWindow()
    
    screen = app.primaryScreen()
    screen_height = int(screen.availableGeometry().height() * 0.8)
    window_width = int(screen_height * 8 / 9)
    w.window_width = window_width
    w.window_height = screen_height
    w._update_font_sizes()
    w.resize(window_width, screen_height)
    w.show()
    try:
        sys.exit(app.exec())
    except Exception:
        _global_excepthook(*sys.exc_info())


if __name__ == "__main__":
    main()
